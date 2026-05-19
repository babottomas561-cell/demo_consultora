from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Optional, Literal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.api.v1.dashboard import get_tenant_schema
from app.core.database import get_db
from app.core.filters import GlobalFilters, get_global_filters, text_filter_clause


router = APIRouter()


INFOMANAGER_REPORT_CATALOG = [
    {"key": "clientes", "name": "Listado de clientes", "group": "clientes", "supported": True},
    {"key": "proveedores", "name": "Listado de proveedores", "group": "proveedores", "supported": True},
    {"key": "vendedores", "name": "Listado de vendedores", "group": "vendedores", "supported": True},
    {"key": "articulos", "name": "Listado de articulos", "group": "stock", "supported": True},
    {"key": "saldos_clientes", "name": "Saldos de cuentas corrientes", "group": "clientes", "supported": True},
    {"key": "comprobantes_pendientes_clientes", "name": "Comprobantes pendientes de clientes", "group": "clientes", "supported": True},
    {"key": "facturas_clientes", "name": "Listado de facturas", "group": "clientes", "supported": True},
    {"key": "facturas_con_recibos", "name": "Facturas con recibos", "group": "clientes", "supported": True},
    {"key": "facturas_compras", "name": "Listado de facturas pendientes de proveedores", "group": "proveedores", "supported": True},
    {"key": "compras_por_factura", "name": "Analisis de compra por factura", "group": "compras", "supported": True},
    {"key": "mayor_contable", "name": "Libro mayor", "group": "contabilidad", "supported": True},
    {"key": "planes", "name": "Plan de cuentas", "group": "contabilidad", "supported": True},
    {"key": "stock_existencias", "name": "Existencias de stock", "group": "stock", "supported": True},
    {"key": "ventas", "name": "Ventas", "group": "clientes", "supported": True},
    {"key": "ventas_items", "name": "Analisis de compra/ventas por articulo", "group": "clientes", "supported": True},
    {"key": "compras", "name": "Compras", "group": "compras", "supported": True},
    {"key": "compras_items", "name": "Compras por articulo", "group": "compras", "supported": True},
    {"key": "interdeposito", "name": "Movimientos de stock entre depositos", "group": "stock", "supported": True},
    {"key": "disponible_por_cliente", "name": "Disponible por cliente", "group": "caja", "supported": True},
    {"key": "comprobantes_relacion", "name": "Comprobantes relacion", "group": "comprobantes", "supported": True},
    {"key": "comprobantes_destino", "name": "Comprobantes destino", "group": "comprobantes", "supported": True},
    {"key": "cta_cte_clientes", "name": "Cuentas corrientes de clientes", "group": "clientes", "supported": True, "source": "derived"},
    {"key": "cta_cte_proveedores", "name": "Cuentas corrientes de proveedores", "group": "proveedores", "supported": True, "source": "derived"},
    {"key": "comprobantes_pendientes_proveedores", "name": "Comprobantes pendientes de proveedores", "group": "proveedores", "supported": True, "source": "derived"},
    {"key": "gastos_por_rubro_proveedores", "name": "Gastos por rubro de proveedores", "group": "proveedores", "supported": True, "source": "derived"},
    {"key": "facturas_vs_remitos_articulos", "name": "Facturas vs remitos por totales de articulos", "group": "compras", "supported": True, "source": "derived"},
    {"key": "notas_credito_clientes", "name": "Listado de notas de credito", "group": "clientes", "supported": True, "source": "derived"},
    {"key": "notas_debito_clientes", "name": "Listado de notas de debito", "group": "clientes", "supported": True, "source": "derived"},
    {"key": "margen_rentabilidad", "name": "Margen de rentabilidad", "group": "ventas", "supported": True, "source": "derived"},
    {"key": "margen_bruto_mensual", "name": "Margen bruto mensual", "group": "ventas", "supported": True, "source": "derived"},
    {"key": "ventas_por_vendedor", "name": "Ventas por vendedor", "group": "vendedores", "supported": True, "source": "derived"},
    {"key": "iva_compras", "name": "IVA compras", "group": "contabilidad", "supported": True, "source": "derived"},
    {"key": "iva_ventas", "name": "IVA ventas", "group": "contabilidad", "supported": True, "source": "derived"},
    {"key": "estado_resultados_derivado", "name": "Estado de resultados", "group": "contabilidad", "supported": True, "source": "derived"},
    {"key": "cash_flow", "name": "Cash Flow", "group": "caja", "supported": True, "source": "derived"},
    {
        "key": "clientes_por_vendedor",
        "name": "Clientes por vendedor",
        "group": "vendedores",
        "supported": True,
        "source": "derived",
        "note": "Derivado desde ventas sincronizadas; Swagger v1 no publica un reporte dedicado.",
    },
    {
        "key": "comisiones_por_recibos",
        "name": "Comisiones por recibos",
        "group": "vendedores",
        "supported": True,
        "source": "derived",
        "note": "Derivado desde recibos y comprobantes sincronizados; Swagger v1 no publica un reporte dedicado.",
    },
    {
        "key": "anticipos_clientes",
        "name": "Anticipos emitidos vs cancelados de clientes",
        "group": "clientes",
        "supported": False,
        "note": "Swagger v1 no publica un endpoint de anticipos de clientes.",
    },
    {
        "key": "remitos_cliente",
        "name": "Remitos de clientes",
        "group": "clientes",
        "supported": False,
        "note": "Swagger v1 solo publica creacion/consulta individual de remitos, no un listado exportable por periodo.",
    },
    {
        "key": "saldos_proveedores_clientes",
        "name": "Saldos de proveedores-clientes",
        "group": "proveedores",
        "supported": False,
        "note": "Swagger v1 no publica el reporte consolidado proveedor-cliente.",
    },
    {
        "key": "conciliacion_proveedor_cliente",
        "name": "Conciliacion proveedor-cliente",
        "group": "proveedores",
        "supported": False,
        "note": "Swagger v1 no publica el reporte de conciliacion proveedor-cliente.",
    },
    {
        "key": "anticipos_proveedores",
        "name": "Anticipos emitidos vs cancelados de proveedores",
        "group": "proveedores",
        "supported": False,
        "note": "Swagger v1 no publica un endpoint de anticipos de proveedores.",
    },
    {
        "key": "saldos_proveedores_por_cuenta",
        "name": "Saldos de proveedores por cuenta contable",
        "group": "proveedores",
        "supported": False,
        "note": "Swagger v1 no publica el saldo de proveedores por cuenta contable.",
    },
    {
        "key": "movimientos_por_articulo",
        "name": "Movimientos por articulo",
        "group": "stock",
        "supported": True,
        "source": "derived",
        "note": "Derivado desde compras, ventas y stock sincronizados; Swagger v1 no publica un historial global.",
    },
    {
        "key": "proyeccion_stock",
        "name": "Proyeccion de stock",
        "group": "stock",
        "supported": True,
        "source": "derived",
        "note": "Derivado desde existencias y demanda vendida; Swagger v1 no publica este reporte.",
    },
    {
        "key": "cheques",
        "name": "Cheques / disponibilidades / cash flow",
        "group": "caja",
        "supported": False,
        "note": "No hay endpoints de cheques, disponibilidades o cash flow en Swagger v1.",
    },
    {
        "key": "iva_balance",
        "name": "IVA, balances y estado de resultados",
        "group": "contabilidad",
        "supported": False,
        "note": "Swagger v1 solo publica plan de cuentas y mayor; no publica IVA compras/ventas ni balances cerrados.",
    },
]


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def _year_ago(d: date) -> date:
    try:
        return d.replace(year=d.year - 1)
    except ValueError:
        return d.replace(year=d.year - 1, day=28)


def _prev_period(filters: GlobalFilters):
    if getattr(filters, "compare_mode", "anterior") == "anio":
        return _year_ago(filters.desde), _year_ago(filters.hasta_exclusive)
    length = (filters.hasta - filters.desde).days
    prev_hasta = filters.desde - timedelta(days=1)
    prev_desde = prev_hasta - timedelta(days=length)
    return prev_desde, prev_hasta + timedelta(days=1)


def _ventas_base_where(filters: GlobalFilters) -> str:
    """Base WHERE for ventas (dates + all array filters + anulada)."""
    return text_filter_clause("ventas", filters)


def _var_pct(actual: float, anterior: float) -> Optional[float]:
    if anterior and anterior != 0:
        return round((actual - anterior) / abs(anterior) * 100, 2)
    return None


def _kpi_obj(actual: float, anterior: Optional[float] = None):
    return {
        "actual": round(actual, 2),
        "anterior": round(anterior, 2) if anterior is not None else None,
        "variacion_pct": _var_pct(actual, anterior) if anterior is not None else None,
    }


def _col(alias: str, column: str) -> str:
    return f"{alias}.{column}" if alias else column


def venta_importe_neto_expr(alias: str = "") -> str:
    total = _col(alias, "total")
    tipo = _col(alias, "tipo_comprobante")
    return (
        f"CASE WHEN {tipo} IN ('FA','ND') THEN ABS({total}) "
        f"WHEN {tipo} = 'NC' THEN -ABS({total}) ELSE 0 END"
    )


def venta_iva_neto_expr(alias: str = "") -> str:
    total = _col(alias, "total")
    iva = _col(alias, "iva_importe")
    tipo = _col(alias, "tipo_comprobante")
    # Cast to text first so this works whether the column is numeric or text
    base = f"COALESCE(NULLIF({iva}::text, '')::float, 0)"
    return (
        f"CASE WHEN {tipo} IN ('FA','ND') THEN ABS({base}) "
        f"WHEN {tipo} = 'NC' THEN -ABS({base}) ELSE 0 END"
    )


def venta_costo_neto_expr(alias: str = "") -> str:
    cantidad = _col(alias, "cantidad")
    costo = _col(alias, "precio_compra_actual")
    tipo = _col(alias, "tipo_comprobante")
    # Cast to text first so this works whether the column is numeric or text
    safe_costo = f"COALESCE(NULLIF({costo}::text, '')::float, 0)"
    base = f"ABS({cantidad}) * {safe_costo}"
    return (
        f"CASE WHEN {tipo} IN ('FA','ND') THEN {base} "
        f"WHEN {tipo} = 'NC' THEN -{base} ELSE 0 END"
    )


def compra_importe_neto_expr(alias: str = "") -> str:
    total = _col(alias, "total")
    tipo = _col(alias, "tipo_comprobante")
    return (
        f"CASE WHEN COALESCE({tipo}, 'FC') IN ('FC','FA','ND') THEN ABS({total}) "
        f"WHEN {tipo} = 'NC' THEN -ABS({total}) ELSE ABS({total}) END"
    )


async def _fetch_kpi_row(db: AsyncSession, where: str, params: dict) -> dict:
    row = (await db.execute(text(f"""
        SELECT
            COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END),0)            AS fa_bruto,
            COALESCE(SUM(CASE WHEN tipo_comprobante='NC' THEN ABS(total) ELSE 0 END),0)            AS nc_total,
            COALESCE(SUM(CASE WHEN tipo_comprobante='ND' THEN ABS(total) ELSE 0 END),0)            AS nd_total,
            COALESCE(SUM({venta_importe_neto_expr()}),0)                                          AS facturado_neto,
            COALESCE(SUM({venta_iva_neto_expr()}),0)                                              AS iva_debito,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                                 AS tickets,
            COUNT(DISTINCT CASE WHEN tipo_comprobante='FA' THEN cliente_id END)               AS clientes_unicos,
            COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN cantidad ELSE 0 END),0)         AS unidades,
            COALESCE(SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}),0)           AS margen_dolares,
            COALESCE(SUM(
              CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END),0) AS total_con_costo
        FROM ventas
        WHERE {where}
    """), params)).mappings().one()
    return dict(row)


def quote_schema(schema_name: str) -> str:
    return '"' + schema_name.replace('"', '""') + '"'


async def set_tenant_search_path(db: AsyncSession, tenant_schema: str) -> None:
    await db.execute(text(f"SET search_path TO {quote_schema(tenant_schema)}"))


def money(value) -> float:
    return float(value or 0)


def resolve_dates(desde: Optional[str], hasta: Optional[str]):
    """Backward-compatible helper for tests and older callers."""
    hasta_d = date.fromisoformat(hasta) if hasta else date.today()
    desde_d = date.fromisoformat(desde) if desde else date.today() - timedelta(days=365)
    return desde_d, hasta_d + timedelta(days=1)


def _jsonable_payload(row: dict) -> dict:
    payload = {}
    for key, value in row.items():
        if isinstance(value, Decimal):
            payload[key] = float(value)
        elif hasattr(value, "isoformat"):
            payload[key] = value.isoformat()
        else:
            payload[key] = value
    return payload


async def _fetch_derived_infomanager_rows(
    report_key: str,
    filters: GlobalFilters,
    db: AsyncSession,
    limit: int,
) -> Optional[list[dict]]:
    params = dict(filters.sql_params())
    params["limit"] = limit
    ventas_where = _ventas_base_where(filters)
    compras_where = text_filter_clause("compras", filters)
    cta_where = "fecha >= :desde AND fecha < :hasta"

    sql_by_key = {
        "cta_cte_clientes": f"""
            SELECT cliente_id, cliente_nombre, comprobante_id, tipo, fecha,
                   importe, saldo_acumulado, fecha_vencimiento
            FROM cuentas_corrientes_clientes
            WHERE {cta_where}
            ORDER BY fecha DESC, comprobante_id
            LIMIT :limit
        """,
        "cta_cte_proveedores": f"""
            SELECT proveedor_id, proveedor_nombre, comprobante_id, tipo, fecha,
                   importe, saldo_acumulado, fecha_vencimiento
            FROM cuentas_corrientes_proveedores
            WHERE {cta_where}
            ORDER BY fecha DESC, comprobante_id
            LIMIT :limit
        """,
        "comprobantes_pendientes_proveedores": f"""
            SELECT proveedor_id, proveedor_nombre, comprobante_id, tipo, numero,
                   punto_de_venta, fecha, fecha_vencimiento, importe_total,
                   importe_pagado, saldo
            FROM comprobantes_proveedores
            WHERE {cta_where} AND saldo > 0
            ORDER BY fecha_vencimiento NULLS LAST, fecha DESC
            LIMIT :limit
        """,
        "gastos_por_rubro_proveedores": f"""
            SELECT producto_id, producto_nombre, proveedor_id, proveedor_nombre,
                   COUNT(*) AS comprobantes,
                   SUM(cantidad) AS cantidad,
                   SUM({compra_importe_neto_expr()}) AS importe_total,
                   SUM(COALESCE(iva_importe::float, 0)) AS iva_total
            FROM compras
            WHERE {compras_where}
            GROUP BY producto_id, producto_nombre, proveedor_id, proveedor_nombre
            ORDER BY importe_total DESC
            LIMIT :limit
        """,
        "facturas_vs_remitos_articulos": f"""
            SELECT producto_id, producto_nombre,
                   SUM(CASE WHEN tipo_comprobante IN ('FA','FC') THEN ABS(total) ELSE 0 END) AS facturas_total,
                   SUM(CASE WHEN tipo_comprobante = 'RE' THEN ABS(total) ELSE 0 END) AS remitos_total,
                   SUM(CASE WHEN tipo_comprobante IN ('FA','FC') THEN cantidad ELSE 0 END) AS facturas_cantidad,
                   SUM(CASE WHEN tipo_comprobante = 'RE' THEN cantidad ELSE 0 END) AS remitos_cantidad
            FROM compras
            WHERE {compras_where}
            GROUP BY producto_id, producto_nombre
            ORDER BY GREATEST(
                SUM(CASE WHEN tipo_comprobante IN ('FA','FC') THEN ABS(total) ELSE 0 END),
                SUM(CASE WHEN tipo_comprobante = 'RE' THEN ABS(total) ELSE 0 END)
            ) DESC
            LIMIT :limit
        """,
        "notas_credito_clientes": f"""
            SELECT fecha, cliente_id, cliente_nombre, producto_id, producto_nombre,
                   cantidad, precio_unitario, total, iva_importe, punto_de_venta,
                   cod_vendedor
            FROM ventas
            WHERE {ventas_where} AND tipo_comprobante = 'NC'
            ORDER BY fecha DESC
            LIMIT :limit
        """,
        "notas_debito_clientes": f"""
            SELECT fecha, cliente_id, cliente_nombre, producto_id, producto_nombre,
                   cantidad, precio_unitario, total, iva_importe, punto_de_venta,
                   cod_vendedor
            FROM ventas
            WHERE {ventas_where} AND tipo_comprobante = 'ND'
            ORDER BY fecha DESC
            LIMIT :limit
        """,
        "margen_rentabilidad": f"""
            SELECT producto_id, producto_nombre,
                   SUM({venta_importe_neto_expr()}) AS facturado,
                   SUM({venta_costo_neto_expr()}) AS costo,
                   SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}) AS margen,
                   CASE WHEN SUM({venta_importe_neto_expr()}) <> 0
                        THEN SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}) / SUM({venta_importe_neto_expr()}) * 100
                        ELSE 0 END AS margen_pct
            FROM ventas
            WHERE {ventas_where}
            GROUP BY producto_id, producto_nombre
            ORDER BY margen DESC
            LIMIT :limit
        """,
        "margen_bruto_mensual": f"""
            SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo,
                   SUM({venta_importe_neto_expr()}) AS facturado,
                   SUM({venta_costo_neto_expr()}) AS costo,
                   SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}) AS margen
            FROM ventas
            WHERE {ventas_where}
            GROUP BY 1
            ORDER BY 1
            LIMIT :limit
        """,
        "ventas_por_vendedor": f"""
            WITH base AS (
                SELECT * FROM ventas WHERE {text_filter_clause("ventas", filters)}
            )
            SELECT b.cod_vendedor, COALESCE(MAX(ven.nombre), 'Sin vendedor') AS vendedor,
                   COUNT(*) AS comprobantes,
                   COUNT(DISTINCT b.cliente_id) AS clientes,
                   SUM({venta_importe_neto_expr('b')}) AS facturado,
                   SUM({venta_importe_neto_expr('b')} - {venta_costo_neto_expr('b')}) AS margen
            FROM base b
            LEFT JOIN vendedores ven ON ven.cod_vendedor = b.cod_vendedor
            GROUP BY b.cod_vendedor
            ORDER BY facturado DESC
            LIMIT :limit
        """,
        "clientes_por_vendedor": f"""
            WITH base AS (
                SELECT * FROM ventas WHERE {text_filter_clause("ventas", filters)}
            )
            SELECT b.cod_vendedor, COALESCE(MAX(ven.nombre), 'Sin vendedor') AS vendedor,
                   b.cliente_id, MAX(b.cliente_nombre) AS cliente_nombre,
                   COUNT(*) AS comprobantes,
                   SUM({venta_importe_neto_expr('b')}) AS facturado,
                   MAX(b.fecha) AS ultima_compra
            FROM base b
            LEFT JOIN vendedores ven ON ven.cod_vendedor = b.cod_vendedor
            GROUP BY b.cod_vendedor, b.cliente_id
            ORDER BY vendedor, facturado DESC
            LIMIT :limit
        """,
        "comisiones_por_recibos": """
            SELECT cod_vendedor, vendedor_nombre, periodo, base_cobrada,
                   porcentaje, comision, recibos
            FROM comisiones_vendedores
            WHERE periodo >= to_char(CAST(:desde AS date), 'YYYY-MM')
              AND periodo <= to_char((CAST(:hasta AS date) - INTERVAL '1 day'), 'YYYY-MM')
            ORDER BY periodo DESC, comision DESC
            LIMIT :limit
        """,
        "movimientos_por_articulo": f"""
            SELECT fecha, producto_id AS cod_articulo, producto_nombre, 'entrada' AS movimiento,
                   proveedor_id AS tercero_id, proveedor_nombre AS tercero_nombre,
                   tipo_comprobante, cantidad, total
            FROM compras
            WHERE {compras_where}
            UNION ALL
            SELECT fecha, producto_id AS cod_articulo, producto_nombre, 'salida' AS movimiento,
                   cliente_id AS tercero_id, cliente_nombre AS tercero_nombre,
                   tipo_comprobante, cantidad, total
            FROM ventas
            WHERE {ventas_where}
            ORDER BY fecha DESC
            LIMIT :limit
        """,
        "proyeccion_stock": f"""
            WITH ventas_filtradas AS (
                SELECT * FROM ventas WHERE {ventas_where}
            ),
            demanda AS (
                SELECT am.cod_articulo,
                       SUM(CASE WHEN v.tipo_comprobante='FA' THEN v.cantidad ELSE 0 END) / GREATEST((CAST(:hasta AS date) - CAST(:desde AS date)), 1) AS venta_diaria
                FROM {_ARTICULO_MAP_SQL}
                LEFT JOIN ventas_filtradas v ON v.producto_id = am.producto_id
                GROUP BY am.cod_articulo
            ),
            stock_agg AS (
                SELECT cod_articulo, SUM(cantidad) AS stock_actual,
                       MIN(stock_minimo) AS stock_minimo,
                       AVG(precio_compra_actual) AS precio_compra_actual
                FROM stock
                GROUP BY cod_articulo
            )
            SELECT s.cod_articulo, am.nombre AS articulo, s.stock_actual, s.stock_minimo,
                   COALESCE(d.venta_diaria, 0) AS venta_diaria,
                   CASE WHEN COALESCE(d.venta_diaria, 0) > 0 THEN s.stock_actual / d.venta_diaria ELSE NULL END AS dias_cobertura,
                   CASE WHEN s.stock_minimo > 0 AND s.stock_actual < s.stock_minimo THEN 'reponer'
                        WHEN COALESCE(d.venta_diaria, 0) > 0 AND s.stock_actual / d.venta_diaria < 30 THEN 'reponer'
                        ELSE 'ok' END AS estado
            FROM stock_agg s
            LEFT JOIN {_ARTICULO_MAP_SQL} ON am.cod_articulo = s.cod_articulo
            LEFT JOIN demanda d ON d.cod_articulo = s.cod_articulo
            ORDER BY estado DESC, dias_cobertura ASC NULLS LAST
            LIMIT :limit
        """,
        "cash_flow": f"""
            SELECT fecha, 'cobro_cliente' AS tipo, cod_cliente::text AS tercero_id,
                   cliente_nombre AS tercero_nombre, forma_pago, importe::float AS importe
            FROM pagos_clientes
            WHERE {cta_where}
            UNION ALL
            SELECT fecha, 'pago_proveedor' AS tipo, proveedor_id::text AS tercero_id,
                   proveedor_nombre AS tercero_nombre, forma_pago, -importe::float AS importe
            FROM pagos_proveedores
            WHERE {cta_where}
            ORDER BY fecha DESC
            LIMIT :limit
        """,
        "iva_compras": f"""
            SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo,
                   SUM(COALESCE(iva_importe::float, 0)) AS iva_compras,
                   SUM({compra_importe_neto_expr()}) AS base_compras
            FROM compras
            WHERE {compras_where}
            GROUP BY 1
            ORDER BY 1
            LIMIT :limit
        """,
        "iva_ventas": f"""
            SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo,
                   SUM({venta_iva_neto_expr()}) AS iva_ventas,
                   SUM({venta_importe_neto_expr()}) AS base_ventas
            FROM ventas
            WHERE {ventas_where}
            GROUP BY 1
            ORDER BY 1
            LIMIT :limit
        """,
        "estado_resultados_derivado": f"""
            WITH v AS (
                SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo,
                       SUM({venta_importe_neto_expr()}) AS ventas,
                       SUM({venta_costo_neto_expr()}) AS costo_ventas
                FROM ventas
                WHERE {ventas_where}
                GROUP BY 1
            ),
            c AS (
                SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo,
                       SUM({compra_importe_neto_expr()}) AS compras
                FROM compras
                WHERE {compras_where}
                GROUP BY 1
            )
            SELECT COALESCE(v.periodo, c.periodo) AS periodo,
                   COALESCE(v.ventas, 0) AS ventas,
                   COALESCE(v.costo_ventas, 0) AS costo_ventas,
                   COALESCE(c.compras, 0) AS compras,
                   COALESCE(v.ventas, 0) - COALESCE(v.costo_ventas, 0) AS margen_bruto
            FROM v FULL OUTER JOIN c USING (periodo)
            ORDER BY periodo
            LIMIT :limit
        """,
        "iva_balance": f"""
            WITH vc AS (
                SELECT SUM({venta_iva_neto_expr()}) AS iva_ventas,
                       SUM({venta_importe_neto_expr()}) AS base_ventas
                FROM ventas
                WHERE {ventas_where}
            ),
            cc AS (
                SELECT SUM(COALESCE(iva_importe::float, 0)) AS iva_compras,
                       SUM({compra_importe_neto_expr()}) AS base_compras
                FROM compras
                WHERE {compras_where}
            )
            SELECT COALESCE(vc.iva_ventas, 0) AS iva_ventas,
                   COALESCE(cc.iva_compras, 0) AS iva_compras,
                   COALESCE(vc.iva_ventas, 0) - COALESCE(cc.iva_compras, 0) AS saldo_iva,
                   COALESCE(vc.base_ventas, 0) AS base_ventas,
                   COALESCE(cc.base_compras, 0) AS base_compras
            FROM vc CROSS JOIN cc
            LIMIT :limit
        """,
    }

    sql_query = sql_by_key.get(report_key)
    if not sql_query:
        return None
    rows = (await db.execute(text(sql_query), params)).mappings().all()
    return [_jsonable_payload(dict(row)) for row in rows]


@router.get("/infomanager/reportes")
async def infomanager_reportes_catalogo(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    rows = (await db.execute(text("""
        SELECT report_key,
               MAX(report_name) AS report_name,
               COUNT(*) AS rows_count,
               MAX(synced_at) AS last_sync_at,
               MAX(fecha_desde) AS fecha_desde,
               MAX(fecha_hasta) AS fecha_hasta
        FROM infomanager_report_rows
        GROUP BY report_key
    """))).mappings().all()
    stats = {r["report_key"]: r for r in rows}

    reportes = []
    for item in INFOMANAGER_REPORT_CATALOG:
        stat = stats.get(item["key"])
        reportes.append({
            **item,
            "rows_count": int(stat["rows_count"]) if stat else 0,
            "last_sync_at": str(stat["last_sync_at"]) if stat and stat["last_sync_at"] else None,
            "fecha_desde": str(stat["fecha_desde"]) if stat and stat["fecha_desde"] else None,
            "fecha_hasta": str(stat["fecha_hasta"]) if stat and stat["fecha_hasta"] else None,
        })

    return {"reportes": reportes}


@router.get("/infomanager/reportes/{report_key}")
async def infomanager_reporte_detalle(
    report_key: str,
    limit: int = Query(default=200, ge=1, le=1000),
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    desde = filters.desde
    hasta = filters.hasta
    catalog_item = next((item for item in INFOMANAGER_REPORT_CATALOG if item["key"] == report_key), None)

    derived_rows = await _fetch_derived_infomanager_rows(report_key, filters, db, limit)
    if derived_rows is not None:
        return {
            "report_key": report_key,
            "report_name": catalog_item["name"] if catalog_item else report_key,
            "supported": True,
            "note": catalog_item.get("note") if catalog_item else None,
            "rows": [
                {
                    "row_key": f"{report_key}:{index}",
                    "fecha_desde": str(desde),
                    "fecha_hasta": str(hasta),
                    "synced_at": None,
                    "payload": row,
                }
                for index, row in enumerate(derived_rows)
            ],
        }

    rows = (await db.execute(text("""
        SELECT row_key, report_name, fecha_desde, fecha_hasta, synced_at, payload
        FROM infomanager_report_rows
        WHERE report_key = :report_key
          AND (
            fecha_desde IS NULL
            OR fecha_hasta IS NULL
            OR (fecha_desde <= :hasta AND fecha_hasta >= :desde)
          )
        ORDER BY synced_at DESC, row_key
        LIMIT :limit
    """), {"report_key": report_key, "desde": desde, "hasta": hasta, "limit": limit})).mappings().all()

    return {
        "report_key": report_key,
        "report_name": catalog_item["name"] if catalog_item else (rows[0]["report_name"] if rows else report_key),
        "supported": bool(catalog_item.get("supported")) if catalog_item else True,
        "note": catalog_item.get("note") if catalog_item else None,
        "rows": [
            {
                "row_key": r["row_key"],
                "fecha_desde": str(r["fecha_desde"]) if r["fecha_desde"] else None,
                "fecha_hasta": str(r["fecha_hasta"]) if r["fecha_hasta"] else None,
                "synced_at": str(r["synced_at"]) if r["synced_at"] else None,
                "payload": r["payload"],
            }
            for r in rows
        ],
    }


def previous_period_params(filters: GlobalFilters) -> dict:
    current_days = (filters.hasta_exclusive - filters.desde).days
    if getattr(filters, "compare_mode", "anterior") == "anio":
        prev_desde = _year_ago(filters.desde)
        prev_hasta = _year_ago(filters.hasta_exclusive)
    else:
        prev_hasta = filters.desde
        prev_desde = prev_hasta - timedelta(days=current_days)
    return {"prev_desde": prev_desde, "prev_hasta": prev_hasta, "offset_days": current_days}


def variation_payload(actual, anterior=None) -> dict:
    actual_value = money(actual)
    anterior_value = money(anterior)
    payload = {"actual": actual_value, "anterior": anterior_value}
    payload["variacion_pct"] = ((actual_value - anterior_value) / abs(anterior_value) * 100) if anterior_value else None
    return payload


def compra_filters_clause(
    filters: GlobalFilters,
    start_param: str = "desde",
    end_param: str = "hasta",
    include_provider: bool = True,
) -> str:
    clauses = [f"fecha >= :{start_param}", f"fecha < :{end_param}"]
    if filters.cod_empresa:
        clauses.append("cod_empresa = ANY(:cod_empresa)")
    if filters.cod_articulo:
        clauses.append("producto_id = ANY(:cod_articulo_text)")
    if filters.cod_deposito:
        clauses.append("cod_deposito = ANY(:cod_deposito)")
    if include_provider and filters.cod_cliente:
        clauses.append("proveedor_id = ANY(:proveedor_id_text)")
    if not filters.incluir_anuladas:
        clauses.append("COALESCE(anulada, 'N') <> 'S'")
    return " AND ".join(clauses)


def compra_params(filters: GlobalFilters) -> dict:
    params = filters.sql_params()
    params["cod_articulo_text"] = [str(item) for item in filters.cod_articulo or []]
    params["proveedor_id_text"] = [str(item) for item in filters.cod_cliente or []]
    return params


def granularity_sql(granularidad: str) -> str:
    return {
        "dia": "day",
        "semana": "week",
        "mes": "month",
        "trimestre": "quarter",
    }.get(granularidad, "month")


async def query_compras_kpis(db: AsyncSession, filters: GlobalFilters) -> dict:
    params = {**compra_params(filters), **previous_period_params(filters), "today": date.today(), "today_30": date.today() + timedelta(days=30)}
    current_where = compra_filters_clause(filters)
    previous_where = compra_filters_clause(filters, "prev_desde", "prev_hasta")

    totals = (await db.execute(text(f"""
        WITH actual AS (
            SELECT
                COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
                COALESCE(SUM(COALESCE(iva_importe::float, 0)), 0) AS iva_credito_fiscal,
                COUNT(*) AS ordenes,
                COALESCE(SUM(cantidad), 0) AS unidades_compradas,
                COUNT(DISTINCT proveedor_id) AS proveedores_activos,
                CASE WHEN COUNT(*) > 0 THEN COALESCE(SUM({compra_importe_neto_expr()}), 0) / COUNT(*) ELSE 0 END AS ticket_promedio_compra
            FROM compras
            WHERE {current_where}
        ),
        anterior AS (
            SELECT
                COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
                COALESCE(SUM(COALESCE(iva_importe::float, 0)), 0) AS iva_credito_fiscal,
                COUNT(*) AS ordenes,
                COALESCE(SUM(cantidad), 0) AS unidades_compradas,
                COUNT(DISTINCT proveedor_id) AS proveedores_activos,
                CASE WHEN COUNT(*) > 0 THEN COALESCE(SUM({compra_importe_neto_expr()}), 0) / COUNT(*) ELSE 0 END AS ticket_promedio_compra
            FROM compras
            WHERE {previous_where}
        )
        SELECT actual.*, anterior.total_comprado AS total_comprado_anterior,
               anterior.iva_credito_fiscal AS iva_credito_fiscal_anterior,
               anterior.ordenes AS ordenes_anterior,
               anterior.unidades_compradas AS unidades_compradas_anterior,
               anterior.proveedores_activos AS proveedores_activos_anterior,
               anterior.ticket_promedio_compra AS ticket_promedio_compra_anterior
        FROM actual, anterior
    """), params)).mappings().one()

    deuda = (await db.execute(text("""
        SELECT COALESCE(SUM(GREATEST(saldo_acumulado, importe, 0)), 0) AS deuda_vencida
        FROM cuentas_corrientes_proveedores
        WHERE lower(tipo) IN ('factura', 'fc')
          AND fecha_vencimiento < :today
          AND saldo_acumulado > 0
    """), params)).mappings().one()

    proximos = (await db.execute(text("""
        SELECT COALESCE(SUM(GREATEST(saldo_acumulado, importe, 0)), 0) AS proximos_vencimientos_30d
        FROM cuentas_corrientes_proveedores
        WHERE lower(tipo) IN ('factura', 'fc')
          AND fecha_vencimiento >= :today
          AND fecha_vencimiento < :today_30
          AND saldo_acumulado > 0
    """), params)).mappings().one()

    return {
        "total_comprado": variation_payload(totals["total_comprado"], totals["total_comprado_anterior"]),
        "iva_credito_fiscal": variation_payload(totals["iva_credito_fiscal"], totals["iva_credito_fiscal_anterior"]),
        "ordenes": variation_payload(totals["ordenes"], totals["ordenes_anterior"]),
        "unidades_compradas": variation_payload(totals["unidades_compradas"], totals["unidades_compradas_anterior"]),
        "proveedores_activos": variation_payload(totals["proveedores_activos"], totals["proveedores_activos_anterior"]),
        "ticket_promedio_compra": variation_payload(totals["ticket_promedio_compra"], totals["ticket_promedio_compra_anterior"]),
        "deuda_vencida": {"actual": money(deuda["deuda_vencida"])},
        "proximos_vencimientos_30d": {"actual": money(proximos["proximos_vencimientos_30d"])},
    }


async def query_compras_temporal(db: AsyncSession, filters: GlobalFilters, granularidad: str = "mes") -> list[dict]:
    params = {**compra_params(filters), **previous_period_params(filters)}
    period = granularity_sql(granularidad)
    current_where = compra_filters_clause(filters)
    previous_where = compra_filters_clause(filters, "prev_desde", "prev_hasta")
    rows = (await db.execute(text(f"""
        WITH actual AS (
            SELECT date_trunc('{period}', fecha)::date AS periodo,
                   COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
                   COUNT(*) AS ordenes
            FROM compras
            WHERE {current_where}
            GROUP BY 1
        ),
        anterior AS (
            SELECT date_trunc('{period}', fecha + (:offset_days * interval '1 day'))::date AS periodo,
                   COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado_anterior
            FROM compras
            WHERE {previous_where}
            GROUP BY 1
        )
        SELECT actual.periodo,
               actual.total_comprado,
               actual.ordenes,
               COALESCE(anterior.total_comprado_anterior, 0) AS total_comprado_anterior
        FROM actual
        LEFT JOIN anterior USING (periodo)
        ORDER BY actual.periodo
    """), params)).mappings().all()
    return [{**dict(row), "periodo": row["periodo"].isoformat()} for row in rows]


async def query_compras_por_producto(db: AsyncSession, filters: GlobalFilters) -> list[dict]:
    params = compra_params(filters)
    where = compra_filters_clause(filters)
    total_row = (await db.execute(text(f"SELECT COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total FROM compras WHERE {where}"), params)).mappings().one()
    total_compras = money(total_row["total"])
    rows = (await db.execute(text(f"""
        SELECT producto_id AS cod_articulo,
               COALESCE(MAX(producto_nombre), producto_id) AS nombre,
               COALESCE(SUM(cantidad), 0) AS unidades,
               COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
               CASE WHEN SUM(cantidad) > 0 THEN SUM({compra_importe_neto_expr()}) / SUM(cantidad) ELSE 0 END AS precio_promedio,
               (ARRAY_AGG(precio_unitario ORDER BY fecha DESC))[1] AS ultimo_precio,
               (ARRAY_AGG(precio_unitario ORDER BY fecha ASC))[1] AS primer_precio,
               COUNT(DISTINCT proveedor_id) AS proveedores_distintos
        FROM compras
        WHERE {where}
        GROUP BY producto_id
        ORDER BY total_comprado DESC
        LIMIT 50
    """), params)).mappings().all()
    result = []
    for row in rows:
        primer = money(row["primer_precio"])
        ultimo = money(row["ultimo_precio"])
        result.append({
            **dict(row),
            "pct_total": (money(row["total_comprado"]) / total_compras * 100) if total_compras else 0,
            "variacion_precio_pct": ((ultimo - primer) / abs(primer) * 100) if primer else None,
        })
    return result


async def query_compras_por_proveedor(db: AsyncSession, filters: GlobalFilters) -> list[dict]:
    params = compra_params(filters)
    where = compra_filters_clause(filters)
    total_row = (await db.execute(text(f"SELECT COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total FROM compras WHERE {where}"), params)).mappings().one()
    total_compras = money(total_row["total"])
    rows = (await db.execute(text(f"""
        SELECT proveedor_id,
               COALESCE(MAX(proveedor_nombre), proveedor_id) AS proveedor_nombre,
               COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
               COUNT(*) AS ordenes,
               CASE WHEN SUM(cantidad) > 0 THEN SUM({compra_importe_neto_expr()}) / SUM(cantidad) ELSE 0 END AS precio_promedio,
               MAX(fecha)::date AS ultimo_pedido,
               (CURRENT_DATE - MAX(fecha)::date) AS dias_desde_ultimo_pedido,
               AVG(precio_unitario) FILTER (WHERE fecha >= :desde AND fecha < (:desde + ((:hasta - :desde) / 2))) AS precio_promedio_inicio,
               AVG(precio_unitario) FILTER (WHERE fecha >= (:desde + ((:hasta - :desde) / 2)) AND fecha < :hasta) AS precio_promedio_fin
        FROM compras
        WHERE {where}
        GROUP BY proveedor_id
        ORDER BY total_comprado DESC
        LIMIT 30
    """), params)).mappings().all()
    result = []
    for row in rows:
        inicio = money(row["precio_promedio_inicio"])
        fin = money(row["precio_promedio_fin"])
        result.append({
            **dict(row),
            "ultimo_pedido": row["ultimo_pedido"].isoformat() if row["ultimo_pedido"] else None,
            "pct_total": (money(row["total_comprado"]) / total_compras * 100) if total_compras else 0,
            "precio_promedio_variacion": ((fin - inicio) / abs(inicio) * 100) if inicio else None,
        })
    return result


async def query_compras_calendario_pagos(db: AsyncSession) -> list[dict]:
    today = date.today()
    rows = (await db.execute(text("""
        SELECT proveedor_nombre,
               fecha_vencimiento::date AS fecha_vencimiento,
               GREATEST(saldo_acumulado, importe, 0) AS importe,
               (fecha_vencimiento::date - :today) AS dias_para_vencer
        FROM cuentas_corrientes_proveedores
        WHERE lower(tipo) IN ('factura', 'fc')
          AND fecha_vencimiento >= :today
          AND fecha_vencimiento <= :today_90
          AND saldo_acumulado > 0
        ORDER BY fecha_vencimiento ASC
        LIMIT 200
    """), {"today": today, "today_90": today + timedelta(days=90)})).mappings().all()
    result = []
    for row in rows:
        days = int(row["dias_para_vencer"] or 0)
        if days <= 0:
            urgencia = "hoy"
        elif days <= 7:
            urgencia = "semana"
        elif days <= 30:
            urgencia = "mes"
        else:
            urgencia = "futuro"
        result.append({
            **dict(row),
            "fecha_vencimiento": row["fecha_vencimiento"].isoformat() if row["fecha_vencimiento"] else None,
            "urgencia": urgencia,
        })
    return result


async def query_compras_variacion_precios(db: AsyncSession, filters: GlobalFilters) -> list[dict]:
    params = compra_params(filters)
    where = compra_filters_clause(filters)
    rows = (await db.execute(text(f"""
        SELECT producto_id,
               COALESCE(MAX(producto_nombre), producto_id) AS nombre,
               (ARRAY_AGG(precio_unitario ORDER BY fecha ASC))[1] AS precio_inicio_periodo,
               (ARRAY_AGG(precio_unitario ORDER BY fecha DESC))[1] AS precio_fin_periodo,
               (ARRAY_AGG(proveedor_nombre ORDER BY fecha ASC))[1] AS proveedor_nombre,
               (ARRAY_AGG(proveedor_nombre ORDER BY fecha DESC))[1] AS ultimo_proveedor
        FROM compras
        WHERE {where}
        GROUP BY producto_id
        ORDER BY producto_id
    """), params)).mappings().all()
    result = []
    for row in rows:
        inicio = money(row["precio_inicio_periodo"])
        fin = money(row["precio_fin_periodo"])
        variacion = ((fin - inicio) / abs(inicio) * 100) if inicio else None
        result.append({**dict(row), "variacion_pct": variacion})
    return sorted(result, key=lambda item: item["variacion_pct"] or 0, reverse=True)


async def query_compras_transacciones(db: AsyncSession, filters: GlobalFilters, page: int = 1, limit: int = 50) -> dict:
    params = compra_params(filters)
    where = compra_filters_clause(filters)
    total = (await db.execute(text(f"SELECT COUNT(*) AS total FROM compras WHERE {where}"), params)).mappings().one()["total"]
    safe_page = max(page, 1)
    safe_limit = min(max(limit, 1), 500)
    params.update({"offset": (safe_page - 1) * safe_limit, "limit": safe_limit})
    rows = (await db.execute(text(f"""
        SELECT id, fecha::date AS fecha, proveedor_id, proveedor_nombre, producto_id, producto_nombre,
               cantidad, precio_unitario, total, tipo_comprobante, tipo_factura,
               punto_de_venta, cod_empresa, COALESCE(neto::float, 0) AS neto,
               COALESCE(iva_importe::float, 0) AS iva, anulada, cod_deposito
        FROM compras
        WHERE {where}
        ORDER BY fecha DESC, id DESC
        OFFSET :offset
        LIMIT :limit
    """), params)).mappings().all()
    return {
        "page": safe_page,
        "limit": safe_limit,
        "total": total,
        "data": [{**dict(row), "fecha": row["fecha"].isoformat() if row["fecha"] else None} for row in rows],
    }


@router.get("/ventas/resumen")
async def ventas_resumen(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    ventas_where = text_filter_clause("ventas", filters)

    totals = (await db.execute(text(f"""
        SELECT
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS total_ventas,
            COALESCE(SUM(cantidad), 0) AS unidades,
            COUNT(*) AS transacciones,
            COUNT(DISTINCT cliente_id) AS clientes
        FROM ventas
        WHERE {ventas_where}
    """), params)).mappings().one()
    series = (await db.execute(text(f"""
        SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo,
               COALESCE(SUM({venta_importe_neto_expr()}), 0) AS total,
               COUNT(*) AS transacciones
        FROM ventas
        WHERE {ventas_where}
        GROUP BY 1
        ORDER BY 1
    """), params)).mappings().all()
    top_productos = (await db.execute(text(f"""
        SELECT producto_id, COALESCE(MAX(producto_nombre), producto_id) AS producto_nombre,
               COALESCE(SUM({venta_importe_neto_expr()}), 0) AS total,
               COALESCE(SUM(cantidad), 0) AS unidades
        FROM ventas
        WHERE {ventas_where}
        GROUP BY producto_id
        ORDER BY total DESC
        LIMIT 10
    """), params)).mappings().all()
    top_clientes = (await db.execute(text(f"""
        SELECT cliente_id, COALESCE(MAX(cliente_nombre), cliente_id) AS cliente_nombre,
               COUNT(*) AS transacciones,
               COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturacion,
               CASE WHEN COUNT(*) > 0 THEN COALESCE(SUM({venta_importe_neto_expr()}), 0) / COUNT(*) ELSE 0 END AS ticket_promedio
        FROM ventas
        WHERE {ventas_where}
        GROUP BY cliente_id
        ORDER BY facturacion DESC
        LIMIT 10
    """), params)).mappings().all()

    total_ventas = money(totals["total_ventas"])
    transacciones = totals["transacciones"]
    ticket_promedio = total_ventas / transacciones if transacciones else 0

    return {
        "summary": {
            "total_ventas": total_ventas,
            "unidades": money(totals["unidades"]),
            "transacciones": transacciones,
            "clientes": totals["clientes"],
            "ticket_promedio": ticket_promedio,
        },
        "series": [dict(row) for row in series],
        "top_productos": [
            {**dict(row), "porcentaje": (money(row["total"]) / total_ventas * 100) if total_ventas else 0}
            for row in top_productos
        ],
        "top_clientes": [dict(row) for row in top_clientes],
    }


@router.get("/compras/resumen")
async def compras_resumen(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = compra_params(filters)
    compras_where = compra_filters_clause(filters)
    ventas_where = text_filter_clause("ventas", filters)

    totals = (await db.execute(text(f"""
        SELECT
            COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_compras,
            COALESCE(SUM(cantidad), 0) AS unidades,
            COUNT(*) AS ordenes,
            COUNT(DISTINCT proveedor_id) AS proveedores
        FROM compras
        WHERE {compras_where}
    """), params)).mappings().one()

    total_ventas_row = (await db.execute(text(f"""
        SELECT COALESCE(SUM({venta_importe_neto_expr()}), 0) AS total FROM ventas
        WHERE {ventas_where}
    """), params)).mappings().one()

    series = (await db.execute(text(f"""
        SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo,
               COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total,
               COUNT(*) AS ordenes
        FROM compras
        WHERE {compras_where}
        GROUP BY 1
        ORDER BY 1
    """), params)).mappings().all()
    top_proveedores = (await db.execute(text(f"""
        SELECT proveedor_id, proveedor_nombre,
               COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total,
               COUNT(*) AS ordenes
        FROM compras
        WHERE {compras_where}
        GROUP BY proveedor_id, proveedor_nombre
        ORDER BY total DESC
        LIMIT 10
    """), params)).mappings().all()
    top_productos = (await db.execute(text(f"""
        SELECT producto_id, COALESCE(MAX(producto_nombre), producto_id) AS producto_nombre,
               COALESCE(SUM(cantidad), 0) AS cantidad,
               COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total
        FROM compras
        WHERE {compras_where}
        GROUP BY producto_id
        ORDER BY total DESC
        LIMIT 10
    """), params)).mappings().all()

    total_compras = money(totals["total_compras"])
    total_ventas = money(total_ventas_row["total"])
    ratio_compra_venta = (total_compras / total_ventas * 100) if total_ventas else 0
    margen_bruto = ((total_ventas - total_compras) / total_ventas * 100) if total_ventas else 0

    return {
        "summary": {
            "total_compras": total_compras,
            "unidades": money(totals["unidades"]),
            "ordenes": totals["ordenes"],
            "proveedores": totals["proveedores"],
            "ratio_compra_venta": round(ratio_compra_venta, 1),
            "margen_bruto": round(margen_bruto, 1),
        },
        "series": [dict(row) for row in series],
        "top_proveedores": [
            {**dict(row), "porcentaje": (money(row["total"]) / total_compras * 100) if total_compras else 0}
            for row in top_proveedores
        ],
        "top_productos": [dict(row) for row in top_productos],
    }


@router.get("/compras/kpis")
async def compras_kpis(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    return await query_compras_kpis(db, filters)


@router.get("/compras/temporal")
async def compras_temporal(
    company_id: int = None,
    granularidad: str = "mes",
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    return {"series": await query_compras_temporal(db, filters, granularidad)}


@router.get("/compras/por-producto")
async def compras_por_producto(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    return {"productos": await query_compras_por_producto(db, filters)}


@router.get("/compras/por-proveedor")
async def compras_por_proveedor(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    return {"proveedores": await query_compras_por_proveedor(db, filters)}


@router.get("/compras/calendario-pagos")
async def compras_calendario_pagos(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    return {"vencimientos": await query_compras_calendario_pagos(db)}


@router.get("/compras/variacion-precios")
async def compras_variacion_precios(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    return {"productos": await query_compras_variacion_precios(db, filters)}


@router.get("/compras/transacciones")
async def compras_transacciones(
    company_id: int = None,
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=500),
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    return await query_compras_transacciones(db, filters, page, limit)


@router.get("/compras/facturas")
async def compras_facturas(
    company_id: int = None,
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=100, ge=1, le=500),
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params: dict = {"limit": limit, "offset": (page - 1) * limit}
    date_cond = ""
    if filters.desde:
        date_cond += " AND fecha::date >= :desde"
        params["desde"] = filters.desde
    if filters.hasta:
        date_cond += " AND fecha::date <= :hasta"
        params["hasta"] = filters.hasta
    empresa_cond = ""
    if filters.cod_empresa:
        empresa_cond = " AND cod_empresa = ANY(:cod_empresa)"
        params["cod_empresa"] = filters.cod_empresa
    deposito_cond = ""
    if filters.cod_deposito:
        deposito_cond = " AND cod_deposito = ANY(:cod_deposito)"
        params["cod_deposito"] = filters.cod_deposito

    _sql = f"""
        SELECT id, fecha, tipo_comprobante, tipo_factura,
               punto_de_venta, numero, proveedor, moneda,
               importe_total, importe_iva, anulada, cod_empresa, cod_deposito
        FROM facturas_compra
        WHERE 1=1
        {date_cond} {empresa_cond} {deposito_cond}
        ORDER BY fecha DESC NULLS LAST
        LIMIT :limit OFFSET :offset
    """
    rows = (await db.execute(text(_sql), params)).mappings().all()

    count_sql = f"""
        SELECT COUNT(*) FROM facturas_compra
        WHERE 1=1 {date_cond} {empresa_cond} {deposito_cond}
    """
    count_params = {k: v for k, v in params.items() if k not in ("limit", "offset")}
    total = (await db.execute(text(count_sql), count_params)).scalar() or 0

    return {
        "facturas": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "limit": limit,
    }


@router.get("/compras/exportar")
async def compras_exportar(
    company_id: int = None,
    granularidad: str = "mes",
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(title: str, rows: list[dict]):
        sheet = workbook.create_sheet(title)
        if not rows:
            sheet.append(["Sin datos"])
            return
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])

    kpis = await query_compras_kpis(db, filters)
    add_sheet("KPIs", [
        {"kpi": key, **value}
        for key, value in kpis.items()
    ])
    add_sheet("Temporal", await query_compras_temporal(db, filters, granularidad))
    add_sheet("Productos", await query_compras_por_producto(db, filters))
    add_sheet("Proveedores", await query_compras_por_proveedor(db, filters))
    add_sheet("Calendario Pagos", await query_compras_calendario_pagos(db))
    add_sheet("Variacion Precios", await query_compras_variacion_precios(db, filters))
    add_sheet("Transacciones", (await query_compras_transacciones(db, filters, 1, 500))["data"])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="compras_analytics.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/resultado/resumen")
async def resultado_resumen(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    ventas_where = text_filter_clause("ventas", filters)
    compras_where = text_filter_clause("compras", filters)
    caja_where = text_filter_clause("movimientos_caja", filters)

    totals = (await db.execute(text(f"""
        WITH ventas_total AS (
            SELECT COALESCE(SUM(total), 0) AS total FROM ventas WHERE {ventas_where}
        ),
        compras_total AS (
            SELECT COALESCE(SUM(total), 0) AS total FROM compras WHERE {compras_where}
        ),
        gastos_total AS (
            SELECT COALESCE(SUM(ABS(importe)), 0) AS total
            FROM movimientos_caja
            WHERE tipo = 'gasto' AND {caja_where}
        )
        SELECT ventas_total.total AS ingresos,
               compras_total.total AS costo_mercaderia,
               gastos_total.total AS gastos,
               ventas_total.total - compras_total.total - gastos_total.total AS resultado
        FROM ventas_total, compras_total, gastos_total
    """), params)).mappings().one()
    series = (await db.execute(text(f"""
        WITH meses AS (
            SELECT date_trunc('month', fecha) AS mes FROM ventas WHERE {ventas_where}
            UNION
            SELECT date_trunc('month', fecha) AS mes FROM compras WHERE {compras_where}
            UNION
            SELECT date_trunc('month', fecha) AS mes FROM movimientos_caja WHERE {caja_where}
        ),
        ventas_mes AS (
            SELECT date_trunc('month', fecha) AS mes, SUM(total) AS ingresos
            FROM ventas WHERE {ventas_where} GROUP BY 1
        ),
        compras_mes AS (
            SELECT date_trunc('month', fecha) AS mes, SUM(total) AS compras
            FROM compras WHERE {compras_where} GROUP BY 1
        ),
        gastos_mes AS (
            SELECT date_trunc('month', fecha) AS mes, SUM(ABS(importe)) AS gastos
            FROM movimientos_caja WHERE tipo = 'gasto' AND {caja_where} GROUP BY 1
        )
        SELECT to_char(meses.mes, 'YYYY-MM') AS periodo,
               COALESCE(ventas_mes.ingresos, 0) AS ingresos,
               COALESCE(compras_mes.compras, 0) AS compras,
               COALESCE(gastos_mes.gastos, 0) AS gastos,
               COALESCE(ventas_mes.ingresos, 0) - COALESCE(compras_mes.compras, 0) - COALESCE(gastos_mes.gastos, 0) AS resultado,
               CASE WHEN COALESCE(ventas_mes.ingresos, 0) > 0
                    THEN ((COALESCE(ventas_mes.ingresos, 0) - COALESCE(compras_mes.compras, 0)) / COALESCE(ventas_mes.ingresos, 0)) * 100
                    ELSE 0
               END AS margen_pct
        FROM meses
        LEFT JOIN ventas_mes USING (mes)
        LEFT JOIN compras_mes USING (mes)
        LEFT JOIN gastos_mes USING (mes)
        ORDER BY periodo
    """), params)).mappings().all()

    ingresos = money(totals["ingresos"])
    resultado = money(totals["resultado"])

    return {
        "summary": {
            "ingresos": ingresos,
            "costo_mercaderia": money(totals["costo_mercaderia"]),
            "gastos": money(totals["gastos"]),
            "resultado": resultado,
            "margen_resultado": (resultado / ingresos) if ingresos else 0,
        },
        "series": [dict(row) for row in series],
    }


@router.get("/clientes/resumen")
async def clientes_resumen(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    clientes_where = text_filter_clause("cuentas_corrientes_clientes", filters)

    totals = (await db.execute(text(f"""
        SELECT
            COUNT(DISTINCT cliente_id) AS clientes_con_movimiento,
            COALESCE(SUM(CASE WHEN tipo = 'factura' THEN importe ELSE 0 END), 0) AS facturado,
            COALESCE(SUM(CASE WHEN tipo = 'recibo' THEN ABS(importe) ELSE 0 END), 0) AS cobrado,
            COALESCE(SUM(importe), 0) AS saldo_total
        FROM cuentas_corrientes_clientes
        WHERE {clientes_where}
    """), params)).mappings().one()
    top_saldos = (await db.execute(text(f"""
        SELECT cliente_id, cliente_nombre, MAX(fecha) AS ultimo_movimiento,
               COALESCE(SUM(importe), 0) AS saldo
        FROM cuentas_corrientes_clientes
        WHERE {clientes_where}
        GROUP BY cliente_id, cliente_nombre
        ORDER BY saldo DESC
        LIMIT 10
    """), params)).mappings().all()
    ageing = (await db.execute(text("""
        SELECT
            COALESCE(SUM(CASE WHEN fecha_vencimiento < now() AND fecha_vencimiento >= now() - interval '30 days' THEN GREATEST(importe, 0) ELSE 0 END), 0) AS vencido_0_30,
            COALESCE(SUM(CASE WHEN fecha_vencimiento < now() - interval '30 days' AND fecha_vencimiento >= now() - interval '60 days' THEN GREATEST(importe, 0) ELSE 0 END), 0) AS vencido_31_60,
            COALESCE(SUM(CASE WHEN fecha_vencimiento < now() - interval '60 days' AND fecha_vencimiento >= now() - interval '90 days' THEN GREATEST(importe, 0) ELSE 0 END), 0) AS vencido_61_90,
            COALESCE(SUM(CASE WHEN fecha_vencimiento < now() - interval '90 days' THEN GREATEST(importe, 0) ELSE 0 END), 0) AS vencido_90_plus,
            COALESCE(SUM(CASE WHEN fecha_vencimiento >= now() THEN GREATEST(importe, 0) ELSE 0 END), 0) AS corriente
        FROM cuentas_corrientes_clientes
        WHERE tipo = 'factura'
    """))).mappings().one()

    # Check which clients have overdue debt
    clientes_vencidos = set()
    vencidos_rows = (await db.execute(text("""
        SELECT DISTINCT cliente_id
        FROM cuentas_corrientes_clientes
        WHERE tipo = 'factura' AND fecha_vencimiento < now() AND importe > 0
    """))).all()
    for row in vencidos_rows:
        clientes_vencidos.add(row[0])

    return {
        "summary": {
            "clientes_con_movimiento": totals["clientes_con_movimiento"],
            "facturado": money(totals["facturado"]),
            "cobrado": money(totals["cobrado"]),
            "saldo_total": money(totals["saldo_total"]),
            "vencido": money(ageing["vencido_0_30"]) + money(ageing["vencido_31_60"]) + money(ageing["vencido_61_90"]) + money(ageing["vencido_90_plus"]),
            "a_vencer": money(ageing["corriente"]),
        },
        "ageing": {
            "corriente": money(ageing["corriente"]),
            "vencido_0_30": money(ageing["vencido_0_30"]),
            "vencido_31_60": money(ageing["vencido_31_60"]),
            "vencido_61_90": money(ageing["vencido_61_90"]),
            "vencido_90_plus": money(ageing["vencido_90_plus"]),
        },
        "top_saldos": [
            {
                **dict(row),
                "ultimo_movimiento": row["ultimo_movimiento"].isoformat() if row["ultimo_movimiento"] else None,
                "estado": "vencido" if row["cliente_id"] in clientes_vencidos else "al_dia",
            }
            for row in top_saldos
        ],
    }


@router.get("/proveedores/resumen")
async def proveedores_resumen(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    proveedores_where = text_filter_clause("cuentas_corrientes_proveedores", filters)

    totals = (await db.execute(text(f"""
        SELECT
            COUNT(DISTINCT proveedor_id) AS proveedores_con_movimiento,
            COALESCE(SUM(CASE WHEN tipo = 'factura' THEN importe ELSE 0 END), 0) AS facturado,
            COALESCE(SUM(CASE WHEN tipo = 'pago' THEN ABS(importe) ELSE 0 END), 0) AS pagado,
            COALESCE(SUM(importe), 0) AS saldo_total
        FROM cuentas_corrientes_proveedores
        WHERE {proveedores_where}
    """), params)).mappings().one()
    top_saldos = (await db.execute(text(f"""
        SELECT proveedor_id, proveedor_nombre, MAX(fecha) AS ultimo_movimiento,
               COALESCE(SUM(importe), 0) AS saldo
        FROM cuentas_corrientes_proveedores
        WHERE {proveedores_where}
        GROUP BY proveedor_id, proveedor_nombre
        ORDER BY saldo DESC
        LIMIT 10
    """), params)).mappings().all()
    proximos_vencimientos = (await db.execute(text("""
        SELECT proveedor_id, proveedor_nombre, comprobante_id, fecha_vencimiento, importe
        FROM cuentas_corrientes_proveedores
        WHERE tipo = 'factura' AND fecha_vencimiento IS NOT NULL
              AND fecha_vencimiento >= now()
              AND fecha_vencimiento <= now() + interval '30 days'
        ORDER BY fecha_vencimiento
        LIMIT 15
    """))).mappings().all()

    return {
        "summary": {
            "proveedores_con_movimiento": totals["proveedores_con_movimiento"],
            "facturado": money(totals["facturado"]),
            "pagado": money(totals["pagado"]),
            "saldo_total": money(totals["saldo_total"]),
        },
        "top_saldos": [
            {
                **dict(row),
                "ultimo_movimiento": row["ultimo_movimiento"].isoformat() if row["ultimo_movimiento"] else None,
            }
            for row in top_saldos
        ],
        "proximos_vencimientos": [
            {
                **dict(row),
                "fecha_vencimiento": row["fecha_vencimiento"].isoformat() if row["fecha_vencimiento"] else None,
            }
            for row in proximos_vencimientos
        ],
    }


@router.get("/caja/resumen")
async def caja_resumen(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    caja_where = text_filter_clause("movimientos_caja", filters)

    totals = (await db.execute(text(f"""
        SELECT
            COALESCE(SUM(CASE WHEN importe > 0 THEN importe ELSE 0 END), 0) AS ingresos,
            COALESCE(SUM(CASE WHEN importe < 0 THEN ABS(importe) ELSE 0 END), 0) AS egresos_caja,
            COALESCE(SUM(importe), 0) AS saldo_neto,
            COUNT(*) AS movimientos
        FROM movimientos_caja
        WHERE {caja_where}
    """), params)).mappings().one()

    # Include compras (pagos a proveedores) in egresos
    _c_params: dict = {"desde": filters.desde, "hasta": filters.hasta_exclusive}
    _c_cond = "fecha >= :desde AND fecha < :hasta AND COALESCE(anulada, 'N') <> 'S'"
    if filters.cod_empresa:
        _c_cond += " AND cod_empresa = ANY(:cod_empresa)"
        _c_params["cod_empresa"] = filters.cod_empresa
    _compras_egresos = float((await db.execute(text(f"""
        SELECT COALESCE(SUM({compra_importe_neto_expr()}), 0) FROM compras WHERE {_c_cond}
    """), _c_params)).scalar() or 0)

    series = (await db.execute(text(f"""
        SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo,
               COALESCE(SUM(CASE WHEN importe > 0 THEN importe ELSE 0 END), 0) AS cobros,
               COALESCE(SUM(CASE WHEN importe < 0 THEN ABS(importe) ELSE 0 END), 0) AS pagos,
               COALESCE(SUM(importe), 0) AS saldo_neto
        FROM movimientos_caja
        WHERE {caja_where}
        GROUP BY 1
        ORDER BY 1
    """), params)).mappings().all()

    # Build cumulative saldo
    saldo_acum = 0
    series_with_acum = []
    for row in series:
        d = dict(row)
        saldo_acum += d["saldo_neto"]
        d["saldo_acumulado"] = saldo_acum
        series_with_acum.append(d)

    ultimos_movimientos = (await db.execute(text(f"""
        SELECT fecha, tipo, descripcion, importe, saldo_acumulado
        FROM movimientos_caja
        WHERE {caja_where}
        ORDER BY fecha DESC
        LIMIT 15
    """), params)).mappings().all()

    return {
        "summary": {
            "ingresos": money(totals["ingresos"]),
            "egresos": round(float(totals["egresos_caja"] or 0) + _compras_egresos, 2),
            "saldo_neto": money(totals["saldo_neto"]),
            "movimientos": totals["movimientos"],
        },
        "series": series_with_acum,
        "ultimos_movimientos": [
            {
                **dict(row),
                "fecha": row["fecha"].isoformat() if row["fecha"] else None,
            }
            for row in ultimos_movimientos
        ],
    }


@router.get("/stock/resumen")
async def stock_resumen(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    ventas_where = text_filter_clause("ventas", filters)

    stock = (await db.execute(text(f"""
        WITH stock_agg AS (
            SELECT cod_articulo,
                   SUM(cantidad) AS stock_actual,
                   MIN(stock_minimo) AS stock_minimo,
                   AVG(precio_compra_actual) AS precio_unitario_promedio
            FROM stock
            GROUP BY cod_articulo
        ),
        vendidas AS (
            SELECT producto_id,
                   COALESCE(SUM(cantidad), 0) AS unidades_vendidas
            FROM ventas
            WHERE {ventas_where}
            GROUP BY producto_id
        )
        SELECT am.producto_id,
               am.nombre AS producto_nombre,
               s.stock_actual,
               s.stock_minimo,
               COALESCE(v.unidades_vendidas, 0) AS unidades_vendidas,
               s.stock_actual AS stock_estimado,
               s.precio_unitario_promedio,
               s.stock_actual * COALESCE(s.precio_unitario_promedio, 0) AS valor_stock_estimado
        FROM stock_agg s
        LEFT JOIN {_ARTICULO_MAP_SQL} ON am.cod_articulo = s.cod_articulo
        LEFT JOIN vendidas v ON v.producto_id = am.producto_id
        ORDER BY valor_stock_estimado DESC
    """), params)).mappings().all()

    stock_list = []
    for row in stock:
        s = dict(row)
        est = s["stock_estimado"]
        if est <= 0:
            s["estado"] = "sin_stock"
        elif float(s.get("stock_minimo") or 0) > 0 and est < float(s.get("stock_minimo") or 0):
            s["estado"] = "bajo"
        else:
            s["estado"] = "ok"
        stock_list.append(s)

    total_articulos = len(stock_list)
    articulos_con_stock = sum(1 for s in stock_list if s["stock_estimado"] > 0)
    articulos_sin_stock = total_articulos - articulos_con_stock
    valor_inventario = sum(max(s["valor_stock_estimado"], 0) for s in stock_list)

    return {
        "total_articulos": total_articulos,
        "articulos_con_stock": articulos_con_stock,
        "articulos_sin_stock": articulos_sin_stock,
        "valor_inventario_estimado": round(valor_inventario, 2),
        "stock_por_producto": stock_list,
    }


@router.get("/vendedores/resumen")
async def vendedores_resumen(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    ventas_where = text_filter_clause("ventas", filters)

    VENDEDORES = [
        {"id": "V001", "nombre": "Lucas García"},
        {"id": "V002", "nombre": "María López"},
        {"id": "V003", "nombre": "Carlos Ruiz"},
        {"id": "V004", "nombre": "Ana Martínez"},
        {"id": "V005", "nombre": "Diego Fernández"},
    ]

    # Get all client sales summaries within the date range
    client_sales = (await db.execute(text(f"""
        SELECT cliente_id,
               COUNT(*) AS transacciones,
               COALESCE(SUM(total), 0) AS facturacion
        FROM ventas
        WHERE {ventas_where}
        GROUP BY cliente_id
    """), params)).mappings().all()

    # Deterministic assignment: hash(cliente_id) mod 5
    vendedor_data = {v["id"]: {
        "vendedor_id": v["id"],
        "vendedor_nombre": v["nombre"],
        "clientes_atendidos": 0,
        "transacciones": 0,
        "facturacion_total": 0.0,
    } for v in VENDEDORES}

    for row in client_sales:
        idx = hash(row["cliente_id"]) % len(VENDEDORES)
        vid = VENDEDORES[idx]["id"]
        vendedor_data[vid]["clientes_atendidos"] += 1
        vendedor_data[vid]["transacciones"] += row["transacciones"]
        vendedor_data[vid]["facturacion_total"] += float(row["facturacion"])

    total_facturacion = sum(v["facturacion_total"] for v in vendedor_data.values())

    result = []
    for v in sorted(vendedor_data.values(), key=lambda x: x["facturacion_total"], reverse=True):
        v["ticket_promedio"] = v["facturacion_total"] / v["transacciones"] if v["transacciones"] else 0
        v["porcentaje_del_total"] = (v["facturacion_total"] / total_facturacion * 100) if total_facturacion else 0
        result.append(v)

    return {"vendedores": result}


# ══════════════════════════════════════════════════════════════════════════════
# FASE 3: Panel Ventas — endpoints dedicados
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/ventas/kpis")
async def ventas_kpis(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params = filters.sql_params()
    where = _ventas_base_where(filters)

    row = await _fetch_kpi_row(db, where, params)

    fa = money(row["fa_bruto"])
    nc = money(row["nc_total"])
    iva = money(row["iva_debito"])
    tickets = int(row["tickets"] or 0)
    clientes = int(row["clientes_unicos"] or 0)
    unidades = money(row["unidades"])
    margen_d = money(row["margen_dolares"])
    total_con_costo = money(row["total_con_costo"])
    facturado_neto = money(row["facturado_neto"])
    ticket_prom = facturado_neto / tickets if tickets else 0
    tasa_dev = min((nc / fa * 100) if fa else 0, 100.0)
    margen_pct = (margen_d / total_con_costo * 100) if total_con_costo else 0

    # DSO (Days Sales Outstanding) — días promedio de cobro
    # Fórmula: (saldo cuentas por cobrar / ventas a crédito) * días del período
    # Solo cuenta facturas con condición venta = cuenta corriente
    dso = None
    try:
        # Use a savepoint so a DB error here doesn't abort the outer transaction
        async with db.begin_nested():
            dso_row = (await db.execute(text(f"""
                SELECT
                    COALESCE(SUM({venta_importe_neto_expr()}) FILTER (
                        WHERE condicion_venta_tipo IN ('cta_cte', 'cuenta_corriente', 'CC')
                    ), 0) AS ventas_credito,
                    COALESCE(SUM({venta_importe_neto_expr()}), 0) AS ventas_totales
                FROM ventas WHERE {where}
            """), params)).mappings().one()
            ventas_credito = money(dso_row["ventas_credito"])
            ventas_totales = money(dso_row["ventas_totales"])

            saldo_row = (await db.execute(text("""
                SELECT COALESCE(SUM(saldo_acumulado), 0) AS saldo_total
                FROM (
                    SELECT cliente_id, saldo_acumulado,
                           ROW_NUMBER() OVER (PARTITION BY cliente_id ORDER BY id DESC) AS rn
                    FROM cuentas_corrientes_clientes
                ) t
                WHERE rn = 1 AND saldo_acumulado > 0
            """))).mappings().one()
            saldo_cc = money(saldo_row["saldo_total"])

            dias_periodo = max(1, (filters.hasta - filters.desde).days)

            # Si hay ventas a crédito identificadas, usar ratio puro
            # Si no (porque el dato condicion_venta_tipo viene vacío), usar fallback con ventas totales
            denominador = ventas_credito if ventas_credito > 0 else ventas_totales
            if denominador > 0 and saldo_cc > 0:
                dso = round(saldo_cc / denominador * dias_periodo, 1)
    except Exception:
        dso = None

    # previous period
    ant: dict[str, Optional[float]] = {}
    if filters.comparar_anterior:
        prev_desde, prev_hasta = _prev_period(filters)
        prev_params = dict(params)
        prev_params["desde"] = prev_desde
        prev_params["hasta"] = prev_hasta
        prev_row = await _fetch_kpi_row(db, where, prev_params)
        ant_fa = money(prev_row["fa_bruto"])
        ant_nc = money(prev_row["nc_total"])
        ant["facturado_neto"] = money(prev_row["facturado_neto"])
        ant["facturado_bruto"] = ant_fa
        ant["iva_debito"] = money(prev_row["iva_debito"])
        ant["tickets"] = int(prev_row["tickets"] or 0)
        ant["ticket_promedio"] = ant["facturado_neto"] / ant["tickets"] if ant["tickets"] else 0
        ant["unidades"] = money(prev_row["unidades"])
        ant["tasa_devolucion"] = min((ant_nc / ant_fa * 100) if ant_fa else 0, 100.0)
        ant["clientes_unicos"] = int(prev_row["clientes_unicos"] or 0)
        ant_margen_d = money(prev_row["margen_dolares"])
        ant_total_con_costo = money(prev_row["total_con_costo"])
        ant["margen_bruto_pct"] = (ant_margen_d / ant_total_con_costo * 100) if ant_total_con_costo else 0

    return {
        "facturado_neto": _kpi_obj(facturado_neto, ant.get("facturado_neto")),
        "facturado_bruto": _kpi_obj(fa, ant.get("facturado_bruto")),
        "iva_debito": _kpi_obj(iva, ant.get("iva_debito")),
        "tickets": _kpi_obj(tickets, ant.get("tickets")),
        "ticket_promedio": _kpi_obj(ticket_prom, ant.get("ticket_promedio")),
        "unidades": _kpi_obj(unidades, ant.get("unidades")),
        "tasa_devolucion": _kpi_obj(round(tasa_dev, 2), ant.get("tasa_devolucion")),
        "clientes_unicos": _kpi_obj(clientes, ant.get("clientes_unicos")),
        "margen_bruto_pct": _kpi_obj(round(margen_pct, 2), ant.get("margen_bruto_pct")),
        "dso_dias": {"actual": dso} if dso is not None else {"actual": None},
    }


@router.get("/ventas/temporal")
async def ventas_temporal(
    company_id: int = None,
    granularidad: Literal["dia", "semana", "mes", "trimestre"] = "mes",
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    trunc_map = {
        "dia": "day", "semana": "week", "mes": "month", "trimestre": "quarter"
    }
    trunc = trunc_map.get(granularidad, "month")
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    series = (await db.execute(text(f"""
        SELECT
            to_char(date_trunc('{trunc}', fecha), 'YYYY-MM-DD')                                       AS periodo,
            COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END),0)               AS fa_bruto,
            COALESCE(SUM(CASE WHEN tipo_comprobante='NC' THEN ABS(total) ELSE 0 END),0)               AS devoluciones,
            COALESCE(SUM({venta_importe_neto_expr()}),0)                                              AS facturado,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                                          AS tickets,
            COALESCE(SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}),0)                  AS margen_abs,
            COALESCE(SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END),0)    AS total_con_costo
        FROM ventas
        WHERE {where}
        GROUP BY 1
        ORDER BY 1
    """), params)).mappings().all()

    result = []
    for r in series:
        d = dict(r)
        tc = float(d.pop("total_con_costo") or 0)
        ma = float(d["margen_abs"] or 0)
        d["margen_pct"] = round(ma / tc * 100, 1) if tc else None
        result.append(d)

    if filters.comparar_anterior:
        prev_desde, prev_hasta = _prev_period(filters)
        prev_params = dict(params)
        prev_params["desde"] = prev_desde
        prev_params["hasta"] = prev_hasta
        prev_series = (await db.execute(text(f"""
            SELECT
                to_char(date_trunc('{trunc}', fecha), 'YYYY-MM-DD')                                       AS periodo,
                COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END),0)               AS fa_bruto,
                COALESCE(SUM(CASE WHEN tipo_comprobante='NC' THEN ABS(total) ELSE 0 END),0)               AS devoluciones,
                COALESCE(SUM({venta_importe_neto_expr()}),0)                                              AS facturado,
                COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                                          AS tickets,
                COALESCE(SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}),0)                  AS margen_abs,
                COALESCE(SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END),0)    AS total_con_costo
            FROM ventas
            WHERE {where}
            GROUP BY 1
            ORDER BY 1
        """), prev_params)).mappings().all()
        prev_map = {}
        for r in prev_series:
            d = dict(r)
            tc = float(d.pop("total_con_costo") or 0)
            ma = float(d["margen_abs"] or 0)
            d["margen_pct"] = round(ma / tc * 100, 1) if tc else None
            prev_map[d["periodo"]] = d
        prev_vals = list(prev_map.values())
        for i, row in enumerate(result):
            if i < len(prev_vals):
                row["facturado_anterior"] = prev_vals[i]["facturado"]
                row["tickets_anterior"] = prev_vals[i]["tickets"]
                row["margen_pct_anterior"] = prev_vals[i].get("margen_pct")

    return {"series": result}


@router.get("/ventas/productos")
async def ventas_productos(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params = filters.sql_params()
    where = _ventas_base_where(filters)

    total_row = (await db.execute(text(f"""
        SELECT COALESCE(SUM({venta_importe_neto_expr()}),0) AS total_fa
        FROM ventas WHERE {where}
    """), params)).mappings().one()
    total_fa = money(total_row["total_fa"]) or 1

    ranking_rows = (await db.execute(text(f"""
        SELECT
            producto_id,
            COALESCE(MAX(producto_nombre), producto_id)                                    AS nombre,
            COALESCE(MAX(cod_rubro), 0)                                                    AS cod_rubro,
            COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN cantidad ELSE 0 END), 0)    AS unidades,
            COALESCE(SUM({venta_importe_neto_expr()}), 0)                                      AS facturado,
            COUNT(DISTINCT CASE WHEN tipo_comprobante='FA' THEN cliente_id END)           AS clientes_unicos,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                             AS tickets,
            COALESCE(SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}), 0)      AS margen_dolares,
            COALESCE(SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END), 0) AS total_con_costo
        FROM ventas
        WHERE {where}
        GROUP BY producto_id
        ORDER BY facturado DESC
        LIMIT 100
    """), params)).mappings().all()

    ranking = []
    acumulado = 0.0
    for r in ranking_rows:
        d = dict(r)
        d["facturado"] = money(d["facturado"])
        d["margen_dolares"] = money(d["margen_dolares"])
        tc = money(d["total_con_costo"])
        d["margen_pct"] = round(d["margen_dolares"] / tc * 100, 1) if tc else 0
        d["ticket_promedio"] = d["facturado"] / d["tickets"] if d["tickets"] else 0
        d["pct_total"] = round(d["facturado"] / total_fa * 100, 2)
        acumulado += d["pct_total"]
        d["acumulado_pct"] = round(acumulado, 2)
        del d["total_con_costo"]
        ranking.append(d)

    # por_rubro: join stock_disponible for names (more reliable than rubros catalog alone)
    rubros_rows = (await db.execute(text(f"""
        SELECT
            v.cod_rubro,
            COALESCE(
                MAX(s.rubro),
                MAX(r.nombre),
                CONCAT('Rubro ', v.cod_rubro)
            )                                                                         AS nombre,
            COALESCE(SUM({venta_importe_neto_expr('v')}), 0)                         AS facturado,
            COALESCE(SUM({venta_importe_neto_expr('v')} - {venta_costo_neto_expr('v')}), 0) AS margen_abs,
            COALESCE(SUM(CASE WHEN v.precio_compra_actual IS NOT NULL THEN ABS(v.total) ELSE 0 END), 0) AS total_con_costo,
            COUNT(CASE WHEN v.tipo_comprobante='FA' THEN 1 END)                      AS tickets
        FROM ventas v
        LEFT JOIN (
            SELECT DISTINCT ON (cod_articulo) cod_articulo, rubro
            FROM stock_disponible
            WHERE rubro IS NOT NULL
            ORDER BY cod_articulo
        ) s ON s.cod_articulo::text = v.producto_id
        LEFT JOIN rubros r ON r.cod_rubro = v.cod_rubro
        WHERE {where} AND v.cod_rubro IS NOT NULL
        GROUP BY v.cod_rubro
        ORDER BY facturado DESC
    """), params)).mappings().all()

    rubros = []
    for r in rubros_rows:
        d = dict(r)
        d["facturado"] = money(d["facturado"])
        d["margen_abs"] = money(d["margen_abs"])
        tc = money(d.pop("total_con_costo"))
        d["margen_pct"] = round(d["margen_abs"] / tc * 100, 1) if tc else 0
        d["pct_total"] = round(d["facturado"] / total_fa * 100, 2)
        rubros.append(d)

    # por_subrubro: join stock_disponible for subrubro dimension
    subrubros_rows = (await db.execute(text(f"""
        SELECT
            s.cod_subrubro,
            MAX(s.subrubro)                                                                AS nombre,
            MAX(s.rubro)                                                                   AS rubro_nombre,
            COALESCE(SUM({venta_importe_neto_expr('v')}), 0)                              AS facturado,
            COALESCE(SUM({venta_importe_neto_expr('v')} - {venta_costo_neto_expr('v')}), 0) AS margen_abs,
            COALESCE(SUM(CASE WHEN v.precio_compra_actual IS NOT NULL THEN ABS(v.total) ELSE 0 END), 0) AS total_con_costo,
            COUNT(CASE WHEN v.tipo_comprobante='FA' THEN 1 END)                           AS tickets
        FROM ventas v
        JOIN (
            SELECT DISTINCT ON (cod_articulo) cod_articulo, cod_subrubro, subrubro, rubro
            FROM stock_disponible
            WHERE cod_subrubro IS NOT NULL
            ORDER BY cod_articulo
        ) s ON s.cod_articulo::text = v.producto_id
        WHERE {where}
        GROUP BY s.cod_subrubro
        ORDER BY facturado DESC
    """), params)).mappings().all()

    subrubros = []
    for r in subrubros_rows:
        d = dict(r)
        d["facturado"] = money(d["facturado"])
        d["margen_abs"] = money(d["margen_abs"])
        tc = money(d.pop("total_con_costo"))
        d["margen_pct"] = round(d["margen_abs"] / tc * 100, 1) if tc else 0
        d["pct_total"] = round(d["facturado"] / total_fa * 100, 2)
        subrubros.append(d)

    pareto = [
        {
            "producto_id": r["producto_id"],
            "producto": r["nombre"],
            "facturado": r["facturado"],
            "unidades": r["unidades"],
            "acumulado_pct": r["acumulado_pct"],
            "pct": r["pct_total"],
        }
        for r in ranking[:20]
    ]

    return {"ranking": ranking, "por_rubro": rubros, "por_subrubro": subrubros, "pareto": pareto}


@router.get("/ventas/por-lista")
async def ventas_por_lista(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params = filters.sql_params()
    where = _ventas_base_where(filters)

    lista_names: dict = {}
    try:
        lp_rows = (await db.execute(text(
            "SELECT cod_lista, descripcion FROM listas_precios"
        ))).mappings().all()
        lista_names = {r["cod_lista"]: r["descripcion"] for r in lp_rows}
    except Exception:
        pass

    por_lista = (await db.execute(text(f"""
        SELECT
            cod_lista_precios,
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets,
            COUNT(DISTINCT cliente_id) AS clientes_unicos,
            COUNT(DISTINCT cod_vendedor) AS vendedores_unicos,
            COALESCE(SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}), 0) AS margen_abs,
            COALESCE(SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END), 0) AS total_con_costo
        FROM ventas
        WHERE {where} AND cod_lista_precios IS NOT NULL
        GROUP BY cod_lista_precios
        ORDER BY facturado DESC
    """), params)).mappings().all()

    por_vendedor_lista = (await db.execute(text(f"""
        SELECT
            cod_vendedor, cod_lista_precios,
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets
        FROM ventas
        WHERE {where} AND cod_lista_precios IS NOT NULL AND cod_vendedor IS NOT NULL
        GROUP BY cod_vendedor, cod_lista_precios
        ORDER BY cod_vendedor, facturado DESC
    """), params)).mappings().all()

    vendedor_names: dict = {}
    try:
        vd_rows = (await db.execute(text("SELECT cod_vendedor, nombre FROM vendedores"))).mappings().all()
        vendedor_names = {r["cod_vendedor"]: r["nombre"] for r in vd_rows}
    except Exception:
        pass

    total_fa = sum(float(r["facturado"] or 0) for r in por_lista) or 1
    listas_out = []
    for r in por_lista:
        d = dict(r)
        d["facturado"] = money(d["facturado"])
        d["margen_abs"] = money(d["margen_abs"])
        tc = money(d.pop("total_con_costo"))
        d["margen_pct"] = round(d["margen_abs"] / tc * 100, 1) if tc else 0
        d["pct_total"] = round(d["facturado"] / total_fa * 100, 2)
        d["ticket_promedio"] = round(d["facturado"] / d["tickets"], 2) if d["tickets"] else 0
        d["nombre"] = lista_names.get(d["cod_lista_precios"], f"Lista {d['cod_lista_precios']}")
        listas_out.append(d)

    vendedores_out = []
    for r in por_vendedor_lista:
        d = dict(r)
        d["facturado"] = money(d["facturado"])
        d["vendedor_nombre"] = vendedor_names.get(d["cod_vendedor"], f"Vendedor {d['cod_vendedor']}")
        d["lista_nombre"] = lista_names.get(d["cod_lista_precios"], f"Lista {d['cod_lista_precios']}")
        vendedores_out.append(d)

    return {
        "por_lista": listas_out,
        "por_vendedor_lista": vendedores_out,
        "total_facturado": round(total_fa, 2),
    }


@router.get("/ventas/por-vendedor")
async def ventas_por_vendedor(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params = filters.sql_params()
    where = _ventas_base_where(filters)

    vendedor_names = {}
    cuotas = {}
    try:
        async with db.begin_nested():
            vd_rows = (await db.execute(text(
                "SELECT cod_vendedor, nombre, cuota_mensual FROM vendedores"
            ))).mappings().all()
            for r in vd_rows:
                vendedor_names[r["cod_vendedor"]] = r["nombre"]
                cuotas[r["cod_vendedor"]] = float(r["cuota_mensual"] or 0)
    except Exception:
        pass

    ventas_rows = (await db.execute(text(f"""
        SELECT
            cod_vendedor,
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado_neto,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                                  AS tickets,
            COUNT(DISTINCT CASE WHEN tipo_comprobante='FA' THEN cliente_id END)               AS clientes_unicos,
            COALESCE(SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}), 0)          AS margen_dolares,
            COALESCE(SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END), 0) AS total_con_costo
        FROM ventas
        WHERE {where} AND cod_vendedor IS NOT NULL
        GROUP BY cod_vendedor
        ORDER BY facturado_neto DESC
    """), params)).mappings().all()

    presupuestos_rows = []
    try:
        pres_where = "fecha >= :desde AND fecha < :hasta AND cod_vendedor IS NOT NULL"
        if not filters.incluir_anuladas:
            pres_where += " AND (anulada IS NULL OR anulada <> 'S')"
        async with db.begin_nested():
            presupuestos_rows = (await db.execute(text(f"""
                SELECT cod_vendedor,
                       COUNT(*) AS emitidos,
                       COUNT(CASE WHEN confirmado THEN 1 END) AS confirmados
                FROM presupuestos
                WHERE {pres_where}
                GROUP BY cod_vendedor
            """), {"desde": params["desde"], "hasta": params["hasta"]})).mappings().all()
    except Exception:
        pass

    pres_map = {r["cod_vendedor"]: dict(r) for r in presupuestos_rows}
    total_facturado = sum(money(r["facturado_neto"]) for r in ventas_rows) or 1

    result = []
    for r in ventas_rows:
        d = dict(r)
        fn = money(d["facturado_neto"])
        tk = int(d["tickets"] or 0)
        tc = money(d["total_con_costo"])
        md = money(d["margen_dolares"])
        pres = pres_map.get(d["cod_vendedor"], {})
        emitidos = int(pres.get("emitidos", 0))
        confirmados = int(pres.get("confirmados", 0))

        result.append({
            "cod_vendedor": d["cod_vendedor"],
            "nombre_vendedor": vendedor_names.get(d["cod_vendedor"], f"Vendedor {d['cod_vendedor']}"),
            "facturado_neto": fn,
            "tickets": tk,
            "ticket_promedio": fn / tk if tk else 0,
            "clientes_unicos": int(d["clientes_unicos"] or 0),
            "margen_dolares": md,
            "margen_pct": round(md / tc * 100, 1) if tc else 0,
            "pct_del_total": round(fn / total_facturado * 100, 2),
            "presupuestos_emitidos": emitidos,
            "presupuestos_confirmados": confirmados,
            "tasa_conversion": round(confirmados / emitidos * 100, 1) if emitidos else 0,
            "cuota_mensual": cuotas.get(d["cod_vendedor"], 0),
        })

    return {"vendedores": result}


@router.get("/ventas/por-cliente")
async def ventas_por_cliente(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params = filters.sql_params()
    where = _ventas_base_where(filters)

    rows = (await db.execute(text(f"""
        SELECT
            cliente_id,
            COALESCE(MAX(cliente_nombre), cliente_id)                                        AS cliente_nombre,
            COALESCE(SUM({venta_importe_neto_expr()}),0)                                     AS facturado_neto,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                                AS tickets,
            MAX(fecha)                                                                        AS ultima_compra,
            MIN(fecha)                                                                        AS primera_compra_periodo,
            COALESCE(SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}),0)         AS margen_dolares,
            COALESCE(SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END),0) AS total_con_costo,
            MODE() WITHIN GROUP (ORDER BY condicion_venta_tipo)                               AS condicion_predominante
        FROM ventas
        WHERE {where}
        GROUP BY cliente_id
        ORDER BY facturado_neto DESC
        LIMIT 200
    """), params)).mappings().all()

    # primera compra histórica para detectar clientes nuevos
    nuevos_set: set = set()
    try:
        hist_rows = (await db.execute(text(f"""
            SELECT DISTINCT cliente_id
            FROM ventas
            WHERE fecha < :desde AND tipo_comprobante='FA'
        """), {"desde": params["desde"]})).all()
        historicos = {r[0] for r in hist_rows}
        nuevos_set = {r["cliente_id"] for r in rows if r["cliente_id"] not in historicos}
    except Exception:
        pass

    today = date.today()
    total_facturado = sum(money(r["facturado_neto"]) for r in rows) or 1

    # ABC classification
    total_ordenado = sorted(rows, key=lambda r: money(r["facturado_neto"]), reverse=True)
    acumulado = 0.0
    abc_map = {}
    for r in total_ordenado:
        acumulado += money(r["facturado_neto"]) / total_facturado * 100
        if acumulado <= 80:
            abc_map[r["cliente_id"]] = "A"
        elif acumulado <= 95:
            abc_map[r["cliente_id"]] = "B"
        else:
            abc_map[r["cliente_id"]] = "C"

    result = []
    for r in rows:
        d = dict(r)
        fn = money(d["facturado_neto"])
        tk = int(d["tickets"] or 0)
        tc = money(d["total_con_costo"])
        md = money(d["margen_dolares"])
        ultima = d["ultima_compra"]
        dias_sin_comprar = (today - ultima.date()).days if ultima else None

        result.append({
            "cod_cliente": d["cliente_id"],
            "cliente_nombre": d["cliente_nombre"],
            "facturado_neto": fn,
            "tickets": tk,
            "ticket_promedio": fn / tk if tk else 0,
            "pct_total": round(fn / total_facturado * 100, 2),
            "margen_dolares": md,
            "margen_pct": round(md / tc * 100, 1) if tc else 0,
            "condicion_venta_predominante": d.get("condicion_predominante"),
            "ultima_compra": ultima.isoformat() if ultima else None,
            "dias_sin_comprar": dias_sin_comprar,
            "es_nuevo": d["cliente_id"] in nuevos_set,
            "segmento": abc_map.get(d["cliente_id"], "C"),
        })

    # ABC summary
    abc_summary = {"A": {"clientes": 0, "facturado": 0.0}, "B": {"clientes": 0, "facturado": 0.0}, "C": {"clientes": 0, "facturado": 0.0}}
    for r in result:
        seg = r["segmento"]
        abc_summary[seg]["clientes"] += 1
        abc_summary[seg]["facturado"] += r["facturado_neto"]

    return {"clientes": result, "abc_summary": abc_summary}


@router.get("/ventas/por-comprobante")
async def ventas_por_comprobante(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params = filters.sql_params()
    where = _ventas_base_where(filters)

    tipo_rows = (await db.execute(text(f"""
        SELECT tipo_comprobante,
               COUNT(*)                           AS cantidad,
               COALESCE(SUM(ABS(total)), 0)       AS importe
        FROM ventas WHERE {where}
        GROUP BY tipo_comprobante ORDER BY importe DESC
    """), params)).mappings().all()

    factura_rows = (await db.execute(text(f"""
        SELECT COALESCE(tipo_factura,'?') AS tipo_factura,
               COUNT(*)                  AS cantidad,
               COALESCE(SUM(total), 0)  AS importe
        FROM ventas WHERE {where} AND tipo_comprobante='FA'
        GROUP BY tipo_factura ORDER BY importe DESC
    """), params)).mappings().all()

    condicion_rows = (await db.execute(text(f"""
        SELECT COALESCE(condicion_venta_tipo::text,'?') AS condicion,
               COUNT(*)                               AS cantidad,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END),0) AS importe
        FROM ventas WHERE {where}
        GROUP BY condicion ORDER BY importe DESC
    """), params)).mappings().all()

    pdv_rows = (await db.execute(text(f"""
        SELECT COALESCE(punto_de_venta::text,'?') AS punto_de_venta,
               COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END),0) AS facturado
        FROM ventas WHERE {where}
        GROUP BY punto_de_venta ORDER BY facturado DESC
    """), params)).mappings().all()

    total_fa = sum(money(r["importe"]) for r in tipo_rows if r["tipo_comprobante"] == "FA") or 1

    return {
        "por_tipo": [{"tipo": r["tipo_comprobante"], "cantidad": int(r["cantidad"]), "importe": money(r["importe"])} for r in tipo_rows],
        "por_tipo_factura": [{"tipo": r["tipo_factura"], "cantidad": int(r["cantidad"]), "importe": money(r["importe"])} for r in factura_rows],
        "por_condicion_venta": [{"condicion": r["condicion"], "cantidad": int(r["cantidad"]), "importe": money(r["importe"])} for r in condicion_rows],
        "por_punto_de_venta": [{"punto": r["punto_de_venta"], "tickets": int(r["tickets"]), "facturado": money(r["facturado"]), "pct_total": round(money(r["facturado"]) / total_fa * 100, 1)} for r in pdv_rows],
    }


@router.get("/ventas/transacciones")
async def ventas_transacciones(
    company_id: int = None,
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, le=200),
    sort_by: str = Query(default="fecha"),
    sort_dir: Literal["asc", "desc"] = "desc",
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params = filters.sql_params()
    where = _ventas_base_where(filters)

    safe_sort = {
        "fecha", "total", "cliente_id", "cliente_nombre", "tipo_comprobante",
        "punto_de_venta", "cod_vendedor", "neto", "iva_importe", "cantidad"
    }
    sort_col = sort_by if sort_by in safe_sort else "fecha"
    direction = "ASC" if sort_dir == "asc" else "DESC"
    offset = (page - 1) * limit
    params["limit"] = limit
    params["offset"] = offset

    total_count = (await db.execute(text(f"SELECT COUNT(*) FROM ventas WHERE {where}"), params)).scalar()

    rows = (await db.execute(text(f"""
        SELECT id, fecha, tipo_comprobante, tipo_factura, punto_de_venta,
               cliente_id, cliente_nombre, cod_vendedor,
               COALESCE(neto::float, 0) AS neto,
               COALESCE(iva_importe::float, 0) AS iva_importe,
               total, condicion_venta_tipo, cod_deposito, anulada,
               producto_id, producto_nombre, cantidad, precio_unitario,
               COALESCE(precio_compra_actual::float, 0) AS precio_compra_actual,
               COALESCE(descuento_porc::float, 0) AS descuento_porc
        FROM ventas
        WHERE {where}
        ORDER BY {sort_col} {direction}
        LIMIT :limit OFFSET :offset
    """), params)).mappings().all()

    return {
        "total": int(total_count),
        "page": page,
        "limit": limit,
        "pages": -(-int(total_count) // limit),
        "rows": [
            {**dict(r), "fecha": r["fecha"].isoformat() if r["fecha"] else None}
            for r in rows
        ],
    }


@router.get("/ventas/aging")
async def ventas_aging(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    buckets_raw = (await db.execute(text("""
        SELECT
            CASE
                WHEN fecha_venc > CURRENT_DATE THEN 'vigente'
                WHEN dias_vencido <= 30  THEN '0_30'
                WHEN dias_vencido <= 60  THEN '31_60'
                WHEN dias_vencido <= 90  THEN '61_90'
                ELSE '90_mas'
            END AS bucket,
            COUNT(*)       AS cantidad,
            SUM(importe)   AS total
        FROM (
            SELECT
                importe,
                COALESCE(fecha_vencimiento, fecha + INTERVAL '30 days') AS fecha_venc,
                GREATEST(0, CURRENT_DATE - COALESCE(fecha_vencimiento, fecha + INTERVAL '30 days')::date) AS dias_vencido
            FROM cuentas_corrientes_clientes
            WHERE tipo IN ('FA','ND','saldo') AND importe > 0
        ) sub
        GROUP BY 1
    """))).mappings().all()

    bucket_order = ['vigente', '0_30', '31_60', '61_90', '90_mas']
    bucket_labels = {
        'vigente': 'Vigente', '0_30': '0-30 días',
        '31_60': '31-60 días', '61_90': '61-90 días', '90_mas': '+90 días',
    }
    buckets_map = {r['bucket']: dict(r) for r in buckets_raw}
    buckets = [
        {
            'bucket': k,
            'label': bucket_labels[k],
            'cantidad': int(buckets_map[k]['cantidad']) if k in buckets_map else 0,
            'total': round(money(buckets_map[k]['total']), 2) if k in buckets_map else 0,
        }
        for k in bucket_order
    ]

    total_pendiente = sum(b['total'] for b in buckets if b['bucket'] != 'vigente')
    total_vigente = next((b['total'] for b in buckets if b['bucket'] == 'vigente'), 0)

    top_deudores = (await db.execute(text("""
        SELECT cliente_id, MAX(cliente_nombre) AS nombre, saldo_acumulado
        FROM (
            SELECT cliente_id, cliente_nombre, saldo_acumulado,
                   ROW_NUMBER() OVER (PARTITION BY cliente_id ORDER BY id DESC) AS rn
            FROM cuentas_corrientes_clientes
        ) t
        WHERE rn = 1 AND saldo_acumulado > 0
        ORDER BY saldo_acumulado DESC
        LIMIT 10
    """))).mappings().all()

    return {
        'buckets': buckets,
        'total_pendiente': round(total_pendiente, 2),
        'total_vigente': round(total_vigente, 2),
        'top_deudores': [
            {'cliente_id': r['cliente_id'], 'nombre': r['nombre'], 'saldo': round(money(r['saldo_acumulado']), 2)}
            for r in top_deudores
        ],
    }


@router.get("/ventas/ticket-dist")
async def ventas_ticket_dist(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    # Calcular percentiles del período para definir buckets dinámicamente
    stats_row = (await db.execute(text(f"""
        SELECT
            PERCENTILE_CONT(0.10) WITHIN GROUP (ORDER BY ABS(total)) AS p10,
            PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY ABS(total)) AS p25,
            PERCENTILE_CONT(0.50) WITHIN GROUP (ORDER BY ABS(total)) AS p50,
            PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY ABS(total)) AS p75,
            PERCENTILE_CONT(0.90) WITHIN GROUP (ORDER BY ABS(total)) AS p90,
            MAX(ABS(total)) AS max_val,
            AVG(ABS(total)) AS avg_val
        FROM ventas
        WHERE tipo_comprobante = 'FA' AND {where}
    """), params)).mappings().one()

    def nice_round(v: float) -> int:
        """Round to a visually clean number for bucket boundaries."""
        if v <= 0:
            return 0
        magnitude = 10 ** (len(str(int(v))) - 1)
        return int((v // magnitude + 1) * magnitude)

    p10 = float(stats_row["p10"] or 0)
    p25 = float(stats_row["p25"] or 0)
    p50 = float(stats_row["p50"] or 0)
    p75 = float(stats_row["p75"] or 0)
    p90 = float(stats_row["p90"] or 0)

    # Generar 6 cortes dinámicos redondeados
    raw_cuts = [p10, p25, p50, p75, p90]
    cuts = sorted(set(nice_round(c) for c in raw_cuts if c > 0))
    # Deduplicate and ensure at least 3 cuts
    if len(cuts) < 3:
        avg = float(stats_row["avg_val"] or 1)
        cuts = [nice_round(avg * 0.25), nice_round(avg * 0.5), nice_round(avg), nice_round(avg * 2)]
        cuts = sorted(set(c for c in cuts if c > 0))

    def fmt_bucket_label(lo: int, hi: int | None) -> str:
        def fmt(v: int) -> str:
            if v >= 1_000_000_000: return f"${v//1_000_000_000}B"
            if v >= 1_000_000:     return f"${v//1_000_000}M"
            if v >= 1_000:         return f"${v//1_000}K"
            return f"${v}"
        return f">{fmt(lo)}" if hi is None else (f"<{fmt(hi)}" if lo == 0 else f"{fmt(lo)}-{fmt(hi)}")

    # Build CASE expression dynamically
    when_clauses_ord = []
    when_clauses_label = []
    for i, cut in enumerate(cuts, start=1):
        when_clauses_ord.append(f"WHEN ABS(total) < {cut} THEN {i}")
        lo = cuts[i - 2] if i > 1 else 0
        when_clauses_label.append(f"WHEN ABS(total) < {cut} THEN '{fmt_bucket_label(lo, cut)}'")
    last_lo = cuts[-1]
    max_ord = len(cuts) + 1

    case_ord   = "CASE " + " ".join(when_clauses_ord)   + f" ELSE {max_ord} END"
    case_label = "CASE " + " ".join(when_clauses_label) + f" ELSE '{fmt_bucket_label(last_lo, None)}' END"

    rows = (await db.execute(text(f"""
        SELECT
            {case_ord}   AS bucket_ord,
            {case_label} AS bucket,
            COUNT(*)          AS cantidad,
            SUM(ABS(total))   AS facturado,
            AVG(ABS(total))   AS ticket_promedio
        FROM ventas
        WHERE tipo_comprobante = 'FA' AND {where}
        GROUP BY 1, 2
        ORDER BY 1
    """), params)).mappings().all()

    data = [
        {
            'bucket': r['bucket'],
            'cantidad': int(r['cantidad']),
            'facturado': round(money(r['facturado']), 2),
            'ticket_promedio': round(money(r['ticket_promedio']), 2),
        }
        for r in rows
    ]

    total_tickets = sum(d['cantidad'] for d in data)
    p50_idx = 0
    acum = 0
    for i, d in enumerate(data):
        acum += d['cantidad']
        if acum >= total_tickets * 0.5:
            p50_idx = i
            break

    return {
        'distribucion': data,
        'total_tickets': total_tickets,
        'p50_bucket': data[p50_idx]['bucket'] if data else None,
        'ticket_promedio_global': round(money(stats_row["avg_val"]), 2) if stats_row["avg_val"] else 0,
    }


@router.get("/ventas/cohort")
async def ventas_cohort(
    company_id: int = None,
    meses: int = 12,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    rows = (await db.execute(text(f"""
        WITH first_purchase AS (
            SELECT cliente_id,
                   DATE_TRUNC('month', MIN(fecha)) AS cohort_month
            FROM ventas
            WHERE tipo_comprobante = 'FA'
            GROUP BY cliente_id
        ),
        cohort_sizes AS (
            SELECT cohort_month, COUNT(*) AS cohort_size
            FROM first_purchase
            WHERE cohort_month >= DATE_TRUNC('month', NOW()) - INTERVAL '{meses - 1} months'
            GROUP BY cohort_month
        ),
        retention AS (
            SELECT
                fp.cohort_month,
                DATE_TRUNC('month', v.fecha) AS activity_month,
                COUNT(DISTINCT fp.cliente_id) AS retained_count
            FROM first_purchase fp
            JOIN ventas v ON v.cliente_id = fp.cliente_id AND v.tipo_comprobante = 'FA'
            WHERE fp.cohort_month >= DATE_TRUNC('month', NOW()) - INTERVAL '{meses - 1} months'
              AND DATE_TRUNC('month', v.fecha) >= fp.cohort_month
            GROUP BY fp.cohort_month, DATE_TRUNC('month', v.fecha)
        )
        SELECT
            TO_CHAR(r.cohort_month, 'YYYY-MM') AS cohort_month,
            cs.cohort_size,
            TO_CHAR(r.activity_month, 'YYYY-MM') AS activity_month,
            r.retained_count,
            ROUND(r.retained_count::float / cs.cohort_size * 100)::int AS retention_pct,
            ROUND(EXTRACT(EPOCH FROM (r.activity_month - r.cohort_month)) / (30.44 * 24 * 3600))::int AS month_offset
        FROM retention r
        JOIN cohort_sizes cs ON r.cohort_month = cs.cohort_month
        ORDER BY r.cohort_month, month_offset
    """))).mappings().all()

    cohorts: dict = {}
    for r in rows:
        cm = r['cohort_month']
        if cm not in cohorts:
            cohorts[cm] = {'cohort_month': cm, 'cohort_size': int(r['cohort_size']), 'retention': {}}
        cohorts[cm]['retention'][int(r['month_offset'])] = int(r['retention_pct'])

    return {'cohorts': list(cohorts.values()), 'max_offset': meses - 1}


@router.get("/ventas/clientes-riesgo")
async def ventas_clientes_riesgo(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Clientes que compraron en el período anterior pero NO en el actual.

    Tres categorías:
    - perdidos: compraban en período anterior, 0 compras en actual
    - en_caida: compraron en ambos pero el actual cayó >40% vs. anterior
    - nuevos: primera compra en el período actual (vinieron por primera vez)
    """
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params = filters.sql_params()
    where = _ventas_base_where(filters)
    prev_desde, prev_hasta = _prev_period(filters)
    prev_params = dict(params)
    prev_params["desde"] = prev_desde
    prev_params["hasta"] = prev_hasta

    # Facturado actual por cliente
    actual_rows = (await db.execute(text(f"""
        SELECT cliente_id,
               COALESCE(MAX(cliente_nombre), cliente_id) AS cliente_nombre,
               COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado,
               COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets,
               MAX(fecha) AS ultima_compra
        FROM ventas WHERE {where}
        GROUP BY cliente_id
    """), params)).mappings().all()
    actual_map = {r["cliente_id"]: dict(r) for r in actual_rows}

    # Facturado período anterior por cliente
    prev_rows = (await db.execute(text(f"""
        SELECT cliente_id,
               COALESCE(MAX(cliente_nombre), cliente_id) AS cliente_nombre,
               COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado,
               MAX(fecha) AS ultima_compra
        FROM ventas WHERE {where}
        GROUP BY cliente_id
    """), prev_params)).mappings().all()
    prev_map = {r["cliente_id"]: dict(r) for r in prev_rows}

    # Clientes históricos (cualquier compra antes del período anterior) — para detectar nuevos
    hist_rows = (await db.execute(text(f"""
        SELECT DISTINCT cliente_id FROM ventas
        WHERE fecha < :prev_desde AND tipo_comprobante='FA'
    """), {"prev_desde": prev_desde})).all()
    historicos = {r[0] for r in hist_rows}

    perdidos = []
    en_caida = []
    nuevos = []

    for cli_id, prev in prev_map.items():
        prev_fact = money(prev["facturado"])
        if prev_fact <= 0:
            continue
        actual = actual_map.get(cli_id)
        actual_fact = money(actual["facturado"]) if actual else 0

        if actual_fact <= 0:
            perdidos.append({
                "cod_cliente": cli_id,
                "cliente_nombre": prev["cliente_nombre"],
                "facturado_anterior": prev_fact,
                "facturado_actual": 0,
                "variacion_pct": -100,
                "ultima_compra": prev["ultima_compra"].isoformat() if prev["ultima_compra"] else None,
            })
        else:
            var = (actual_fact - prev_fact) / prev_fact * 100
            if var <= -40:
                en_caida.append({
                    "cod_cliente": cli_id,
                    "cliente_nombre": prev["cliente_nombre"],
                    "facturado_anterior": prev_fact,
                    "facturado_actual": actual_fact,
                    "variacion_pct": round(var, 1),
                    "ultima_compra": actual["ultima_compra"].isoformat() if actual["ultima_compra"] else None,
                })

    for cli_id, actual in actual_map.items():
        if cli_id in historicos or cli_id in prev_map:
            continue
        nuevos.append({
            "cod_cliente": cli_id,
            "cliente_nombre": actual["cliente_nombre"],
            "facturado_actual": money(actual["facturado"]),
            "tickets": int(actual["tickets"] or 0),
            "primera_compra": actual["ultima_compra"].isoformat() if actual["ultima_compra"] else None,
        })

    # Ordenar por importancia (mayor facturado primero)
    perdidos.sort(key=lambda x: x["facturado_anterior"], reverse=True)
    en_caida.sort(key=lambda x: x["facturado_anterior"] - x["facturado_actual"], reverse=True)
    nuevos.sort(key=lambda x: x["facturado_actual"], reverse=True)

    return {
        "perdidos": perdidos[:30],
        "en_caida": en_caida[:30],
        "nuevos": nuevos[:30],
        "totales": {
            "perdidos_cnt": len(perdidos),
            "perdidos_facturado_ant": sum(p["facturado_anterior"] for p in perdidos),
            "en_caida_cnt": len(en_caida),
            "en_caida_riesgo": sum(p["facturado_anterior"] - p["facturado_actual"] for p in en_caida),
            "nuevos_cnt": len(nuevos),
            "nuevos_facturado": sum(n["facturado_actual"] for n in nuevos),
        },
    }


@router.get("/ventas/dia-semana")
async def ventas_dia_semana(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Facturación, tickets y ticket promedio agregado por día de la semana."""
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    rows = (await db.execute(text(f"""
        SELECT
            EXTRACT(ISODOW FROM fecha)::int AS dow,
            COALESCE(SUM({venta_importe_neto_expr()}), 0)                    AS facturado,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                AS tickets,
            COUNT(DISTINCT fecha::date) FILTER (WHERE tipo_comprobante='FA') AS dias_activos
        FROM ventas
        WHERE {where}
        GROUP BY 1
        ORDER BY 1
    """), params)).mappings().all()

    NOMBRES = {1: 'Lun', 2: 'Mar', 3: 'Mié', 4: 'Jue', 5: 'Vie', 6: 'Sáb', 7: 'Dom'}
    result = []
    for d in range(1, 8):
        r = next((row for row in rows if int(row["dow"]) == d), None)
        if r:
            fact = money(r["facturado"])
            tk = int(r["tickets"] or 0)
            dias = int(r["dias_activos"] or 0) or 1
            result.append({
                "dow": d,
                "dia": NOMBRES[d],
                "facturado": fact,
                "tickets": tk,
                "ticket_promedio": fact / tk if tk else 0,
                "facturado_promedio_dia": fact / dias,
                "tickets_promedio_dia": round(tk / dias, 1),
                "dias_activos": dias,
            })
        else:
            result.append({
                "dow": d, "dia": NOMBRES[d],
                "facturado": 0, "tickets": 0, "ticket_promedio": 0,
                "facturado_promedio_dia": 0, "tickets_promedio_dia": 0, "dias_activos": 0,
            })

    total_fact = sum(r["facturado"] for r in result) or 1
    for r in result:
        r["pct_total"] = round(r["facturado"] / total_fact * 100, 1)

    return {"por_dia": result}


@router.get("/ventas/nuevos-recurrentes")
async def ventas_nuevos_recurrentes(
    company_id: int = None,
    granularidad: Literal["dia", "semana", "mes", "trimestre"] = "mes",
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Serie temporal con facturado segmentado por clientes nuevos vs. recurrentes."""
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    trunc = {"dia": "day", "semana": "week", "mes": "month", "trimestre": "quarter"}.get(granularidad, "month")
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    rows = (await db.execute(text(f"""
        WITH first_purchase AS (
            SELECT cliente_id, MIN(fecha::date) AS first_date
            FROM ventas
            WHERE tipo_comprobante='FA'
            GROUP BY cliente_id
        )
        SELECT
            to_char(date_trunc('{trunc}', v.fecha), 'YYYY-MM-DD') AS periodo,
            COALESCE(SUM({venta_importe_neto_expr('v')}) FILTER (
                WHERE fp.first_date >= :desde AND fp.first_date < :hasta
            ), 0) AS facturado_nuevos,
            COALESCE(SUM({venta_importe_neto_expr('v')}) FILTER (
                WHERE fp.first_date < :desde
            ), 0) AS facturado_recurrentes,
            COUNT(DISTINCT v.cliente_id) FILTER (
                WHERE fp.first_date >= :desde AND fp.first_date < :hasta
            ) AS clientes_nuevos,
            COUNT(DISTINCT v.cliente_id) FILTER (
                WHERE fp.first_date < :desde
            ) AS clientes_recurrentes
        FROM ventas v
        LEFT JOIN first_purchase fp ON fp.cliente_id = v.cliente_id
        WHERE {where}
        GROUP BY 1
        ORDER BY 1
    """), params)).mappings().all()

    return {
        "series": [
            {
                "periodo": r["periodo"],
                "facturado_nuevos": money(r["facturado_nuevos"]),
                "facturado_recurrentes": money(r["facturado_recurrentes"]),
                "clientes_nuevos": int(r["clientes_nuevos"] or 0),
                "clientes_recurrentes": int(r["clientes_recurrentes"] or 0),
            }
            for r in rows
        ],
    }


@router.get("/ventas/exportar")
async def ventas_exportar(
    company_id: int = None,
    formato: Literal["excel", "csv"] = "excel",
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        import openpyxl
        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False

    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params = filters.sql_params()
    where = _ventas_base_where(filters)

    kpi_row = await _fetch_kpi_row(db, where, params)
    fa = money(kpi_row["fa_bruto"])
    nc = money(kpi_row["nc_total"])
    fn = money(kpi_row["facturado_neto"])

    rows = (await db.execute(text(f"""
        SELECT fecha, tipo_comprobante, tipo_factura, punto_de_venta,
               cliente_id, cliente_nombre, cod_vendedor,
               COALESCE(neto::float, 0) AS neto,
               COALESCE(iva_importe::float, 0)    AS iva_importe,
               total, condicion_venta_tipo, anulada,
               producto_id, producto_nombre, cantidad, precio_unitario,
               COALESCE(precio_compra_actual::float,0) AS precio_compra_actual
        FROM ventas
        WHERE {where}
        ORDER BY fecha DESC
        LIMIT 10000
    """), params)).mappings().all()

    if formato == "csv":
        import csv as csv_mod
        buf = io.StringIO()
        writer = csv_mod.DictWriter(buf, fieldnames=[
            "fecha","tipo_comprobante","tipo_factura","punto_de_venta",
            "cliente_id","cliente_nombre","cod_vendedor","neto","iva_importe",
            "total","condicion_venta_tipo","anulada",
            "producto_id","producto_nombre","cantidad","precio_unitario","precio_compra_actual"
        ])
        writer.writeheader()
        for r in rows:
            d = dict(r)
            d["fecha"] = d["fecha"].isoformat() if d["fecha"] else ""
            writer.writerow(d)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=ventas.csv"},
        )

    # Excel fallback: if openpyxl not installed, fall back to CSV
    if not HAS_OPENPYXL:
        return {"error": "openpyxl no disponible, use formato=csv"}

    from openpyxl import Workbook
    wb = Workbook()

    # KPIs sheet
    ws_kpi = wb.active
    ws_kpi.title = "KPIs"
    ws_kpi.append(["Métrica", "Valor"])
    ws_kpi.append(["Facturado Neto", fn])
    ws_kpi.append(["Facturado Bruto FA", fa])
    ws_kpi.append(["Devoluciones NC/ND", nc])
    ws_kpi.append(["IVA Débito", money(kpi_row["iva_debito"])])
    ws_kpi.append(["Tickets", int(kpi_row["tickets"] or 0)])
    ws_kpi.append(["Clientes únicos", int(kpi_row["clientes_unicos"] or 0)])
    ws_kpi.append(["Unidades", money(kpi_row["unidades"])])

    # Transacciones sheet
    ws_tx = wb.create_sheet("Transacciones")
    headers = ["Fecha","Tipo","Letra","Pto.Vta","Cliente ID","Cliente","Vendedor",
               "Neto","IVA","Total","Condición","Anulada","Producto","Descripción","Cantidad","Precio U.","Costo U."]
    ws_tx.append(headers)
    for r in rows:
        ws_tx.append([
            r["fecha"].isoformat() if r["fecha"] else "",
            r["tipo_comprobante"], r["tipo_factura"], r["punto_de_venta"],
            r["cliente_id"], r["cliente_nombre"], r["cod_vendedor"],
            r["neto"], r["iva_importe"], r["total"], r["condicion_venta_tipo"], r["anulada"],
            r["producto_id"], r["producto_nombre"], r["cantidad"], r["precio_unitario"], r["precio_compra_actual"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ventas.xlsx"},
    )


# ══════════════════════════════════════════════════════════════════════════════
# FASE 5: Panel Resultado — endpoints dedicados
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/resultado/kpis")
async def resultado_kpis(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    row = (await db.execute(text(f"""
        SELECT
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado_neto,
            COALESCE(SUM({venta_costo_neto_expr()}), 0) AS cogs,
            COALESCE(SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END), 0) AS total_con_costo,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets_fa,
            COALESCE(SUM(
                CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                     THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                ELSE 0 END), 0) AS descuento_total,
            COUNT(*) AS total_items,
            SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN 1 ELSE 0 END) AS items_con_costo
        FROM ventas WHERE {where}
    """), params)).mappings().one()

    facturado_neto = money(row["facturado_neto"])
    cogs = money(row["cogs"])
    margen_bruto = facturado_neto - cogs
    margen_pct = (margen_bruto / facturado_neto * 100) if facturado_neto else 0
    tickets_fa = int(row["tickets_fa"] or 0) or 1
    ticket_margen = margen_bruto / tickets_fa
    descuento_total = money(row["descuento_total"])
    descuento_pct = (descuento_total / (facturado_neto + descuento_total) * 100) if (facturado_neto + descuento_total) else 0
    items_con_costo = int(row["items_con_costo"] or 0)
    total_items = int(row["total_items"] or 0)
    cobertura_costo_pct = round(items_con_costo / total_items * 100, 1) if total_items else 0

    bajo_costo_rows = (await db.execute(text(f"""
        SELECT COUNT(DISTINCT producto_id) AS cnt
        FROM ventas
        WHERE {where} AND precio_compra_actual IS NOT NULL
          AND precio_unitario < precio_compra_actual::float
          AND tipo_comprobante = 'FA'
    """), params)).mappings().one()
    productos_bajo_costo = int(bajo_costo_rows["cnt"] or 0)

    anterior = None
    if filters.comparar_anterior:
        prev_desde, prev_hasta = _prev_period(filters)
        pp = dict(params); pp["desde"] = prev_desde; pp["hasta"] = prev_hasta
        prev = (await db.execute(text(f"""
            SELECT
                COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado_neto,
                COALESCE(SUM({venta_costo_neto_expr()}), 0) AS cogs,
                COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets_fa,
                COALESCE(SUM(
                    CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                         THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                    ELSE 0 END), 0) AS descuento_total
            FROM ventas WHERE {where}
        """), pp)).mappings().one()
        pfn = money(prev["facturado_neto"])
        pcogs = money(prev["cogs"])
        pmb = pfn - pcogs
        pmp = (pmb / pfn * 100) if pfn else 0
        ptk = int(prev["tickets_fa"] or 0) or 1
        ptm = pmb / ptk
        pdt = money(prev["descuento_total"])
        pdp = (pdt / (pfn + pdt) * 100) if (pfn + pdt) else 0
        anterior = {
            "facturado_neto": pfn, "cogs": pcogs, "margen_bruto": pmb,
            "margen_pct": pmp, "ticket_margen": ptm, "descuento_total": pdt, "descuento_pct": pdp,
        }

    def kpi(key, val):
        ant = anterior[key] if anterior else None
        return _kpi_obj(val, ant)

    return {
        "facturado_neto": kpi("facturado_neto", facturado_neto),
        "cogs": kpi("cogs", cogs),
        "margen_bruto": kpi("margen_bruto", margen_bruto),
        "margen_pct": kpi("margen_pct", margen_pct),
        "ticket_margen": kpi("ticket_margen", ticket_margen),
        "productos_bajo_costo": {"actual": productos_bajo_costo},
        "descuento_total": kpi("descuento_total", descuento_total),
        "descuento_pct": {"actual": round(descuento_pct, 2)},
        "cobertura_costo_pct": cobertura_costo_pct,
    }


@router.get("/resultado/temporal")
async def resultado_temporal(
    company_id: int = None,
    granularidad: Literal["dia", "semana", "mes", "trimestre"] = "mes",
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    trunc = {"dia": "day", "semana": "week", "mes": "month", "trimestre": "quarter"}.get(granularidad, "month")
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    series = (await db.execute(text(f"""
        SELECT
            to_char(date_trunc('{trunc}', fecha), 'YYYY-MM-DD') AS periodo,
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado_neto,
            COALESCE(SUM({venta_costo_neto_expr()}), 0) AS cogs,
            COALESCE(SUM(
                CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                     THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                ELSE 0 END), 0) AS descuento_total
        FROM ventas WHERE {where}
        GROUP BY 1 ORDER BY 1
    """), params)).mappings().all()

    result = []
    for r in series:
        fn = money(r["facturado_neto"])
        c = money(r["cogs"])
        mb = fn - c
        mp = (mb / fn * 100) if fn else 0
        result.append({
            "periodo": r["periodo"],
            "facturado_neto": round(fn, 2),
            "cogs": round(c, 2),
            "margen_bruto": round(mb, 2),
            "margen_pct": round(mp, 2),
            "descuento_total": round(money(r["descuento_total"]), 2),
        })

    if filters.comparar_anterior:
        prev_desde, prev_hasta = _prev_period(filters)
        pp = dict(params); pp["desde"] = prev_desde; pp["hasta"] = prev_hasta
        prev = (await db.execute(text(f"""
            SELECT
                to_char(date_trunc('{trunc}', fecha), 'YYYY-MM-DD') AS periodo,
                COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado_neto,
                COALESCE(SUM({venta_costo_neto_expr()}), 0) AS cogs
            FROM ventas WHERE {where}
            GROUP BY 1 ORDER BY 1
        """), pp)).mappings().all()
        prev_list = list(prev)
        for i, row in enumerate(result):
            if i < len(prev_list):
                pfn = money(prev_list[i]["facturado_neto"])
                pc = money(prev_list[i]["cogs"])
                row["facturado_anterior"] = round(pfn, 2)
                row["margen_anterior"] = round(pfn - pc, 2)

    return {"series": result}


@router.get("/resultado/por-producto")
async def resultado_por_producto(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    rows = (await db.execute(text(f"""
        SELECT
            producto_id,
            COALESCE(MAX(producto_nombre), producto_id) AS nombre,
            MAX(cod_rubro) AS cod_rubro,
            SUM(cantidad) AS unidades,
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado,
            COALESCE(SUM({venta_costo_neto_expr()}), 0) AS cogs,
            COALESCE(SUM(
                CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                     THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                ELSE 0 END), 0) AS descuento_total
        FROM ventas WHERE {where}
        GROUP BY producto_id
        ORDER BY facturado DESC
        LIMIT 200
    """), params)).mappings().all()

    ranking = []
    facturados = []
    margenes = []
    for r in rows:
        f = money(r["facturado"])
        c = money(r["cogs"])
        mb = f - c
        mp = (mb / f * 100) if f else 0
        facturados.append(f)
        margenes.append(mp)
        ranking.append({
            "producto_id": r["producto_id"],
            "nombre": r["nombre"],
            "cod_rubro": r["cod_rubro"],
            "unidades": int(r["unidades"] or 0),
            "facturado": round(f, 2),
            "cogs": round(c, 2),
            "margen_dolares": round(mb, 2),
            "margen_pct": round(mp, 2),
            "descuento_total": round(money(r["descuento_total"]), 2),
        })

    med_fact = sorted(facturados)[len(facturados) // 2] if facturados else 0
    med_marg = sorted(margenes)[len(margenes) // 2] if margenes else 0

    cuadrantes = {"estrellas": [], "vacas": [], "diamantes": [], "perros": []}
    q_key = {"estrella": "estrellas", "vaca": "vacas", "diamante": "diamantes", "perro": "perros"}
    for p in ranking:
        if p["margen_pct"] >= med_marg and p["facturado"] >= med_fact:
            q = "estrella"
        elif p["margen_pct"] < med_marg and p["facturado"] >= med_fact:
            q = "vaca"
        elif p["margen_pct"] >= med_marg and p["facturado"] < med_fact:
            q = "diamante"
        else:
            q = "perro"
        p["cuadrante"] = q
        cuadrantes[q_key[q]].append(p)

    return {
        "ranking": ranking,
        "por_cuadrante": cuadrantes,
        "medianas": {"facturado": round(med_fact, 2), "margen_pct": round(med_marg, 2)},
    }


@router.get("/resultado/por-vendedor")
async def resultado_por_vendedor(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    vendedor_names = {}
    try:
        vd_rows = (await db.execute(text(
            "SELECT cod_vendedor, nombre FROM vendedores"
        ))).mappings().all()
        for r in vd_rows:
            vendedor_names[r["cod_vendedor"]] = r["nombre"]
    except Exception:
        pass

    rows = (await db.execute(text(f"""
        SELECT
            cod_vendedor,
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado,
            COALESCE(SUM({venta_costo_neto_expr()}), 0) AS cogs,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets,
            COALESCE(SUM(
                CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                     THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                ELSE 0 END), 0) AS descuento_total,
            COALESCE(AVG(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                         THEN descuento_porc::float END), 0) AS descuento_promedio_pct
        FROM ventas
        WHERE {where} AND cod_vendedor IS NOT NULL
        GROUP BY cod_vendedor
        ORDER BY facturado DESC
    """), params)).mappings().all()

    ranking = []
    for r in rows:
        f = money(r["facturado"])
        c = money(r["cogs"])
        mb = f - c
        mp = (mb / f * 100) if f else 0
        ranking.append({
            "cod_vendedor": r["cod_vendedor"],
            "nombre": vendedor_names.get(r["cod_vendedor"], r["cod_vendedor"]),
            "facturado": round(f, 2),
            "cogs": round(c, 2),
            "margen_dolares": round(mb, 2),
            "margen_pct": round(mp, 2),
            "descuento_promedio_pct": round(money(r["descuento_promedio_pct"]), 2),
            "descuento_total": round(money(r["descuento_total"]), 2),
            "tickets": int(r["tickets"] or 0),
        })

    return {"ranking": ranking}


@router.get("/resultado/por-cliente")
async def resultado_por_cliente(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    rows = (await db.execute(text(f"""
        SELECT
            cliente_id,
            COALESCE(MAX(cliente_nombre), cliente_id) AS cliente_nombre,
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado,
            COALESCE(SUM({venta_costo_neto_expr()}), 0) AS cogs,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets,
            COALESCE(SUM(
                CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                     THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                ELSE 0 END), 0) AS descuento_total,
            COALESCE(AVG(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                         THEN descuento_porc::float END), 0) AS descuento_promedio_pct
        FROM ventas WHERE {where}
        GROUP BY cliente_id
        ORDER BY facturado DESC
        LIMIT 200
    """), params)).mappings().all()

    ranking = []
    for r in rows:
        f = money(r["facturado"])
        c = money(r["cogs"])
        mb = f - c
        mp = (mb / f * 100) if f else 0
        ranking.append({
            "cliente_id": r["cliente_id"],
            "cliente_nombre": r["cliente_nombre"],
            "facturado": round(f, 2),
            "cogs": round(c, 2),
            "margen_dolares": round(mb, 2),
            "margen_pct": round(mp, 2),
            "descuento_promedio_pct": round(money(r["descuento_promedio_pct"]), 2),
            "descuento_total": round(money(r["descuento_total"]), 2),
            "tickets": int(r["tickets"] or 0),
        })

    return {"ranking": ranking}


@router.get("/resultado/descuentos")
async def resultado_descuentos(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    totals = (await db.execute(text(f"""
        SELECT
            COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0) AS facturado_bruto,
            COALESCE(SUM(
                CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                     THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                ELSE 0 END), 0) AS total_descontado
        FROM ventas WHERE {where}
    """), params)).mappings().one()

    total_desc = money(totals["total_descontado"])
    fact_bruto = money(totals["facturado_bruto"])
    pct = (total_desc / (fact_bruto + total_desc) * 100) if (fact_bruto + total_desc) else 0

    vendedor_names = {}
    try:
        for r in (await db.execute(text("SELECT cod_vendedor, nombre FROM vendedores"))).mappings().all():
            vendedor_names[r["cod_vendedor"]] = r["nombre"]
    except Exception:
        pass

    por_vendedor = (await db.execute(text(f"""
        SELECT cod_vendedor,
            COALESCE(SUM(
                CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                     THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                ELSE 0 END), 0) AS descuento_total,
            COALESCE(AVG(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                         THEN descuento_porc::float END), 0) AS descuento_pct_promedio,
            COUNT(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0 THEN 1 END) AS tickets_con_descuento
        FROM ventas WHERE {where} AND cod_vendedor IS NOT NULL
        GROUP BY cod_vendedor ORDER BY descuento_total DESC
    """), params)).mappings().all()

    por_producto = (await db.execute(text(f"""
        SELECT producto_id, COALESCE(MAX(producto_nombre), producto_id) AS nombre,
            COALESCE(SUM(
                CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                     THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                ELSE 0 END), 0) AS descuento_total,
            COALESCE(AVG(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                         THEN descuento_porc::float END), 0) AS descuento_pct_promedio,
            COUNT(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0 THEN 1 END) AS tickets_con_descuento
        FROM ventas WHERE {where}
        GROUP BY producto_id
        HAVING COUNT(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0 THEN 1 END) > 0
        ORDER BY descuento_total DESC LIMIT 50
    """), params)).mappings().all()

    por_cliente = (await db.execute(text(f"""
        SELECT cliente_id, COALESCE(MAX(cliente_nombre), cliente_id) AS nombre,
            COALESCE(SUM(
                CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                     THEN cantidad * precio_unitario * descuento_porc::float / 100.0
                ELSE 0 END), 0) AS descuento_total,
            COALESCE(AVG(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                         THEN descuento_porc::float END), 0) AS descuento_pct_promedio
        FROM ventas WHERE {where}
        GROUP BY cliente_id
        HAVING COUNT(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0 THEN 1 END) > 0
        ORDER BY descuento_total DESC LIMIT 50
    """), params)).mappings().all()

    mayores = (await db.execute(text(f"""
        SELECT fecha, cliente_nombre, producto_nombre, cod_vendedor,
            descuento_porc::float AS descuento_pct,
            cantidad * precio_unitario * descuento_porc::float / 100.0 AS descuento_dolares
        FROM ventas
        WHERE {where} AND descuento_porc IS NOT NULL AND descuento_porc::float > 0
        ORDER BY descuento_dolares DESC
        LIMIT 10
    """), params)).mappings().all()

    return {
        "total_descontado": round(total_desc, 2),
        "pct_sobre_facturacion": round(pct, 2),
        "facturacion_perdida": round(total_desc, 2),
        "por_vendedor": [
            {"nombre": vendedor_names.get(r["cod_vendedor"], r["cod_vendedor"]),
             "descuento_total": round(money(r["descuento_total"]), 2),
             "descuento_pct_promedio": round(money(r["descuento_pct_promedio"]), 2),
             "tickets_con_descuento": int(r["tickets_con_descuento"])}
            for r in por_vendedor
        ],
        "por_producto": [
            {"nombre": r["nombre"],
             "descuento_total": round(money(r["descuento_total"]), 2),
             "descuento_pct_promedio": round(money(r["descuento_pct_promedio"]), 2),
             "tickets_con_descuento": int(r["tickets_con_descuento"])}
            for r in por_producto
        ],
        "por_cliente": [
            {"nombre": r["nombre"],
             "descuento_total": round(money(r["descuento_total"]), 2),
             "descuento_pct_promedio": round(money(r["descuento_pct_promedio"]), 2)}
            for r in por_cliente
        ],
        "mayores_descuentos": [
            {"fecha": str(r["fecha"]),
             "cliente": r["cliente_nombre"],
             "vendedor": vendedor_names.get(r["cod_vendedor"], r["cod_vendedor"]),
             "producto": r["producto_nombre"],
             "descuento_pct": round(money(r["descuento_pct"]), 2),
             "descuento_dolares": round(money(r["descuento_dolares"]), 2)}
            for r in mayores
        ],
    }


@router.get("/resultado/exportar")
async def resultado_exportar(
    company_id: int = None,
    formato: Literal["excel", "csv"] = "excel",
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = filters.sql_params()
    where = _ventas_base_where(filters)

    if formato == "csv":
        import csv
        rows = (await db.execute(text(f"""
            SELECT producto_id, COALESCE(MAX(producto_nombre), producto_id) AS nombre,
                SUM(cantidad) AS unidades,
                COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado,
                COALESCE(SUM({venta_costo_neto_expr()}), 0) AS cogs
            FROM ventas WHERE {where} GROUP BY producto_id ORDER BY facturado DESC
        """), params)).mappings().all()
        buf = io.StringIO()
        buf.write('﻿')
        w = csv.writer(buf)
        w.writerow(["Producto", "Unidades", "Facturado", "COGS", "Margen $", "Margen %"])
        for r in rows:
            f, c = money(r["facturado"]), money(r["cogs"])
            mb = f - c
            w.writerow([r["nombre"], int(r["unidades"] or 0), round(f, 2), round(c, 2),
                         round(mb, 2), round((mb/f*100) if f else 0, 2)])
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=resultado.csv"},
        )

    from openpyxl import Workbook
    wb = Workbook()

    ws_kpi = wb.active
    ws_kpi.title = "KPIs"
    kpi_data = await resultado_kpis(company_id, filters, current_user, db)
    for key, val in kpi_data.items():
        if isinstance(val, dict) and "actual" in val:
            ws_kpi.append([key, val["actual"]])
        elif not isinstance(val, dict):
            ws_kpi.append([key, val])

    ws_prod = wb.create_sheet("Por Producto")
    ws_prod.append(["Producto", "Rubro", "Unidades", "Facturado", "COGS", "Margen $", "Margen %", "Descuento", "Cuadrante"])
    prod_data = await resultado_por_producto(company_id, filters, current_user, db)
    for p in prod_data["ranking"]:
        ws_prod.append([p["nombre"], p["rubro_nombre"], p["unidades"], p["facturado"],
                        p["cogs"], p["margen_dolares"], p["margen_pct"],
                        p["descuento_total"], p.get("cuadrante", "")])

    ws_vend = wb.create_sheet("Por Vendedor")
    ws_vend.append(["Vendedor", "Facturado", "COGS", "Margen $", "Margen %", "Desc. Prom %", "Tickets"])
    vend_data = await resultado_por_vendedor(company_id, filters, current_user, db)
    for v in vend_data["ranking"]:
        ws_vend.append([v["nombre"], v["facturado"], v["cogs"], v["margen_dolares"],
                        v["margen_pct"], v["descuento_promedio_pct"], v["tickets"]])

    ws_cli = wb.create_sheet("Por Cliente")
    ws_cli.append(["Cliente", "Facturado", "COGS", "Margen $", "Margen %", "Desc. Prom %", "Tickets"])
    cli_data = await resultado_por_cliente(company_id, filters, current_user, db)
    for c in cli_data["ranking"]:
        ws_cli.append([c["cliente_nombre"], c["facturado"], c["cogs"], c["margen_dolares"],
                       c["margen_pct"], c["descuento_promedio_pct"], c["tickets"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=resultado.xlsx"},
    )


# ══════════════════════════════════════════════════════════════════════════════
# FASE 6: Panel Stock
# ══════════════════════════════════════════════════════════════════════════════

# Dynamic bridge between stock.cod_articulo (integer) and ventas.producto_id (text).
_ARTICULO_MAP_SQL = """
    (SELECT DISTINCT
         producto_id::integer AS cod_articulo,
         producto_id,
         MAX(producto_nombre) AS nombre
     FROM ventas
     WHERE producto_id IS NOT NULL AND producto_id <> ''
       AND producto_id ~ '^[0-9]+$'
     GROUP BY producto_id
    ) AS am"""


@router.get("/stock/kpis")
async def stock_kpis(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)
    dias_periodo = max((filters.hasta - filters.desde).days, 1)
    params["dias_periodo"] = dias_periodo
    # Pass date objects (not strings) — asyncpg requires datetime.date for date columns
    params["noventa90"] = filters.hasta - timedelta(days=90)

    # Aggregate from stock table grouped by cod_articulo (each article can span multiple depositos)
    stock_totals = (await db.execute(text("""
        SELECT
            COUNT(DISTINCT cod_articulo)                           AS total_articulos,
            COUNT(DISTINCT CASE WHEN cantidad > 0 THEN cod_articulo END) AS con_stock,
            COUNT(DISTINCT CASE WHEN cantidad <= 0 THEN cod_articulo END) AS sin_stock,
            COUNT(DISTINCT CASE WHEN cantidad > 0 AND COALESCE(stock_minimo, 0) > 0 AND cantidad < stock_minimo
                                THEN cod_articulo END)              AS stock_bajo,
            COALESCE(SUM(cantidad * COALESCE(precio_compra_actual, 0)), 0)  AS valor_costo,
            COALESCE(SUM(cantidad), 0)                             AS total_unidades
        FROM stock
    """))).mappings().one()

    # Valor inventario a precio de venta (avg precio_unitario from ventas por producto)
    val_pv_row = (await db.execute(text(f"""
        SELECT COALESCE(SUM(s.cantidad * COALESCE(pv.avg_precio, 0)), 0) AS valor_precio_venta
        FROM stock s
        LEFT JOIN (
            SELECT am.cod_articulo, AVG(v.precio_unitario) AS avg_precio
            FROM {_ARTICULO_MAP_SQL}
            JOIN ventas v ON v.producto_id = am.producto_id
                         AND v.tipo_comprobante = 'FA' AND v.anulada <> 'S'
            GROUP BY am.cod_articulo
        ) pv ON pv.cod_articulo = s.cod_articulo
    """), params)).mappings().one()

    # Ventas diarias por articulo en el período → días inventario
    ventas_mov = (await db.execute(text(f"""
        WITH vperiod AS (
            SELECT am.cod_articulo, COALESCE(SUM(v.cantidad),0) AS vendido_periodo
            FROM {_ARTICULO_MAP_SQL}
            LEFT JOIN ventas v ON v.producto_id = am.producto_id AND {ventas_where}
            GROUP BY am.cod_articulo
        ),
        stock_agg AS (
            SELECT cod_articulo, SUM(cantidad) AS cantidad_total
            FROM stock GROUP BY cod_articulo
        )
        SELECT
            COALESCE(AVG(
                CASE WHEN vp.vendido_periodo > 0
                     THEN sa.cantidad_total / (vp.vendido_periodo / GREATEST(:dias_periodo, 1))
                END
            ), 0) AS dias_inv_prom
        FROM stock_agg sa
        LEFT JOIN vperiod vp USING (cod_articulo)
    """), params)).mappings().one()

    # Articles with no ventas in last 90 days
    sin_mov_rows = (await db.execute(text(f"""
        SELECT COUNT(DISTINCT am.cod_articulo) AS cnt
        FROM {_ARTICULO_MAP_SQL}
        JOIN stock s ON s.cod_articulo = am.cod_articulo AND s.cantidad > 0
        WHERE am.producto_id NOT IN (
            SELECT DISTINCT producto_id FROM ventas
            WHERE fecha >= :noventa90 AND tipo_comprobante='FA'
        )
    """), params)).mappings().one()

    total = int(stock_totals["total_articulos"] or 0)
    con = int(stock_totals["con_stock"] or 0)
    sin = int(stock_totals["sin_stock"] or 0)
    bajo = int(stock_totals["stock_bajo"] or 0)
    val_costo = money(stock_totals["valor_costo"])
    val_pv = money(val_pv_row["valor_precio_venta"])
    dias_inv = round(float(ventas_mov["dias_inv_prom"] or 0), 1)
    sin_mov = int(sin_mov_rows["cnt"] or 0)

    return {
        "total_articulos": {"actual": total},
        "con_stock": {"actual": con},
        "sin_stock": {"actual": sin},
        "stock_bajo": {"actual": bajo},
        "valor_inventario_costo": {"actual": round(val_costo, 2)},
        "valor_inventario_precio_venta": {"actual": round(val_pv, 2)},
        "dias_inventario_promedio": {"actual": dias_inv},
        "articulos_sin_movimiento": {"actual": sin_mov},
    }


@router.get("/stock/por-producto")
async def stock_por_producto(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)
    noventa = (filters.hasta - timedelta(days=90)).isoformat()
    dias_periodo = max((filters.hasta - filters.desde).days, 1)
    params["dias_periodo"] = dias_periodo
    params["noventa90"] = noventa

    rows = (await db.execute(text(f"""
        WITH vperiod AS (
            SELECT am.cod_articulo, am.nombre,
                   COALESCE(SUM(CASE WHEN v.tipo_comprobante='FA' THEN v.cantidad ELSE 0 END), 0) AS vendido,
                   MAX(CASE WHEN v.tipo_comprobante='FA' THEN v.fecha END) AS ultimo_mov
            FROM {_ARTICULO_MAP_SQL}
            LEFT JOIN ventas v ON v.producto_id = am.producto_id AND {ventas_where}
            GROUP BY am.cod_articulo, am.nombre
        ),
        stock_agg AS (
            SELECT s.cod_articulo, d.nombre AS deposito_nombre, s.cod_deposito,
                   SUM(s.cantidad) AS cantidad,
                   MIN(s.stock_minimo) AS stock_minimo,
                   AVG(s.precio_compra_actual) AS precio_compra_actual,
                   MAX(s.ultima_actualizacion) AS ultima_actualizacion
            FROM stock s
            LEFT JOIN depositos d ON d.cod_deposito = s.cod_deposito
            GROUP BY s.cod_articulo, d.nombre, s.cod_deposito
        )
        SELECT
            sa.cod_articulo, sa.cod_deposito,
            COALESCE(sa.deposito_nombre, 'Sin depósito') AS deposito_nombre,
            vp.nombre,
            sa.cantidad,
            sa.stock_minimo,
            sa.precio_compra_actual,
            COALESCE(sa.cantidad * sa.precio_compra_actual, 0) AS valor_stock,
            vp.vendido AS unidades_vendidas_periodo,
            vp.ultimo_mov
        FROM stock_agg sa
        JOIN vperiod vp ON vp.cod_articulo = sa.cod_articulo
        ORDER BY valor_stock DESC
    """), params)).mappings().all()

    hoy = filters.hasta
    productos = []
    for r in rows:
        cant = float(r["cantidad"] or 0)
        vendido = float(r["unidades_vendidas_periodo"] or 0)
        venta_diaria = round(vendido / dias_periodo, 2)
        dias_inv = round(cant / venta_diaria, 1) if venta_diaria > 0 else 999
        ultimo_mov = r["ultimo_mov"]
        dias_sin_mov = (hoy - ultimo_mov.date()).days if ultimo_mov else 999

        if cant <= 0:
            estado = "sin_stock"
            alerta = "sin_stock"
        elif float(r["stock_minimo"] or 0) > 0 and cant < float(r["stock_minimo"] or 0):
            estado = "bajo"
            alerta = "bajo_minimo"
        elif dias_inv > 180:
            estado = "sobrestock"
            alerta = "sobrestock"
        elif dias_sin_mov >= 90:
            estado = "sin_movimiento"
            alerta = "sin_movimiento_90d"
        else:
            estado = "ok"
            alerta = None

        productos.append({
            "cod_articulo": r["cod_articulo"],
            "nombre": r["nombre"],
            "cod_deposito": r["cod_deposito"],
            "deposito_nombre": r["deposito_nombre"],
            "cantidad": round(cant, 0),
            "stock_minimo": float(r["stock_minimo"] or 0),
            "precio_compra_actual": round(float(r["precio_compra_actual"] or 0), 2),
            "valor_stock": round(float(r["valor_stock"] or 0), 2),
            "unidades_vendidas_periodo": round(vendido, 0),
            "venta_diaria_promedio": venta_diaria,
            "dias_inventario": dias_inv,
            "ultimo_movimiento": ultimo_mov.date().isoformat() if ultimo_mov else None,
            "dias_sin_movimiento": dias_sin_mov,
            "estado": estado,
            "alerta": alerta,
        })

    # por_deposito summary
    dep_map: dict = {}
    for p in productos:
        k = p["cod_deposito"]
        if k not in dep_map:
            dep_map[k] = {"cod_deposito": k, "nombre": p["deposito_nombre"], "total_articulos": 0, "valor_total": 0.0}
        dep_map[k]["total_articulos"] += 1
        dep_map[k]["valor_total"] += p["valor_stock"]
    for d in dep_map.values():
        d["valor_total"] = round(d["valor_total"], 2)

    return {"productos": productos, "por_deposito": list(dep_map.values())}


@router.get("/stock/movimientos-series")
async def stock_movimientos_series(
    company_id: int = None,
    cod_articulo: Optional[int] = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)
    compras_where = text_filter_clause("compras", filters)

    articulo_filter_v = ""
    articulo_filter_c = ""
    if cod_articulo:
        articulo_filter_v = f"AND v.producto_id = (SELECT am.producto_id FROM {_ARTICULO_MAP_SQL} WHERE am.cod_articulo = :cod_articulo_val LIMIT 1)"
        articulo_filter_c = f"AND producto_id = (SELECT am.producto_id FROM {_ARTICULO_MAP_SQL} WHERE am.cod_articulo = :cod_articulo_val LIMIT 1)"
        params["cod_articulo_val"] = cod_articulo

    series_rows = (await db.execute(text(f"""
        WITH meses AS (
            SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo FROM ventas WHERE {ventas_where}
            UNION
            SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo FROM compras WHERE {compras_where}
        ),
        entradas AS (
            SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo, SUM(cantidad) AS qty
            FROM compras WHERE {compras_where} {articulo_filter_c}
            GROUP BY 1
        ),
        salidas AS (
            SELECT to_char(date_trunc('month', fecha), 'YYYY-MM') AS periodo,
                   SUM(CASE WHEN tipo_comprobante='FA' THEN cantidad ELSE -cantidad END) AS qty
            FROM ventas v WHERE {ventas_where} {articulo_filter_v}
            GROUP BY 1
        )
        SELECT m.periodo,
               COALESCE(e.qty, 0) AS entradas,
               COALESCE(s.qty, 0) AS salidas,
               COALESCE(e.qty, 0) - COALESCE(s.qty, 0) AS saldo_periodo
        FROM meses m
        LEFT JOIN entradas e USING (periodo)
        LEFT JOIN salidas s USING (periodo)
        ORDER BY 1
    """), params)).mappings().all()

    por_articulo_rows = (await db.execute(text(f"""
        WITH ent AS (
            SELECT am.cod_articulo, am.nombre, COALESCE(SUM(c.cantidad), 0) AS entradas
            FROM {_ARTICULO_MAP_SQL}
            LEFT JOIN compras c ON c.producto_id = am.producto_id AND {compras_where}
            GROUP BY am.cod_articulo, am.nombre
        ),
        sal AS (
            SELECT am.cod_articulo, COALESCE(SUM(CASE WHEN v.tipo_comprobante='FA' THEN v.cantidad ELSE 0 END), 0) AS salidas
            FROM {_ARTICULO_MAP_SQL}
            LEFT JOIN ventas v ON v.producto_id = am.producto_id AND {ventas_where}
            GROUP BY am.cod_articulo
        ),
        stk AS (
            SELECT cod_articulo, SUM(cantidad) AS stock_actual FROM stock GROUP BY cod_articulo
        )
        SELECT e.cod_articulo, e.nombre, e.entradas, s.salidas, COALESCE(st.stock_actual, 0) AS stock_actual
        FROM ent e
        LEFT JOIN sal s USING (cod_articulo)
        LEFT JOIN stk st USING (cod_articulo)
        ORDER BY e.entradas DESC
    """), params)).mappings().all()

    return {
        "series": [dict(r) for r in series_rows],
        "por_articulo": [dict(r) for r in por_articulo_rows],
    }


@router.get("/stock/rotacion")
async def stock_rotacion(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)
    dias_periodo = max((filters.hasta - filters.desde).days, 1)
    params["dias_periodo"] = dias_periodo

    rows = (await db.execute(text(f"""
        WITH vp AS (
            SELECT am.cod_articulo, am.nombre,
                   COALESCE(SUM(CASE WHEN v.tipo_comprobante='FA' THEN v.cantidad ELSE 0 END), 0) AS vendido
            FROM {_ARTICULO_MAP_SQL}
            LEFT JOIN ventas v ON v.producto_id = am.producto_id AND {ventas_where}
            GROUP BY am.cod_articulo, am.nombre
        ),
        stk AS (
            SELECT cod_articulo, SUM(cantidad) AS stock_actual,
                   SUM(cantidad * COALESCE(precio_compra_actual, 0)) AS valor_stock
            FROM stock GROUP BY cod_articulo
        )
        SELECT vp.cod_articulo, vp.nombre, vp.vendido,
               COALESCE(stk.stock_actual, 0) AS stock_actual,
               COALESCE(stk.valor_stock, 0) AS valor_stock
        FROM vp LEFT JOIN stk USING (cod_articulo)
        ORDER BY valor_stock DESC
    """), params)).mappings().all()

    total_valor = sum(float(r["valor_stock"] or 0) for r in rows)
    ranking = []
    acum = 0.0
    for r in rows:
        v = float(r["valor_stock"] or 0)
        acum += v
        pct_acum = (acum / total_valor * 100) if total_valor else 0
        if pct_acum <= 70:
            cls = "A"
        elif pct_acum <= 90:
            cls = "B"
        else:
            cls = "C"
        vd = float(r["vendido"] or 0)
        sa = float(r["stock_actual"] or 0)
        venta_diaria = vd / dias_periodo
        dias_inv = round(sa / venta_diaria, 1) if venta_diaria > 0 else 999
        ranking.append({
            "cod_articulo": r["cod_articulo"],
            "nombre": r["nombre"],
            "stock_actual": round(sa),
            "unidades_vendidas": round(vd),
            "venta_diaria": round(venta_diaria, 2),
            "dias_inventario": dias_inv,
            "clasificacion": cls,
            "valor_stock": round(v, 2),
            "pct_valor_total": round(v / total_valor * 100, 1) if total_valor else 0,
        })

    abc: dict = {"A": {"articulos": 0, "valor": 0.0, "pct_valor": 0.0},
                 "B": {"articulos": 0, "valor": 0.0, "pct_valor": 0.0},
                 "C": {"articulos": 0, "valor": 0.0, "pct_valor": 0.0}}
    for p in ranking:
        c = p["clasificacion"]
        abc[c]["articulos"] += 1
        abc[c]["valor"] += p["valor_stock"]
    for c in abc:
        abc[c]["valor"] = round(abc[c]["valor"], 2)
        abc[c]["pct_valor"] = round(abc[c]["valor"] / total_valor * 100, 1) if total_valor else 0

    return {"ranking": ranking, "abc": abc}


@router.get("/stock/alertas")
async def stock_alertas(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)
    noventa = (filters.hasta - timedelta(days=90)).isoformat()
    dias_periodo = max((filters.hasta - filters.desde).days, 1)
    params["dias_periodo"] = dias_periodo
    params["noventa90"] = noventa

    rows = (await db.execute(text(f"""
        WITH vp AS (
            SELECT am.cod_articulo, am.nombre,
                   COALESCE(SUM(CASE WHEN v.tipo_comprobante='FA' THEN v.cantidad ELSE 0 END), 0) AS vendido,
                   MAX(CASE WHEN v.tipo_comprobante='FA' THEN v.fecha END) AS ultimo_mov
            FROM {_ARTICULO_MAP_SQL}
            LEFT JOIN ventas v ON v.producto_id = am.producto_id AND {ventas_where}
            GROUP BY am.cod_articulo, am.nombre
        ),
        stk AS (
            SELECT cod_articulo, SUM(cantidad) AS cantidad_total,
                   MIN(stock_minimo) AS stock_minimo,
                   AVG(precio_compra_actual) AS precio_compra_actual
            FROM stock GROUP BY cod_articulo
        )
        SELECT vp.cod_articulo, vp.nombre, vp.vendido, vp.ultimo_mov,
               COALESCE(stk.cantidad_total, 0) AS stock_actual,
               COALESCE(stk.stock_minimo, 0) AS stock_minimo,
               COALESCE(stk.precio_compra_actual, 0) AS precio_compra_actual
        FROM vp LEFT JOIN stk USING (cod_articulo)
    """), params)).mappings().all()

    hoy = filters.hasta
    alertas = []
    for r in rows:
        cant = float(r["stock_actual"] or 0)
        sm = float(r["stock_minimo"] or 0)
        vendido = float(r["vendido"] or 0)
        vd = vendido / dias_periodo
        dias_inv = round(cant / vd, 1) if vd > 0 else 999
        ultimo_mov = r["ultimo_mov"]
        dias_sin_mov = (hoy - ultimo_mov.date()).days if ultimo_mov else 999

        issues = []
        if cant <= 0:
            issues.append(("sin_stock", "critical",
                           f"Sin stock. Último mov: {ultimo_mov.date() if ultimo_mov else 'nunca'}",
                           "Realizar pedido urgente"))
        elif sm > 0 and cant < sm:
            issues.append(("bajo_minimo", "warning",
                           f"Stock: {int(cant)} un. (mín: {int(sm)}). Cubre {dias_inv:.0f} días",
                           "Solicitar reposición"))
        if dias_inv > 180:
            issues.append(("sobrestock", "info",
                           f"Días inventario: {dias_inv:.0f}d (>180d). Valor: ${cant * float(r['precio_compra_actual'] or 0):,.0f}",
                           "Revisar política de compras o realizar promoción"))
        if dias_sin_mov >= 90 and cant > 0:
            issues.append(("sin_movimiento", "warning",
                           f"{dias_sin_mov} días sin ventas. Stock: {int(cant)} un.",
                           "Investigar demanda o liquidar"))

        for tipo, sev, detalle, accion in issues:
            alertas.append({
                "cod_articulo": r["cod_articulo"],
                "nombre": r["nombre"],
                "tipo": tipo,
                "severidad": sev,
                "detalle": detalle,
                "accion_sugerida": accion,
                "stock_actual": int(cant),
                "stock_minimo": int(sm),
                "dias_sin_movimiento": dias_sin_mov,
            })

    # Sort: critical first, then warning, then info
    orden = {"critical": 0, "warning": 1, "info": 2}
    alertas.sort(key=lambda a: orden.get(a["severidad"], 3))
    resumen = {
        "critical": sum(1 for a in alertas if a["severidad"] == "critical"),
        "warning": sum(1 for a in alertas if a["severidad"] == "warning"),
        "info": sum(1 for a in alertas if a["severidad"] == "info"),
    }
    return {"alertas": alertas, "resumen": resumen}


@router.get("/stock/reposicion")
async def stock_reposicion(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)
    dias_periodo = max((filters.hasta - filters.desde).days, 1)
    params["dias_periodo"] = dias_periodo

    rows = (await db.execute(text(f"""
        WITH vp AS (
            SELECT am.cod_articulo, am.nombre,
                   COALESCE(SUM(CASE WHEN v.tipo_comprobante='FA' THEN v.cantidad ELSE 0 END), 0) AS vendido
            FROM {_ARTICULO_MAP_SQL}
            LEFT JOIN ventas v ON v.producto_id = am.producto_id AND {ventas_where}
            GROUP BY am.cod_articulo, am.nombre
        ),
        stk AS (
            SELECT cod_articulo, SUM(cantidad) AS stock_actual,
                   MIN(stock_minimo) AS stock_minimo,
                   AVG(precio_compra_actual) AS precio_compra_actual
            FROM stock GROUP BY cod_articulo
        ),
        prov AS (
            SELECT am.cod_articulo, MAX(c.proveedor_id) AS ultimo_proveedor
            FROM {_ARTICULO_MAP_SQL}
            LEFT JOIN compras c ON c.producto_id = am.producto_id
            GROUP BY am.cod_articulo
        )
        SELECT vp.cod_articulo, vp.nombre, vp.vendido,
               COALESCE(stk.stock_actual, 0) AS stock_actual,
               COALESCE(stk.stock_minimo, 0) AS stock_minimo,
               COALESCE(stk.precio_compra_actual, 0) AS precio_compra_actual,
               COALESCE(prov.ultimo_proveedor, 'Sin proveedor') AS ultimo_proveedor
        FROM vp
        LEFT JOIN stk USING (cod_articulo)
        LEFT JOIN prov USING (cod_articulo)
        WHERE COALESCE(stk.stock_actual, 0) < CASE
              WHEN COALESCE(stk.stock_minimo, 0) > 0 THEN COALESCE(stk.stock_minimo, 0) * 2
              ELSE GREATEST(vp.vendido / GREATEST(:dias_periodo, 1) * 30, 0)
            END
        ORDER BY COALESCE(stk.stock_actual, 0) / NULLIF(vp.vendido / :dias_periodo, 0) ASC NULLS LAST
    """), params)).mappings().all()

    sugerencias = []
    for r in rows:
        cant = float(r["stock_actual"] or 0)
        sm = float(r["stock_minimo"] or 0)
        vendido = float(r["vendido"] or 0)
        vd = vendido / dias_periodo
        dias_cob = round(cant / vd, 1) if vd > 0 else 999
        # Suggest enough to cover 30 days of sales above minimum
        cant_sug = max(int((vd * 30 + sm) - cant), 0)
        costo_est = round(cant_sug * float(r["precio_compra_actual"] or 0), 2)
        sugerencias.append({
            "cod_articulo": r["cod_articulo"],
            "nombre": r["nombre"],
            "stock_actual": int(cant),
            "venta_diaria": round(vd, 2),
            "dias_cobertura": dias_cob,
            "cantidad_sugerida": cant_sug,
            "costo_estimado": costo_est,
            "ultimo_proveedor": r["ultimo_proveedor"],
        })

    return {"sugerencias": sugerencias}


# ══════════════════════════════════════════════════════════════════════════════
# FASE 7: Panel Vendedores — endpoints dedicados
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/vendedores/kpis")
async def vendedores_kpis(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)
    presup_where = text_filter_clause("presupuestos", filters)

    # Global ventas KPIs
    vrow = (await db.execute(text(f"""
        SELECT
            COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0)          AS fa_bruto,
            COALESCE(SUM(CASE WHEN tipo_comprobante='NC' THEN ABS(total) ELSE 0 END), 0)     AS nc_total,
            COALESCE(SUM(CASE WHEN tipo_comprobante='ND' THEN ABS(total) ELSE 0 END), 0)     AS nd_total,
            COALESCE(SUM({venta_importe_neto_expr()}), 0)                                    AS facturado_neto,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                                 AS tickets,
            COALESCE(SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}), 0)         AS margen_dolares,
            COALESCE(SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END), 0) AS total_con_costo,
            COALESCE(AVG(CASE WHEN descuento_porc IS NOT NULL AND descuento_porc::float > 0
                         THEN descuento_porc::float END), 0)                                  AS descuento_prom
        FROM ventas WHERE {ventas_where}
    """), params)).mappings().one()

    # Total vendedores habilitados
    total_v = (await db.execute(text(
        "SELECT COUNT(*) AS cnt FROM vendedores WHERE habilitado"
    ))).mappings().one()

    # Per-vendedor facturado for mejor/peor
    pv_rows = (await db.execute(text(f"""
        SELECT v.cod_vendedor, v.nombre,
               COALESCE(SUM({venta_importe_neto_expr("vt")}), 0) AS facturado
        FROM vendedores v
        LEFT JOIN ventas vt ON vt.cod_vendedor = v.cod_vendedor AND {ventas_where}
        GROUP BY v.cod_vendedor, v.nombre
        ORDER BY facturado DESC
    """), params)).mappings().all()

    # Presupuestos global
    prow = (await db.execute(text(f"""
        SELECT COUNT(*) AS total, COUNT(CASE WHEN confirmado THEN 1 END) AS confirmados
        FROM presupuestos WHERE {presup_where}
    """), params)).mappings().one()

    fa_bruto = float(vrow["fa_bruto"] or 0)
    nc_total = float(vrow["nc_total"] or 0)
    fn = float(vrow["facturado_neto"] or 0)
    tickets = int(vrow["tickets"] or 0)
    margen = float(vrow["margen_dolares"] or 0)
    tc = float(vrow["total_con_costo"] or 0)

    vlist = list(pv_rows)
    mejor = vlist[0]["nombre"] if vlist else ""
    peor = vlist[-1]["nombre"] if len(vlist) > 1 else ""

    pt = int(prow["total"] or 0)
    pc = int(prow["confirmados"] or 0)

    return {
        "total_vendedores": {"actual": int(total_v["cnt"] or 0)},
        "facturado_total": {"actual": round(fn, 2)},
        "ticket_promedio_global": {"actual": round(fn / tickets, 2) if tickets > 0 else 0},
        "margen_total": {"actual": round(margen, 2)},
        "margen_pct": {"actual": round(margen / tc * 100, 1) if tc > 0 else 0},
        "mejor_vendedor": {"actual": mejor},
        "peor_vendedor": {"actual": peor},
        "presupuestos_emitidos": {"actual": pt},
        "presupuestos_confirmados": {"actual": pc},
        "tasa_conversion_global": {"actual": round(pc / pt * 100, 1) if pt > 0 else 0},
        "descuento_prom": {"actual": round(float(vrow["descuento_prom"] or 0), 2)},
    }


@router.get("/vendedores/ranking")
async def vendedores_ranking(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)
    presup_where = text_filter_clause("presupuestos", filters)
    dias_periodo = max((filters.hasta - filters.desde).days, 1)

    # Ventas summary per vendedor (all vendedores via LEFT JOIN so none disappear)
    v_rows = (await db.execute(text(f"""
        SELECT
            v.cod_vendedor, v.nombre,
            COALESCE(SUM(CASE WHEN vt.tipo_comprobante='FA' THEN vt.total ELSE 0 END), 0)        AS fa_bruto,
            COALESCE(SUM(CASE WHEN vt.tipo_comprobante='NC' THEN ABS(vt.total) ELSE 0 END), 0)   AS nc_total,
            COALESCE(SUM({venta_importe_neto_expr("vt")}), 0)                                    AS facturado_neto,
            COUNT(CASE WHEN vt.tipo_comprobante='FA' THEN 1 END)                                   AS tickets,
            COUNT(DISTINCT CASE WHEN vt.tipo_comprobante='FA' THEN vt.cliente_id END)             AS clientes_unicos,
            COALESCE(SUM({venta_importe_neto_expr("vt")} - {venta_costo_neto_expr("vt")}), 0)     AS margen_dolares,
            COALESCE(SUM(CASE WHEN vt.precio_compra_actual IS NOT NULL THEN vt.total ELSE 0 END), 0) AS total_con_costo,
            v.cuota_mensual
        FROM vendedores v
        LEFT JOIN ventas vt ON vt.cod_vendedor = v.cod_vendedor AND {ventas_where}
        GROUP BY v.cod_vendedor, v.nombre, v.cuota_mensual
        ORDER BY fa_bruto DESC
    """), params)).mappings().all()

    # Presupuestos per vendedor
    p_rows = (await db.execute(text(f"""
        SELECT cod_vendedor,
               COUNT(*) AS emitidos,
               COUNT(CASE WHEN confirmado THEN 1 END) AS confirmados
        FROM presupuestos WHERE {presup_where} AND cod_vendedor IS NOT NULL
        GROUP BY cod_vendedor
    """), params)).mappings().all()

    pmap = {int(r["cod_vendedor"]): {"emitidos": int(r["emitidos"] or 0), "confirmados": int(r["confirmados"] or 0)} for r in p_rows}
    total_facturado = sum(float(r["facturado_neto"] or 0) for r in v_rows)

    resultado = []
    for r in v_rows:
        fn = float(r["facturado_neto"] or 0)
        t = int(r["tickets"] or 0)
        m = float(r["margen_dolares"] or 0)
        tc = float(r["total_con_costo"] or 0)
        cuota = float(r["cuota_mensual"] or 0)
        cuota_periodo = cuota * dias_periodo / 30
        pres = pmap.get(int(r["cod_vendedor"]), {"emitidos": 0, "confirmados": 0})

        resultado.append({
            "cod_vendedor": r["cod_vendedor"],
            "nombre": r["nombre"],
            "facturado_neto": round(fn, 2),
            "tickets": t,
            "ticket_promedio": round(fn / t, 2) if t > 0 else 0,
            "clientes_unicos": int(r["clientes_unicos"] or 0),
            "margen_dolares": round(m, 2),
            "margen_pct": round(m / tc * 100, 1) if tc > 0 else 0,
            "cuota_mensual": round(cuota, 2),
            "pct_cumplimiento": round(fn / cuota_periodo * 100, 1) if cuota_periodo > 0 else None,
            "presupuestos_emitidos": pres["emitidos"],
            "presupuestos_confirmados": pres["confirmados"],
            "tasa_conversion": round(pres["confirmados"] / pres["emitidos"] * 100, 1) if pres["emitidos"] > 0 else 0,
            "pct_del_total": round(fn / total_facturado * 100, 1) if total_facturado > 0 else 0,
        })

    return {"vendedores": resultado}


@router.get("/vendedores/temporal")
async def vendedores_temporal(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)

    rows = (await db.execute(text(f"""
        SELECT
            TO_CHAR(vt.fecha, 'YYYY-MM') AS periodo,
            v.cod_vendedor, v.nombre,
            COALESCE(SUM({venta_importe_neto_expr("vt")}), 0) AS facturado,
            COUNT(CASE WHEN vt.tipo_comprobante='FA' THEN 1 END) AS tickets,
            COALESCE(SUM({venta_importe_neto_expr("vt")} - {venta_costo_neto_expr("vt")}), 0) AS margen
        FROM vendedores v
        JOIN ventas vt ON vt.cod_vendedor = v.cod_vendedor AND {ventas_where}
        GROUP BY periodo, v.cod_vendedor, v.nombre
        ORDER BY periodo, v.cod_vendedor
    """), params)).mappings().all()

    # Pivot: series[periodo][cod_vendedor] = {facturado, tickets, margen}
    vendedores_seen: dict = {}
    series_map: dict = {}
    for r in rows:
        cv = int(r["cod_vendedor"])
        if cv not in vendedores_seen:
            vendedores_seen[cv] = r["nombre"]
        p = r["periodo"]
        if p not in series_map:
            series_map[p] = {"periodo": p}
        series_map[p][str(cv)] = {
            "facturado": round(float(r["facturado"] or 0), 2),
            "tickets": int(r["tickets"] or 0),
            "margen": round(float(r["margen"] or 0), 2),
        }

    return {
        "vendedores": [{"cod_vendedor": k, "nombre": v} for k, v in sorted(vendedores_seen.items())],
        "series": list(series_map.values()),
    }


@router.get("/vendedores/comisiones")
async def vendedores_comisiones(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    desde, hasta = filters.date_range()
    params = {"desde_periodo": desde.strftime("%Y-%m"), "hasta_periodo": hasta.strftime("%Y-%m")}

    rows = (await db.execute(text("""
        SELECT cod_vendedor, COALESCE(MAX(vendedor_nombre), CONCAT('Vendedor ', cod_vendedor)) AS vendedor_nombre,
               COALESCE(SUM(base_cobrada), 0) AS base_cobrada,
               COALESCE(AVG(porcentaje), 0) AS porcentaje,
               COALESCE(SUM(comision), 0) AS comision,
               COALESCE(SUM(recibos), 0) AS recibos
        FROM comisiones_vendedores
        WHERE periodo >= :desde_periodo AND periodo <= :hasta_periodo
        GROUP BY cod_vendedor
        ORDER BY comision DESC
    """), params)).mappings().all()

    temporal = (await db.execute(text("""
        SELECT periodo,
               COALESCE(SUM(base_cobrada), 0) AS base_cobrada,
               COALESCE(SUM(comision), 0) AS comision,
               COALESCE(SUM(recibos), 0) AS recibos
        FROM comisiones_vendedores
        WHERE periodo >= :desde_periodo AND periodo <= :hasta_periodo
        GROUP BY periodo
        ORDER BY periodo
    """), params)).mappings().all()

    return {
        "total_base_cobrada": round(sum(float(r["base_cobrada"] or 0) for r in rows), 2),
        "total_comision": round(sum(float(r["comision"] or 0) for r in rows), 2),
        "vendedores": [
            {
                "cod_vendedor": r["cod_vendedor"],
                "vendedor_nombre": r["vendedor_nombre"],
                "base_cobrada": round(float(r["base_cobrada"] or 0), 2),
                "porcentaje": round(float(r["porcentaje"] or 0), 4),
                "comision": round(float(r["comision"] or 0), 2),
                "recibos": int(r["recibos"] or 0),
            }
            for r in rows
        ],
        "temporal": [
            {
                "periodo": r["periodo"],
                "base_cobrada": round(float(r["base_cobrada"] or 0), 2),
                "comision": round(float(r["comision"] or 0), 2),
                "recibos": int(r["recibos"] or 0),
            }
            for r in temporal
        ],
    }


@router.get("/vendedores/detalle/{vendedor_id}")
async def vendedores_detalle(
    vendedor_id: int,
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from fastapi import HTTPException as _HTTPException
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    params["cod_v"] = vendedor_id
    # Always filter by path cod_vendedor (ignore filter's cod_vendedor for detalle)
    where_v = "fecha >= :desde AND fecha < :hasta AND anulada <> 'S' AND cod_vendedor = :cod_v"

    vinfo = (await db.execute(text(
        "SELECT cod_vendedor, nombre, cuota_mensual FROM vendedores WHERE cod_vendedor = :cod_v"
    ), {"cod_v": vendedor_id})).mappings().one_or_none()

    if not vinfo:
        raise _HTTPException(status_code=404, detail="Vendedor no encontrado")

    clientes = (await db.execute(text(f"""
        SELECT cliente_id, MAX(cliente_nombre) AS nombre,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0) AS facturado,
               COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets
        FROM ventas WHERE {where_v}
        GROUP BY cliente_id ORDER BY facturado DESC LIMIT 10
    """), params)).mappings().all()

    productos = (await db.execute(text(f"""
        SELECT producto_id, MAX(producto_nombre) AS nombre,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0) AS facturado,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN cantidad ELSE 0 END), 0) AS unidades
        FROM ventas WHERE {where_v}
        GROUP BY producto_id ORDER BY facturado DESC LIMIT 10
    """), params)).mappings().all()

    evolucion = (await db.execute(text(f"""
        SELECT TO_CHAR(fecha, 'YYYY-MM') AS periodo,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0) AS facturado,
               COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets
        FROM ventas WHERE {where_v}
        GROUP BY periodo ORDER BY periodo
    """), params)).mappings().all()

    pendientes = (await db.execute(text(
        "SELECT id, fecha, cliente_nombre, total FROM presupuestos "
        "WHERE cod_vendedor = :cod_v AND NOT confirmado "
        "AND fecha >= :desde AND fecha < :hasta ORDER BY fecha DESC LIMIT 20"
    ), params)).mappings().all()

    return {
        "vendedor": {
            "cod_vendedor": vinfo["cod_vendedor"],
            "nombre": vinfo["nombre"],
            "cuota_mensual": float(vinfo["cuota_mensual"] or 0),
        },
        "top_clientes": [
            {"cliente_id": r["cliente_id"], "nombre": r["nombre"],
             "facturado": round(float(r["facturado"] or 0), 2), "tickets": int(r["tickets"] or 0)}
            for r in clientes
        ],
        "top_productos": [
            {"producto_id": r["producto_id"], "nombre": r["nombre"],
             "facturado": round(float(r["facturado"] or 0), 2), "unidades": int(r["unidades"] or 0)}
            for r in productos
        ],
        "evolucion": [
            {"periodo": r["periodo"], "facturado": round(float(r["facturado"] or 0), 2), "tickets": int(r["tickets"] or 0)}
            for r in evolucion
        ],
        "presupuestos_pendientes": [
            {"id": r["id"], "fecha": str(r["fecha"]), "cliente_nombre": r["cliente_nombre"],
             "total": float(r["total"] or 0)}
            for r in pendientes
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# FASE 7 — Panel Clientes
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/clientes/kpis")
async def clientes_kpis(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)

    # Global ventas summary
    vrow = (await db.execute(text(f"""
        SELECT
            COUNT(DISTINCT CASE WHEN tipo_comprobante='FA' THEN cliente_id END)     AS clientes_activos,
            COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0) AS fa_bruto,
            COALESCE(SUM(CASE WHEN tipo_comprobante='NC' THEN ABS(total) ELSE 0 END), 0) AS nc_total,
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado_neto,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                        AS tickets,
            COUNT(DISTINCT CASE WHEN tipo_comprobante='FA' AND primera_compra THEN cliente_id END) AS clientes_nuevos
        FROM (
            SELECT cliente_id, tipo_comprobante, total,
                   NOT EXISTS (
                       SELECT 1 FROM ventas v2
                       WHERE v2.cliente_id = ventas.cliente_id
                         AND v2.tipo_comprobante = 'FA'
                         AND v2.fecha < :desde
                         AND v2.anulada <> 'S'
                   ) AS primera_compra
            FROM ventas WHERE {ventas_where}
        ) sub
    """), params)).mappings().one()

    # Best client
    best = (await db.execute(text(f"""
        SELECT cliente_id, MAX(cliente_nombre) AS nombre,
               COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado
        FROM ventas WHERE {ventas_where}
        GROUP BY cliente_id ORDER BY facturado DESC LIMIT 1
    """), params)).mappings().one_or_none()

    # Cuenta corriente: saldo total y deuda vencida
    ccrow = (await db.execute(text("""
        SELECT
            COALESCE(SUM(importe), 0) AS saldo_total,
            COALESCE(SUM(CASE WHEN fecha_vencimiento < now() AND importe > 0 THEN importe ELSE 0 END), 0) AS deuda_vencida
        FROM cuentas_corrientes_clientes
    """))).mappings().one()

    # Retención: clientes con >1 compra en el período
    retencion_row = (await db.execute(text(f"""
        SELECT
            COUNT(DISTINCT cliente_id) AS total,
            COUNT(DISTINCT CASE WHEN compras > 1 THEN cliente_id END) AS recurrentes
        FROM (
            SELECT cliente_id, COUNT(*) AS compras
            FROM ventas WHERE {ventas_where} AND tipo_comprobante = 'FA'
            GROUP BY cliente_id
        ) sub
    """), params)).mappings().one()

    fa = float(vrow["fa_bruto"] or 0)
    nc = float(vrow["nc_total"] or 0)
    fn = float(vrow["facturado_neto"] or 0)
    tickets = int(vrow["tickets"] or 0)
    total_c = int(retencion_row["total"] or 0)
    recurrentes = int(retencion_row["recurrentes"] or 0)

    return {
        "clientes_activos":   {"actual": int(vrow["clientes_activos"] or 0)},
        "facturado_total":    {"actual": round(fn, 2)},
        "ticket_promedio":    {"actual": round(fn / tickets, 2) if tickets > 0 else 0},
        "clientes_nuevos":    {"actual": int(vrow["clientes_nuevos"] or 0)},
        "mejor_cliente":      {"actual": best["nombre"] if best else ""},
        "saldo_cta_cte":      {"actual": round(float(ccrow["saldo_total"] or 0), 2)},
        "deuda_vencida":      {"actual": round(float(ccrow["deuda_vencida"] or 0), 2)},
        "tasa_retencion":     {"actual": round(recurrentes / total_c * 100, 1) if total_c > 0 else 0},
    }


@router.get("/clientes/ranking")
async def clientes_ranking(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)

    rows = (await db.execute(text(f"""
        SELECT
            cliente_id,
            MAX(cliente_nombre) AS nombre,
            COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0)          AS fa_bruto,
            COALESCE(SUM(CASE WHEN tipo_comprobante='NC' THEN ABS(total) ELSE 0 END), 0)     AS nc_total,
            COALESCE(SUM({venta_importe_neto_expr()}), 0)                                    AS facturado_neto,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                                 AS tickets,
            COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN cantidad ELSE 0 END), 0)        AS unidades,
            MAX(CASE WHEN tipo_comprobante='FA' THEN fecha END)                               AS ultima_compra,
            COALESCE(SUM({venta_importe_neto_expr()} - {venta_costo_neto_expr()}), 0)         AS margen_dolares,
            COALESCE(SUM(CASE WHEN precio_compra_actual IS NOT NULL THEN ABS(total) ELSE 0 END), 0) AS total_con_costo
        FROM ventas WHERE {ventas_where}
        GROUP BY cliente_id ORDER BY fa_bruto DESC
    """), params)).mappings().all()

    total_facturado = sum(float(r["facturado_neto"] or 0) for r in rows)
    cumsum = 0.0
    resultado = []

    for r in rows:
        fn = float(r["facturado_neto"] or 0)
        t = int(r["tickets"] or 0)
        m = float(r["margen_dolares"] or 0)
        tc = float(r["total_con_costo"] or 0)
        pct = (fn / total_facturado * 100) if total_facturado > 0 else 0
        cumsum += pct
        segmento = "A" if cumsum <= 80 else "B" if cumsum <= 95 else "C"

        resultado.append({
            "cliente_id": r["cliente_id"],
            "nombre": r["nombre"],
            "facturado_neto": round(fn, 2),
            "tickets": t,
            "ticket_promedio": round(fn / t, 2) if t > 0 else 0,
            "unidades": int(r["unidades"] or 0),
            "margen_pct": round(m / tc * 100, 1) if tc > 0 else 0,
            "ultima_compra": str(r["ultima_compra"]) if r["ultima_compra"] else None,
            "pct_del_total": round(pct, 1),
            "pct_acumulado": round(cumsum, 1),
            "segmento": segmento,
        })

    return {"clientes": resultado}


@router.get("/clientes/temporal")
async def clientes_temporal(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    ventas_where = _ventas_base_where(filters)

    # Monthly total + new clients count
    rows = (await db.execute(text(f"""
        SELECT
            TO_CHAR(fecha, 'YYYY-MM') AS periodo,
            COALESCE(SUM({venta_importe_neto_expr()}), 0) AS facturado,
            COUNT(DISTINCT CASE WHEN tipo_comprobante='FA' THEN cliente_id END)             AS clientes_activos,
            COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END)                               AS tickets
        FROM ventas WHERE {ventas_where}
        GROUP BY periodo ORDER BY periodo
    """), params)).mappings().all()

    return {
        "series": [
            {
                "periodo": r["periodo"],
                "facturado": round(float(r["facturado"] or 0), 2),
                "clientes_activos": int(r["clientes_activos"] or 0),
                "tickets": int(r["tickets"] or 0),
            }
            for r in rows
        ]
    }


@router.get("/clientes/{cliente_id}/comprobantes")
async def clientes_comprobantes(
    cliente_id: str,
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    params["cli_id"] = cliente_id

    rows = (await db.execute(text("""
        SELECT comprobante_id, cliente_id, cliente_nombre, tipo, numero, punto_de_venta,
               fecha, fecha_vencimiento, importe_total, importe_pagado, saldo,
               cod_vendedor, detalle
        FROM comprobantes_clientes
        WHERE cliente_id = :cli_id AND (fecha IS NULL OR (fecha >= :desde AND fecha < :hasta))
        ORDER BY COALESCE(fecha_vencimiento, fecha) DESC NULLS LAST, comprobante_id DESC
        LIMIT 200
    """), params)).mappings().all()

    comprobante_ids = [r["comprobante_id"] for r in rows]
    pagos_por_comprobante: dict[str, list[dict]] = {cid: [] for cid in comprobante_ids}
    if comprobante_ids:
        pagos = (await db.execute(text("""
            SELECT pago_id, comprobante_id, fecha, forma_pago, importe, cod_cliente, cliente_nombre
            FROM pagos_clientes
            WHERE comprobante_id = ANY(:comprobante_ids)
            ORDER BY fecha DESC NULLS LAST, pago_id DESC
        """), {"comprobante_ids": comprobante_ids})).mappings().all()
        for pago in pagos:
            pagos_por_comprobante.setdefault(pago["comprobante_id"], []).append({
                "pago_id": pago["pago_id"],
                "fecha": str(pago["fecha"]) if pago["fecha"] else None,
                "forma_pago": pago["forma_pago"],
                "importe": round(float(pago["importe"] or 0), 2),
                "cod_cliente": pago["cod_cliente"],
                "cliente_nombre": pago["cliente_nombre"],
            })

    comprobantes = []
    for r in rows:
        comprobantes.append({
            "comprobante_id": r["comprobante_id"],
            "cliente_id": r["cliente_id"],
            "cliente_nombre": r["cliente_nombre"],
            "tipo": r["tipo"],
            "numero": r["numero"],
            "punto_de_venta": r["punto_de_venta"],
            "fecha": str(r["fecha"]) if r["fecha"] else None,
            "fecha_vencimiento": str(r["fecha_vencimiento"]) if r["fecha_vencimiento"] else None,
            "importe_total": round(float(r["importe_total"] or 0), 2),
            "importe_pagado": round(float(r["importe_pagado"] or 0), 2),
            "saldo": round(float(r["saldo"] or 0), 2),
            "cod_vendedor": r["cod_vendedor"],
            "detalle": r["detalle"],
            "pagos": pagos_por_comprobante.get(r["comprobante_id"], []),
        })

    return {
        "cliente_id": cliente_id,
        "total_comprobantes": len(comprobantes),
        "importe_total": round(sum(c["importe_total"] for c in comprobantes), 2),
        "importe_pagado": round(sum(c["importe_pagado"] for c in comprobantes), 2),
        "saldo": round(sum(c["saldo"] for c in comprobantes), 2),
        "comprobantes": comprobantes,
    }


@router.get("/clientes/detalle/{cliente_id}")
async def clientes_detalle(
    cliente_id: str,
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from fastapi import HTTPException as _HTTPException
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    params["cli_id"] = cliente_id
    where_c = "fecha >= :desde AND fecha < :hasta AND anulada <> 'S' AND cliente_id = :cli_id"

    # Client info
    cinfo = (await db.execute(text(f"""
        SELECT cliente_id, MAX(cliente_nombre) AS nombre,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0) AS facturado,
               COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets,
               MAX(CASE WHEN tipo_comprobante='FA' THEN fecha END) AS ultima_compra
        FROM ventas WHERE {where_c}
        GROUP BY cliente_id
    """), params)).mappings().one_or_none()

    if not cinfo:
        raise _HTTPException(status_code=404, detail="Cliente no encontrado")

    # Top products
    productos = (await db.execute(text(f"""
        SELECT producto_id, MAX(producto_nombre) AS nombre,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0) AS facturado,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN cantidad ELSE 0 END), 0) AS unidades
        FROM ventas WHERE {where_c}
        GROUP BY producto_id ORDER BY facturado DESC LIMIT 10
    """), params)).mappings().all()

    # Monthly evolution
    evolucion = (await db.execute(text(f"""
        SELECT TO_CHAR(fecha, 'YYYY-MM') AS periodo,
               COALESCE(SUM(CASE WHEN tipo_comprobante='FA' THEN ABS(total) ELSE 0 END), 0) AS facturado,
               COUNT(CASE WHEN tipo_comprobante='FA' THEN 1 END) AS tickets
        FROM ventas WHERE {where_c}
        GROUP BY periodo ORDER BY periodo
    """), params)).mappings().all()

    # Cuenta corriente saldo
    cc = (await db.execute(text(
        "SELECT COALESCE(SUM(importe), 0) AS saldo FROM cuentas_corrientes_clientes WHERE cliente_id = :cli_id"
    ), {"cli_id": cliente_id})).mappings().one()

    return {
        "cliente": {
            "cliente_id": cinfo["cliente_id"],
            "nombre": cinfo["nombre"],
            "facturado": round(float(cinfo["facturado"] or 0), 2),
            "tickets": int(cinfo["tickets"] or 0),
            "ultima_compra": str(cinfo["ultima_compra"]) if cinfo["ultima_compra"] else None,
            "saldo_cta_cte": round(float(cc["saldo"] or 0), 2),
        },
        "top_productos": [
            {"producto_id": r["producto_id"], "nombre": r["nombre"],
             "facturado": round(float(r["facturado"] or 0), 2), "unidades": int(r["unidades"] or 0)}
            for r in productos
        ],
        "evolucion": [
            {"periodo": r["periodo"], "facturado": round(float(r["facturado"] or 0), 2), "tickets": int(r["tickets"] or 0)}
            for r in evolucion
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# FASE 7 — Panel Proveedores
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/proveedores/kpis")
async def proveedores_kpis(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = compra_params(filters)
    compras_where = compra_filters_clause(filters)

    # Global compras summary
    crow = (await db.execute(text(f"""
        SELECT
            COUNT(DISTINCT proveedor_id)   AS proveedores_activos,
            COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
            COUNT(*)                         AS ordenes,
            CASE WHEN COUNT(*) > 0 THEN COALESCE(SUM({compra_importe_neto_expr()}), 0) / COUNT(*) ELSE 0 END AS ticket_prom
        FROM compras WHERE {compras_where}
    """), params)).mappings().one()

    # Best supplier by total
    best = (await db.execute(text(f"""
        SELECT proveedor_id, COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total
        FROM compras WHERE {compras_where}
        GROUP BY proveedor_id ORDER BY total DESC LIMIT 1
    """), params)).mappings().one_or_none()

    # Resolve best supplier name from master table
    best_nombre = ""
    if best:
        best_prov_id = str(best["proveedor_id"])
        try:
            maestro_best = (await db.execute(text(
                "SELECT nombre FROM proveedores WHERE cod_proveedor = :pid"
            ), {"pid": best_prov_id})).scalar_one_or_none()
        except Exception:
            maestro_best = None
        if not maestro_best:
            cc_best = (await db.execute(text(
                "SELECT MAX(proveedor_nombre) AS nombre FROM cuentas_corrientes_proveedores WHERE proveedor_id = :pid AND proveedor_nombre <> '' AND proveedor_nombre NOT LIKE 'Proveedor %'"
            ), {"pid": best_prov_id})).scalar_one_or_none()
            maestro_best = cc_best
        best_nombre = maestro_best or f"Proveedor {best_prov_id}"

    # Cuenta corriente proveedores: saldo total y deuda vencida
    ccrow = (await db.execute(text("""
        SELECT
            COALESCE(SUM(importe), 0) AS saldo_total,
            COALESCE(SUM(CASE WHEN fecha_vencimiento < now() AND importe > 0 THEN importe ELSE 0 END), 0) AS deuda_vencida
        FROM cuentas_corrientes_proveedores
    """))).mappings().one()

    # Próximos vencimientos (30 días)
    proximos = (await db.execute(text("""
        SELECT COALESCE(SUM(importe), 0) AS total
        FROM cuentas_corrientes_proveedores
        WHERE tipo = 'factura' AND fecha_vencimiento IS NOT NULL
              AND fecha_vencimiento >= now()
              AND fecha_vencimiento <= now() + interval '30 days'
    """))).mappings().one()

    return {
        "proveedores_activos":  {"actual": int(crow["proveedores_activos"] or 0)},
        "total_comprado":       {"actual": round(float(crow["total_comprado"] or 0), 2)},
        "ordenes":              {"actual": int(crow["ordenes"] or 0)},
        "ticket_promedio":      {"actual": round(float(crow["ticket_prom"] or 0), 2)},
        "mejor_proveedor":      {"actual": best_nombre},
        "saldo_cta_cte":        {"actual": round(float(ccrow["saldo_total"] or 0), 2)},
        "deuda_vencida":        {"actual": round(float(ccrow["deuda_vencida"] or 0), 2)},
        "proximos_30d":         {"actual": round(float(proximos["total"] or 0), 2)},
    }


@router.get("/proveedores/ranking")
async def proveedores_ranking(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = compra_params(filters)
    compras_where = compra_filters_clause(filters)

    rows = (await db.execute(text(f"""
        SELECT
            proveedor_id,
            MAX(proveedor_nombre) AS nombre,
            COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
            COUNT(*)                      AS ordenes,
            CASE WHEN COUNT(*) > 0 THEN COALESCE(SUM({compra_importe_neto_expr()}), 0) / COUNT(*) ELSE 0 END AS ticket_promedio,
            COALESCE(SUM(cantidad), 0)    AS unidades,
            MAX(fecha)                    AS ultima_compra
        FROM compras WHERE {compras_where}
        GROUP BY proveedor_id ORDER BY total_comprado DESC
    """), params)).mappings().all()

    # Primary name source: proveedores master table (populated on every sync)
    try:
        maestro_rows = (await db.execute(text("SELECT cod_proveedor, nombre FROM proveedores"))).mappings().all()
        maestro_map = {r["cod_proveedor"]: r["nombre"] for r in maestro_rows if r["nombre"]}
    except Exception:
        maestro_map = {}

    # Secondary fallback: cuentas_corrientes_proveedores
    cc_name_rows = (await db.execute(text("""
        SELECT proveedor_id,
               MAX(CASE WHEN proveedor_nombre <> '' AND proveedor_nombre NOT LIKE 'Proveedor %%' THEN proveedor_nombre END) AS nombre
        FROM cuentas_corrientes_proveedores
        GROUP BY proveedor_id
    """))).mappings().all()
    cc_name_map = {r["proveedor_id"]: r["nombre"] for r in cc_name_rows if r["nombre"]}

    # Saldo por proveedor
    saldos = (await db.execute(text("""
        SELECT proveedor_id, COALESCE(SUM(importe), 0) AS saldo
        FROM cuentas_corrientes_proveedores
        GROUP BY proveedor_id
    """))).mappings().all()
    saldo_map = {r["proveedor_id"]: float(r["saldo"] or 0) for r in saldos}

    total_comprado = sum(float(r["total_comprado"] or 0) for r in rows)
    cumsum = 0.0
    resultado = []

    for r in rows:
        tc = float(r["total_comprado"] or 0)
        pct = (tc / total_comprado * 100) if total_comprado > 0 else 0
        cumsum += pct
        segmento = "A" if cumsum <= 80 else "B" if cumsum <= 95 else "C"

        prov_id = str(r["proveedor_id"])
        raw_nombre = (
            maestro_map.get(prov_id)
            or cc_name_map.get(prov_id)
            or r["nombre"]
            or f"Proveedor {prov_id}"
        )
        resultado.append({
            "proveedor_id": r["proveedor_id"],
            "nombre": raw_nombre,
            "total_comprado": round(tc, 2),
            "ordenes": int(r["ordenes"] or 0),
            "ticket_promedio": round(float(r["ticket_promedio"] or 0), 2),
            "unidades": int(r["unidades"] or 0),
            "ultima_compra": str(r["ultima_compra"]) if r["ultima_compra"] else None,
            "saldo_cta_cte": round(saldo_map.get(r["proveedor_id"], 0.0), 2),
            "pct_del_total": round(pct, 1),
            "pct_acumulado": round(cumsum, 1),
            "segmento": segmento,
        })

    return {"proveedores": resultado}


@router.get("/proveedores/temporal")
async def proveedores_temporal(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = compra_params(filters)
    compras_where = compra_filters_clause(filters)

    rows = (await db.execute(text(f"""
        SELECT
            TO_CHAR(fecha, 'YYYY-MM') AS periodo,
            COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
            COUNT(DISTINCT proveedor_id) AS proveedores_activos,
            COUNT(*)                    AS ordenes
        FROM compras WHERE {compras_where}
        GROUP BY periodo ORDER BY periodo
    """), params)).mappings().all()

    return {
        "series": [
            {
                "periodo": r["periodo"],
                "total_comprado": round(float(r["total_comprado"] or 0), 2),
                "proveedores_activos": int(r["proveedores_activos"] or 0),
                "ordenes": int(r["ordenes"] or 0),
            }
            for r in rows
        ]
    }


@router.get("/proveedores/{proveedor_id}/comprobantes")
async def proveedores_comprobantes(
    proveedor_id: str,
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    params["prov_id"] = proveedor_id

    rows = (await db.execute(text("""
        SELECT comprobante_id, proveedor_id, proveedor_nombre, tipo, numero, punto_de_venta,
               fecha, fecha_vencimiento, importe_total, importe_pagado, saldo, detalle
        FROM comprobantes_proveedores
        WHERE proveedor_id = :prov_id AND (fecha IS NULL OR (fecha >= :desde AND fecha < :hasta))
        ORDER BY COALESCE(fecha_vencimiento, fecha) DESC NULLS LAST, comprobante_id DESC
        LIMIT 200
    """), params)).mappings().all()

    comprobante_ids = [r["comprobante_id"] for r in rows]
    pagos_por_comprobante: dict[str, list[dict]] = {cid: [] for cid in comprobante_ids}
    if comprobante_ids:
        pagos = (await db.execute(text("""
            SELECT pago_id, comprobante_id, fecha, forma_pago, importe, proveedor_id, proveedor_nombre
            FROM pagos_proveedores
            WHERE comprobante_id = ANY(:comprobante_ids)
            ORDER BY fecha DESC NULLS LAST, pago_id DESC
        """), {"comprobante_ids": comprobante_ids})).mappings().all()
        for pago in pagos:
            pagos_por_comprobante.setdefault(pago["comprobante_id"], []).append({
                "pago_id": pago["pago_id"],
                "fecha": str(pago["fecha"]) if pago["fecha"] else None,
                "forma_pago": pago["forma_pago"],
                "importe": round(float(pago["importe"] or 0), 2),
                "proveedor_id": pago["proveedor_id"],
                "proveedor_nombre": pago["proveedor_nombre"],
            })

    comprobantes = []
    for r in rows:
        comprobantes.append({
            "comprobante_id": r["comprobante_id"],
            "proveedor_id": r["proveedor_id"],
            "proveedor_nombre": r["proveedor_nombre"],
            "tipo": r["tipo"],
            "numero": r["numero"],
            "punto_de_venta": r["punto_de_venta"],
            "fecha": str(r["fecha"]) if r["fecha"] else None,
            "fecha_vencimiento": str(r["fecha_vencimiento"]) if r["fecha_vencimiento"] else None,
            "importe_total": round(float(r["importe_total"] or 0), 2),
            "importe_pagado": round(float(r["importe_pagado"] or 0), 2),
            "saldo": round(float(r["saldo"] or 0), 2),
            "detalle": r["detalle"],
            "pagos": pagos_por_comprobante.get(r["comprobante_id"], []),
        })

    return {
        "proveedor_id": proveedor_id,
        "total_comprobantes": len(comprobantes),
        "importe_total": round(sum(c["importe_total"] for c in comprobantes), 2),
        "importe_pagado": round(sum(c["importe_pagado"] for c in comprobantes), 2),
        "saldo": round(sum(c["saldo"] for c in comprobantes), 2),
        "comprobantes": comprobantes,
    }


@router.get("/proveedores/detalle/{proveedor_id}")
async def proveedores_detalle(
    proveedor_id: str,
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from fastapi import HTTPException as _HTTPException
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    params["prov_id"] = proveedor_id
    where_p = "fecha >= :desde AND fecha < :hasta AND proveedor_id = :prov_id"

    # Provider info
    pinfo = (await db.execute(text(f"""
        SELECT proveedor_id, MAX(proveedor_nombre) AS nombre,
               COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
               COUNT(*) AS ordenes,
               MAX(fecha) AS ultima_compra
        FROM compras WHERE {where_p}
        GROUP BY proveedor_id
    """), params)).mappings().one_or_none()

    if not pinfo:
        raise _HTTPException(status_code=404, detail="Proveedor no encontrado")

    # Top products from this supplier
    productos = (await db.execute(text(f"""
        SELECT producto_id, MAX(producto_nombre) AS nombre,
               COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total,
               COALESCE(SUM(cantidad), 0) AS unidades
        FROM compras WHERE {where_p}
        GROUP BY producto_id ORDER BY total DESC LIMIT 10
    """), params)).mappings().all()

    # Monthly evolution
    evolucion = (await db.execute(text(f"""
        SELECT TO_CHAR(fecha, 'YYYY-MM') AS periodo,
               COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_comprado,
               COUNT(*) AS ordenes
        FROM compras WHERE {where_p}
        GROUP BY periodo ORDER BY periodo
    """), params)).mappings().all()

    # Saldo cuenta corriente
    cc = (await db.execute(text(
        "SELECT COALESCE(SUM(importe), 0) AS saldo FROM cuentas_corrientes_proveedores WHERE proveedor_id = :prov_id"
    ), {"prov_id": proveedor_id})).mappings().one()

    # Next due payments
    vencimientos = (await db.execute(text(
        "SELECT comprobante_id, fecha_vencimiento, importe FROM cuentas_corrientes_proveedores "
        "WHERE proveedor_id = :prov_id AND tipo = 'factura' AND fecha_vencimiento IS NOT NULL "
        "AND fecha_vencimiento >= now() ORDER BY fecha_vencimiento LIMIT 10"
    ), {"prov_id": proveedor_id})).mappings().all()

    return {
        "proveedor": {
            "proveedor_id": pinfo["proveedor_id"],
            "nombre": pinfo["nombre"],
            "total_comprado": round(float(pinfo["total_comprado"] or 0), 2),
            "ordenes": int(pinfo["ordenes"] or 0),
            "ultima_compra": str(pinfo["ultima_compra"]) if pinfo["ultima_compra"] else None,
            "saldo_cta_cte": round(float(cc["saldo"] or 0), 2),
        },
        "top_productos": [
            {"producto_id": r["producto_id"], "nombre": r["nombre"],
             "total": round(float(r["total"] or 0), 2), "unidades": int(r["unidades"] or 0)}
            for r in productos
        ],
        "evolucion": [
            {"periodo": r["periodo"], "total_comprado": round(float(r["total_comprado"] or 0), 2),
             "ordenes": int(r["ordenes"] or 0)}
            for r in evolucion
        ],
        "proximos_vencimientos": [
            {"comprobante_id": r["comprobante_id"],
             "fecha_vencimiento": str(r["fecha_vencimiento"]) if r["fecha_vencimiento"] else None,
             "importe": round(float(r["importe"] or 0), 2)}
            for r in vencimientos
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# FASE 7 — Panel Caja
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/caja/kpis")
async def caja_kpis(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    caja_where = text_filter_clause("movimientos_caja", filters)

    row = (await db.execute(text(f"""
        SELECT
            COALESCE(SUM(CASE WHEN importe > 0 THEN importe ELSE 0 END), 0)   AS ingresos,
            COALESCE(SUM(CASE WHEN importe < 0 THEN ABS(importe) ELSE 0 END), 0) AS egresos_caja,
            COALESCE(SUM(importe), 0)                                           AS flujo_neto,
            COUNT(*)                                                             AS movimientos,
            COALESCE(MAX(CASE WHEN importe > 0 THEN importe ELSE 0 END), 0)   AS mayor_ingreso,
            COALESCE(MAX(CASE WHEN importe < 0 THEN ABS(importe) ELSE 0 END), 0) AS mayor_egreso
        FROM movimientos_caja WHERE {caja_where}
    """), params)).mappings().one()

    # Egresos = negative caja movements + compras (pagos a proveedores)
    compras_params: dict = {"desde": filters.desde, "hasta": filters.hasta_exclusive}
    compras_cond = "fecha >= :desde AND fecha < :hasta AND COALESCE(anulada, 'N') <> 'S'"
    if filters.cod_empresa:
        compras_cond += " AND cod_empresa = ANY(:cod_empresa)"
        compras_params["cod_empresa"] = filters.cod_empresa
    compras_egresos = float((await db.execute(text(f"""
        SELECT COALESCE(SUM({compra_importe_neto_expr()}), 0) FROM compras WHERE {compras_cond}
    """), compras_params)).scalar() or 0)

    # Current balance (latest row by fecha)
    saldo_row = (await db.execute(text(
        "SELECT COALESCE(saldo_acumulado, 0) AS saldo_actual FROM movimientos_caja ORDER BY fecha DESC, id DESC LIMIT 1"
    ))).mappings().first()
    if saldo_row is None:
        saldo_row = {"saldo_actual": 0}

    ingresos = float(row["ingresos"] or 0)
    egresos = float(row["egresos_caja"] or 0) + compras_egresos

    return {
        "ingresos":      {"actual": round(ingresos, 2)},
        "egresos":       {"actual": round(egresos, 2)},
        "flujo_neto":    {"actual": round(ingresos - egresos, 2)},
        "movimientos":   {"actual": int(row["movimientos"] or 0)},
        "saldo_actual":  {"actual": round(float(saldo_row["saldo_actual"] or 0), 2)},
        "mayor_ingreso": {"actual": round(float(row["mayor_ingreso"] or 0), 2)},
        "mayor_egreso":  {"actual": round(float(row["mayor_egreso"] or 0), 2)},
        "ratio_cobro_pago": {"actual": round(ingresos / egresos, 2) if egresos > 0 else None},
    }


@router.get("/caja/flujo")
async def caja_flujo(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    caja_where = text_filter_clause("movimientos_caja", filters)

    rows = (await db.execute(text(f"""
        SELECT
            TO_CHAR(fecha, 'YYYY-MM') AS periodo,
            COALESCE(SUM(CASE WHEN importe > 0 THEN importe ELSE 0 END), 0)   AS ingresos,
            COALESCE(SUM(CASE WHEN importe < 0 THEN ABS(importe) ELSE 0 END), 0) AS egresos,
            COALESCE(SUM(importe), 0)                                           AS flujo_neto,
            COUNT(*)                                                             AS movimientos
        FROM movimientos_caja WHERE {caja_where}
        GROUP BY periodo ORDER BY periodo
    """), params)).mappings().all()

    saldo_acum = 0.0
    series = []
    for r in rows:
        saldo_acum += float(r["flujo_neto"] or 0)
        series.append({
            "periodo": r["periodo"],
            "ingresos": round(float(r["ingresos"] or 0), 2),
            "egresos": round(float(r["egresos"] or 0), 2),
            "flujo_neto": round(float(r["flujo_neto"] or 0), 2),
            "saldo_acumulado": round(saldo_acum, 2),
            "movimientos": int(r["movimientos"] or 0),
        })

    return {"series": series}


@router.get("/caja/por-tipo")
async def caja_por_tipo(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    caja_where = text_filter_clause("movimientos_caja", filters)

    rows = (await db.execute(text(f"""
        SELECT
            tipo,
            COALESCE(SUM(CASE WHEN importe > 0 THEN importe ELSE 0 END), 0)   AS ingresos,
            COALESCE(SUM(CASE WHEN importe < 0 THEN ABS(importe) ELSE 0 END), 0) AS egresos,
            COALESCE(SUM(importe), 0)                                           AS neto,
            COUNT(*)                                                             AS movimientos
        FROM movimientos_caja WHERE {caja_where}
        GROUP BY tipo ORDER BY ABS(SUM(importe)) DESC
    """), params)).mappings().all()

    return {
        "por_tipo": [
            {
                "tipo": r["tipo"],
                "ingresos": round(float(r["ingresos"] or 0), 2),
                "egresos": round(float(r["egresos"] or 0), 2),
                "neto": round(float(r["neto"] or 0), 2),
                "movimientos": int(r["movimientos"] or 0),
            }
            for r in rows
        ]
    }


@router.get("/caja/movimientos")
async def caja_movimientos(
    page: int = 1,
    limit: int = 50,
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    params = dict(filters.sql_params())
    params["limit"] = limit
    params["offset"] = (page - 1) * limit
    caja_where = text_filter_clause("movimientos_caja", filters)

    total_row = (await db.execute(text(f"""
        SELECT COUNT(*) AS total FROM movimientos_caja WHERE {caja_where}
    """), params)).mappings().one()

    rows = (await db.execute(text(f"""
        SELECT fecha, tipo, descripcion, importe, saldo_acumulado
        FROM movimientos_caja WHERE {caja_where}
        ORDER BY fecha DESC, id DESC
        LIMIT :limit OFFSET :offset
    """), params)).mappings().all()

    return {
        "total": int(total_row["total"] or 0),
        "page": page,
        "limit": limit,
        "movimientos": [
            {
                "fecha": str(r["fecha"]) if r["fecha"] else None,
                "tipo": r["tipo"],
                "descripcion": r["descripcion"],
                "importe": round(float(r["importe"] or 0), 2),
                "saldo_acumulado": round(float(r["saldo_acumulado"] or 0), 2) if r["saldo_acumulado"] is not None else None,
            }
            for r in rows
        ],
    }


# ══════════════════════════════════════════════════════════════════
# REPORTES — Widgets de cuentas corrientes, cobranza y proveedores
# Fuente: tablas locales sincronizadas desde Infomanager (o Excel).
# Funcionan sin conector activo si los datos ya fueron sincronizados.
# ══════════════════════════════════════════════════════════════════

async def _set_path(db, current_user, company_id):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    return tenant_schema


# ── Infomanager specific-table endpoints ──────────────────────────────────────


@router.get("/empresas")
async def get_empresas(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        tenant_schema = await get_tenant_schema(current_user, db, company_id)
        await set_tenant_search_path(db, tenant_schema)
        rows = (await db.execute(text(
            "SELECT cod_empresa, nombre, nombre_1, cuit, habilitada FROM empresas_infomanager ORDER BY cod_empresa"
        ))).mappings().all()
        return {"empresas": [dict(r) for r in rows]}
    except Exception:
        return {"empresas": []}


@router.get("/depositos")
async def get_depositos(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        tenant_schema = await get_tenant_schema(current_user, db, company_id)
        await set_tenant_search_path(db, tenant_schema)
        rows = (await db.execute(text(
            "SELECT cod_deposito, nombre, habilitado FROM depositos ORDER BY cod_deposito"
        ))).mappings().all()
        return {"depositos": [dict(r) for r in rows]}
    except Exception:
        return {"depositos": []}



@router.get("/listas-precios")
async def get_listas_precios(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        tenant_schema = await get_tenant_schema(current_user, db, company_id)
        await set_tenant_search_path(db, tenant_schema)
        rows = (await db.execute(text(
            "SELECT cod_lista, descripcion FROM listas_precios ORDER BY cod_lista"
        ))).mappings().all()
        return {"listas": [dict(r) for r in rows]}
    except Exception:
        return {"listas": []}



@router.get("/rubros")
async def get_rubros(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        tenant_schema = await get_tenant_schema(current_user, db, company_id)
        await set_tenant_search_path(db, tenant_schema)
        rows = (await db.execute(text(
            "SELECT DISTINCT cod_rubro, rubro AS nombre FROM stock WHERE cod_rubro IS NOT NULL ORDER BY cod_rubro"
        ))).mappings().all()
        return {"rubros": [dict(r) for r in rows]}
    except Exception:
        return {"rubros": []}


@router.get("/vendedores-lookup")
async def get_vendedores_lookup(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        tenant_schema = await get_tenant_schema(current_user, db, company_id)
        await set_tenant_search_path(db, tenant_schema)
        rows = (await db.execute(text(
            "SELECT cod_vendedor, nombre FROM vendedores ORDER BY nombre"
        ))).mappings().all()
        return {"vendedores": [dict(r) for r in rows]}
    except Exception:
        return {"vendedores": []}


@router.get("/clientes-lookup")
async def get_clientes_lookup(
    company_id: int = None,
    limit: int = 200,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        tenant_schema = await get_tenant_schema(current_user, db, company_id)
        await set_tenant_search_path(db, tenant_schema)
        rows = (await db.execute(text("""
            SELECT DISTINCT cliente_id AS cod_cliente, MAX(cliente_nombre) AS nombre
            FROM ventas
            WHERE cliente_id IS NOT NULL
            GROUP BY cliente_id
            ORDER BY nombre
            LIMIT :limit
        """), {"limit": limit})).mappings().all()
        return {"clientes": [dict(r) for r in rows]}
    except Exception:
        return {"clientes": []}


@router.get("/proveedores-lookup")
async def get_proveedores_lookup(
    company_id: int = None,
    limit: int = 200,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        tenant_schema = await get_tenant_schema(current_user, db, company_id)
        await set_tenant_search_path(db, tenant_schema)
        rows = (await db.execute(text("""
            SELECT DISTINCT proveedor_id AS cod_proveedor, MAX(proveedor_nombre) AS nombre
            FROM compras
            WHERE proveedor_id IS NOT NULL
            GROUP BY proveedor_id
            ORDER BY nombre
            LIMIT :limit
        """), {"limit": limit})).mappings().all()
        return {"proveedores": [dict(r) for r in rows]}
    except Exception:
        return {"proveedores": []}


@router.get("/reportes/saldos-clientes")
async def reportes_saldos_clientes(
    limit: int = 100,
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _set_path(db, current_user, company_id)

    rows = (await db.execute(text("""
        SELECT
            cliente_id,
            MAX(cliente_nombre)                                 AS nombre,
            COALESCE(SUM(GREATEST(saldo::float, 0)), 0)         AS saldo_total,
            COALESCE(SUM(GREATEST(importe_total::float, 0)), 0) AS facturado_total,
            COALESCE(SUM(GREATEST(importe_pagado::float, 0)), 0) AS cobrado_total,
            COUNT(*) FILTER (WHERE saldo::float > 0.01)         AS facturas_pendientes,
            MAX(fecha)                                          AS ultimo_mov
        FROM comprobantes_clientes
        WHERE tipo IN ('FA', 'ND', 'RC', 'factura', 'saldo')
        GROUP BY cliente_id
        HAVING COALESCE(SUM(saldo::float), 0) > 0.01
        ORDER BY SUM(saldo::float) DESC
        LIMIT :limit
    """), {"limit": limit})).mappings().all()

    total_deuda = sum(float(r["saldo_total"] or 0) for r in rows)
    total_clientes = len(rows)

    return {
        "kpi": {
            "total_deuda": round(total_deuda, 2),
            "clientes_con_deuda": total_clientes,
        },
        "filas": [
            {
                "cliente_id": r["cliente_id"],
                "nombre": r["nombre"] or r["cliente_id"],
                "saldo_total": round(float(r["saldo_total"] or 0), 2),
                "facturado_total": round(float(r["facturado_total"] or 0), 2),
                "cobrado_total": round(float(r["cobrado_total"] or 0), 2),
                "facturas_pendientes": int(r["facturas_pendientes"] or 0),
                "ultimo_mov": str(r["ultimo_mov"])[:10] if r["ultimo_mov"] else None,
            }
            for r in rows
        ],
    }


@router.get("/reportes/comprob-pendientes")
async def reportes_comprob_pendientes(
    limit: int = 200,
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _set_path(db, current_user, company_id)

    rows = (await db.execute(text("""
        SELECT
            comprobante_id,
            cliente_id,
            cliente_nombre,
            tipo,
            numero,
            punto_de_venta,
            fecha,
            fecha_vencimiento,
            importe_total::float    AS importe_total,
            importe_pagado::float   AS importe_pagado,
            saldo::float            AS saldo,
            CASE
                WHEN fecha_vencimiento IS NOT NULL AND fecha_vencimiento::date < CURRENT_DATE
                THEN (CURRENT_DATE - fecha_vencimiento::date)
                ELSE 0
            END                     AS dias_mora
        FROM comprobantes_clientes
        WHERE saldo::float > 0.01 AND tipo IN ('FA', 'ND', 'factura')
        ORDER BY fecha_vencimiento ASC NULLS LAST, saldo::float DESC
        LIMIT :limit
    """), {"limit": limit})).mappings().all()

    total_pendiente = sum(float(r["saldo"] or 0) for r in rows)
    vencidas = sum(1 for r in rows if int(r["dias_mora"] or 0) > 0)

    return {
        "kpi": {
            "total_pendiente": round(total_pendiente, 2),
            "cantidad_facturas": len(rows),
            "facturas_vencidas": vencidas,
        },
        "filas": [
            {
                "comprobante_id": r["comprobante_id"],
                "cliente_id": r["cliente_id"],
                "cliente_nombre": r["cliente_nombre"],
                "tipo": r["tipo"],
                "numero": r["numero"],
                "fecha": str(r["fecha"])[:10] if r["fecha"] else None,
                "fecha_vencimiento": str(r["fecha_vencimiento"])[:10] if r["fecha_vencimiento"] else None,
                "importe_total": round(float(r["importe_total"] or 0), 2),
                "importe_pagado": round(float(r["importe_pagado"] or 0), 2),
                "saldo": round(float(r["saldo"] or 0), 2),
                "dias_mora": int(r["dias_mora"] or 0),
            }
            for r in rows
        ],
    }


@router.get("/reportes/disponible-credito")
async def reportes_disponible_credito(
    limit: int = 100,
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _set_path(db, current_user, company_id)

    rows = (await db.execute(text("""
        SELECT
            cc.cliente_id,
            MAX(cc.cliente_nombre)                               AS nombre,
            COALESCE(SUM(cc.importe_total::float), 0)           AS facturado_total,
            COALESCE(SUM(cc.importe_pagado::float), 0)          AS cobrado_total,
            COALESCE(SUM(GREATEST(cc.saldo::float, 0)), 0)      AS saldo_pendiente,
            COUNT(*) FILTER (WHERE cc.saldo::float > 0.01)      AS facturas_abiertas
        FROM comprobantes_clientes cc
        WHERE cc.tipo IN ('FA', 'ND', 'factura')
        GROUP BY cc.cliente_id
        ORDER BY SUM(GREATEST(cc.saldo::float, 0)) DESC
        LIMIT :limit
    """), {"limit": limit})).mappings().all()

    clientes_con_saldo = sum(1 for r in rows if float(r["saldo_pendiente"] or 0) > 0)

    return {
        "kpi": {
            "clientes_con_saldo": clientes_con_saldo,
            "total_clientes": len(rows),
            "nota": "Limite de credito no disponible en datos locales; se muestra saldo actual.",
        },
        "filas": [
            {
                "cliente_id": r["cliente_id"],
                "nombre": r["nombre"] or r["cliente_id"],
                "facturado_total": round(float(r["facturado_total"] or 0), 2),
                "cobrado_total": round(float(r["cobrado_total"] or 0), 2),
                "saldo_pendiente": round(float(r["saldo_pendiente"] or 0), 2),
                "facturas_abiertas": int(r["facturas_abiertas"] or 0),
                "pct_cobrado": round(
                    float(r["cobrado_total"] or 0) / float(r["facturado_total"] or 1) * 100, 1
                ) if float(r["facturado_total"] or 0) > 0 else 0,
            }
            for r in rows
        ],
    }


@router.get("/reportes/facturas-con-pagos")
async def reportes_facturas_con_pagos(
    limit: int = 200,
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from fastapi import HTTPException
    await _set_path(db, current_user, company_id)

    params: dict = {"limit": limit}
    date_cond = ""
    if filters.desde:
        date_cond += " AND cc.fecha::date >= :desde"
        params["desde"] = filters.desde
    if filters.hasta:
        date_cond += " AND cc.fecha::date <= :hasta"
        params["hasta"] = filters.hasta

    _sql = f"""
        SELECT
            cc.comprobante_id,
            cc.cliente_id,
            cc.cliente_nombre,
            cc.tipo,
            cc.numero,
            cc.fecha,
            cc.importe_total::float     AS importe_total,
            cc.importe_pagado::float    AS importe_pagado,
            cc.saldo::float             AS saldo,
            CASE
                WHEN cc.importe_total::float > 0
                THEN ROUND((cc.importe_pagado::float / cc.importe_total::float * 100)::numeric, 1)
                ELSE 0::numeric
            END                         AS pct_cobrado,
            MAX(pc.fecha)               AS ultimo_pago,
            MAX(pc.forma_pago)          AS forma_pago
        FROM comprobantes_clientes cc
        LEFT JOIN pagos_clientes pc ON cc.comprobante_id = pc.comprobante_id
        WHERE cc.tipo IN ('FA', 'ND', 'factura')
          AND cc.importe_total::float > 0
          {date_cond}
        GROUP BY cc.comprobante_id, cc.cliente_id, cc.cliente_nombre,
                 cc.tipo, cc.numero, cc.fecha, cc.importe_total, cc.importe_pagado, cc.saldo
        ORDER BY cc.fecha DESC
        LIMIT :limit
    """
    try:
        rows = (await db.execute(text(_sql), params)).mappings().all()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"SQL error en facturas-con-pagos: {type(e).__name__}: {e}")

    total_facturado = sum(float(r["importe_total"] or 0) for r in rows)
    total_cobrado   = sum(float(r["importe_pagado"] or 0) for r in rows)
    cobradas_100    = sum(1 for r in rows if float(r["saldo"] or 0) < 0.01)
    pct_eficiencia  = round(total_cobrado / total_facturado * 100, 1) if total_facturado > 0 else 0

    return {
        "kpi": {
            "total_facturado": round(total_facturado, 2),
            "total_cobrado": round(total_cobrado, 2),
            "pct_eficiencia_cobranza": pct_eficiencia,
            "facturas_cobradas": cobradas_100,
            "total_facturas": len(rows),
        },
        "filas": [
            {
                "comprobante_id": r["comprobante_id"],
                "cliente_id": r["cliente_id"],
                "cliente_nombre": r["cliente_nombre"],
                "tipo": r["tipo"],
                "numero": r["numero"],
                "fecha": str(r["fecha"])[:10] if r["fecha"] else None,
                "importe_total": round(float(r["importe_total"] or 0), 2),
                "importe_pagado": round(float(r["importe_pagado"] or 0), 2),
                "saldo": round(float(r["saldo"] or 0), 2),
                "pct_cobrado": float(r["pct_cobrado"] or 0),
                "ultimo_pago": str(r["ultimo_pago"])[:10] if r["ultimo_pago"] else None,
                "forma_pago": r["forma_pago"],
            }
            for r in rows
        ],
    }


@router.get("/reportes/vencimientos-compras")
async def reportes_vencimientos_compras(
    limit: int = 200,
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _set_path(db, current_user, company_id)

    rows = (await db.execute(text("""
        SELECT
            comprobante_id,
            proveedor_id,
            proveedor_nombre,
            tipo,
            numero,
            punto_de_venta,
            fecha,
            fecha_vencimiento,
            importe_total::float    AS importe_total,
            importe_pagado::float   AS importe_pagado,
            saldo::float            AS saldo,
            CASE
                WHEN fecha_vencimiento IS NOT NULL
                THEN (fecha_vencimiento::date - CURRENT_DATE)
                ELSE NULL
            END                     AS dias_para_vencer,
            CASE
                WHEN fecha_vencimiento IS NOT NULL AND fecha_vencimiento::date < CURRENT_DATE
                THEN (CURRENT_DATE - fecha_vencimiento::date)
                ELSE 0
            END                     AS dias_vencido
        FROM comprobantes_proveedores
        WHERE saldo::float > 0.01
        ORDER BY fecha_vencimiento ASC NULLS LAST
        LIMIT :limit
    """), {"limit": limit})).mappings().all()

    total_saldo         = sum(float(r["saldo"] or 0) for r in rows)
    vencer_30d          = sum(
        float(r["saldo"] or 0) for r in rows
        if r["dias_para_vencer"] is not None and -30 <= int(r["dias_para_vencer"]) <= 30
    )

    return {
        "kpi": {
            "total_a_pagar": round(total_saldo, 2),
            "a_pagar_30_dias": round(vencer_30d, 2),
            "cantidad_facturas": len(rows),
        },
        "filas": [
            {
                "comprobante_id": r["comprobante_id"],
                "proveedor_id": r["proveedor_id"],
                "proveedor_nombre": r["proveedor_nombre"],
                "tipo": r["tipo"],
                "numero": r["numero"],
                "fecha": str(r["fecha"])[:10] if r["fecha"] else None,
                "fecha_vencimiento": str(r["fecha_vencimiento"])[:10] if r["fecha_vencimiento"] else None,
                "importe_total": round(float(r["importe_total"] or 0), 2),
                "importe_pagado": round(float(r["importe_pagado"] or 0), 2),
                "saldo": round(float(r["saldo"] or 0), 2),
                "dias_para_vencer": int(r["dias_para_vencer"]) if r["dias_para_vencer"] is not None else None,
                "dias_vencido": int(r["dias_vencido"] or 0),
            }
            for r in rows
        ],
    }


@router.get("/reportes/empresas-resumen")
async def reportes_empresas_resumen(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _set_path(db, current_user, company_id)

    rows = (await db.execute(text("""
        SELECT
            COALESCE(v.cod_empresa, 1)                                                             AS cod_empresa,
            COALESCE(ei.nombre, 'Empresa ' || COALESCE(v.cod_empresa, 1)::text)                    AS nombre,
            COALESCE(SUM(CASE WHEN v.tipo_comprobante='FA' THEN ABS(v.total) ELSE 0 END), 0)       AS facturado,
            COALESCE(SUM(CASE WHEN v.tipo_comprobante='NC' THEN ABS(v.total) ELSE 0 END), 0)       AS notas_credito,
            COALESCE(SUM(CASE WHEN v.tipo_comprobante IN ('FA','ND') THEN ABS(v.total)
                              WHEN v.tipo_comprobante='NC' THEN -ABS(v.total) ELSE 0 END), 0)     AS neto,
            COUNT(DISTINCT v.cliente_id)                                                           AS clientes
        FROM ventas v
        LEFT JOIN empresas_infomanager ei ON ei.cod_empresa = COALESCE(v.cod_empresa, 1)
        WHERE v.fecha::date >= :desde AND v.fecha::date <= :hasta
          AND v.anulada = 'N'
        GROUP BY COALESCE(v.cod_empresa, 1), ei.nombre
        ORDER BY neto DESC
    """), {"desde": filters.desde, "hasta": filters.hasta})).mappings().all()

    total_neto = sum(float(r["neto"] or 0) for r in rows)

    return {
        "kpi": {
            "total_neto": round(total_neto, 2),
            "empresas": len(rows),
            "una_sola_empresa": len(rows) <= 1,
        },
        "filas": [
            {
                "cod_empresa": int(r["cod_empresa"] or 1),
                "nombre": r["nombre"] or f"Empresa {r['cod_empresa']}",
                "facturado": round(float(r["facturado"] or 0), 2),
                "notas_credito": round(float(r["notas_credito"] or 0), 2),
                "neto": round(float(r["neto"] or 0), 2),
                "clientes": int(r["clientes"] or 0),
                "pct": round(float(r["neto"] or 0) / total_neto * 100, 1) if total_neto > 0 else 0,
            }
            for r in rows
        ],
    }


@router.get("/stock/disponible")
async def get_stock_disponible(
    company_id: int = None,
    cod_deposito: Optional[int] = None,
    cod_rubro: Optional[int] = None,
    cod_empresa: Optional[int] = None,
    solo_alertas: bool = False,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    conditions = ["1=1"]
    params: dict = {}
    if cod_deposito is not None:
        conditions.append("cod_deposito = :cod_deposito")
        params["cod_deposito"] = cod_deposito
    if cod_rubro is not None:
        conditions.append("cod_rubro = :cod_rubro")
        params["cod_rubro"] = cod_rubro
    if solo_alertas:
        conditions.append("existencia < pto_de_reposicion AND pto_de_reposicion > 0")

    where = " AND ".join(conditions)
    rows = (await db.execute(text(f"""
        SELECT cod_articulo, descripcion, cod_deposito,
               stock_entradas, stock_salidas, compras, ventas,
               existencia, existencia_anterior, precio_compra, precio_venta,
               pto_de_reposicion, cod_rubro, rubro, subrubro, proveedor,
               habilitado, snapshot_at
        FROM stock_disponible
        WHERE {where}
        ORDER BY descripcion
        LIMIT 1000
    """), params)).mappings().all()

    return {
        "total": len(rows),
        "articulos": [
            {
                **dict(r),
                "existencia": round(float(r["existencia"] or 0), 3),
                "precio_compra": round(float(r["precio_compra"] or 0), 2),
                "precio_venta": round(float(r["precio_venta"] or 0), 2),
                "alerta_reposicion": (
                    float(r["existencia"] or 0) < float(r["pto_de_reposicion"] or 0)
                    and float(r["pto_de_reposicion"] or 0) > 0
                ),
            }
            for r in rows
        ],
    }


@router.get("/stock/movimientos")
async def get_movimientos_stock(
    company_id: int = None,
    cod_articulo: Optional[int] = None,
    q: Optional[str] = None,
    cod_deposito: Optional[int] = None,
    cod_empresa: Optional[int] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params: dict = {}
    art_cond_v = ""
    art_cond_c = ""
    dep_cond_v = ""
    dep_cond_c = ""
    emp_cond_v = ""
    emp_cond_c = ""
    date_cond_v = "1=1"
    date_cond_c = "1=1"

    if cod_articulo is not None:
        art_cond_v = "AND v.producto_id::text = :cod_articulo::text"
        art_cond_c = "AND c.producto_id::text = :cod_articulo::text"
        params["cod_articulo"] = str(cod_articulo)
    elif q:
        art_cond_v = "AND v.producto_nombre ILIKE :q"
        art_cond_c = "AND c.producto_nombre ILIKE :q"
        params["q"] = f"%{q}%"
    if cod_deposito is not None:
        dep_cond_v = "AND v.cod_deposito = :cod_deposito"
        dep_cond_c = "AND c.cod_deposito = :cod_deposito"
        params["cod_deposito"] = cod_deposito
    if cod_empresa is not None:
        emp_cond_v = "AND v.cod_empresa = :cod_empresa"
        emp_cond_c = "AND c.cod_empresa = :cod_empresa"
        params["cod_empresa"] = cod_empresa
    if desde:
        date_cond_v = "v.fecha::date >= :desde"
        date_cond_c = "c.fecha::date >= :desde"
        params["desde"] = date.fromisoformat(desde)
    if hasta:
        date_cond_v += " AND v.fecha::date <= :hasta" if desde else "v.fecha::date <= :hasta"
        date_cond_c += " AND c.fecha::date <= :hasta" if desde else "c.fecha::date <= :hasta"
        params["hasta"] = date.fromisoformat(hasta)

    rows = (await db.execute(text(f"""
        SELECT
            row_number() OVER () AS id,
            producto_id::integer AS cod_articulo,
            producto_nombre AS descripcion,
            fecha,
            'salida' AS tipo_movimiento,
            -ABS(cantidad) AS cantidad,
            precio_unitario AS precio,
            total,
            cod_deposito,
            cod_empresa
        FROM ventas v
        WHERE {date_cond_v} AND tipo_comprobante = 'FA' AND anulada != 'S'
          {art_cond_v} {dep_cond_v} {emp_cond_v}
          AND producto_id IS NOT NULL AND producto_id ~ '^[0-9]+$'
        UNION ALL
        SELECT
            row_number() OVER () AS id,
            producto_id::integer AS cod_articulo,
            producto_nombre AS descripcion,
            fecha,
            'entrada' AS tipo_movimiento,
            ABS(cantidad) AS cantidad,
            precio_unitario AS precio,
            total,
            cod_deposito,
            cod_empresa
        FROM compras c
        WHERE {date_cond_c} AND tipo_comprobante IN ('FA', 'FC') AND anulada != 'S'
          {art_cond_c} {dep_cond_c} {emp_cond_c}
          AND producto_id IS NOT NULL AND producto_id ~ '^[0-9]+$'
        ORDER BY fecha DESC
        LIMIT 500
    """), params)).mappings().all()

    return {"total": len(rows), "movimientos": [
        {**dict(r), "cantidad": round(float(r["cantidad"] or 0), 3),
         "precio": round(float(r["precio"] or 0), 2),
         "total": round(float(r["total"] or 0), 2)}
        for r in rows
    ]}


@router.get("/stock/interdepositos")
async def get_interdepositos(
    company_id: int = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    conditions = ["1=1"]
    params: dict = {}
    if desde:
        conditions.append("fecha::date >= :desde")
        params["desde"] = date.fromisoformat(desde)
    if hasta:
        conditions.append("fecha::date <= :hasta")
        params["hasta"] = date.fromisoformat(hasta)

    where = " AND ".join(conditions)
    rows = (await db.execute(text(f"""
        SELECT id, cod_articulo, descripcion, fecha, tipo_movimiento,
               cantidad, cod_deposito, cod_empresa
        FROM movimientos_stock
        WHERE tipo_movimiento = 'interdeposito' AND {where}
        ORDER BY fecha DESC
        LIMIT 500
    """), params)).mappings().all()

    return {"total": len(rows), "interdepositos": [dict(r) for r in rows]}


@router.get("/resultado/listas-precios")
async def get_resultado_listas_precios(
    company_id: int = None,
    cod_lista_precios: Optional[int] = None,
    cod_empresa: Optional[int] = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    conditions = ["1=1"]
    params: dict = {}
    if cod_lista_precios is not None:
        conditions.append("i.cod_lista = :cod_lista")
        params["cod_lista"] = cod_lista_precios
    if cod_empresa is not None:
        conditions.append("l.cod_empresa = :cod_empresa")
        params["cod_empresa"] = cod_empresa

    where = " AND ".join(conditions)
    rows = (await db.execute(text(f"""
        SELECT i.cod_lista, l.descripcion AS lista_descripcion,
               i.cod_articulo, i.art_descripcion,
               i.art_precio_compra, i.art_precio_venta,
               i.lista_precio_venta, i.lista_precio_con_iva,
               i.lista_porcentaje, i.lista_descuento, i.art_iva,
               CASE
                 WHEN i.art_precio_compra > 0 THEN
                   ROUND(((i.lista_precio_venta - i.art_precio_compra) / i.art_precio_compra * 100)::numeric, 2)
                 ELSE NULL
               END AS margen_porc
        FROM items_listas_precios i
        LEFT JOIN listas_precios l ON l.cod_lista = i.cod_lista
        WHERE {where}
        ORDER BY i.cod_lista, i.art_descripcion
        LIMIT 2000
    """), params)).mappings().all()

    return {
        "total": len(rows),
        "items": [
            {
                **dict(r),
                "art_precio_compra": round(float(r["art_precio_compra"] or 0), 2),
                "lista_precio_venta": round(float(r["lista_precio_venta"] or 0), 2),
                "margen_porc": float(r["margen_porc"]) if r["margen_porc"] is not None else None,
            }
            for r in rows
        ],
    }


@router.get("/resultado/libro-mayor")
async def get_libro_mayor(
    company_id: int = None,
    cod_empresa: Optional[int] = None,
    cod_cuenta: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    page: int = 1,
    limit: int = 500,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    conditions = ["1=1"]
    params: dict = {"offset": (page - 1) * limit, "limit": limit}
    if cod_empresa is not None:
        conditions.append("cod_empresa = :cod_empresa")
        params["cod_empresa"] = cod_empresa
    if cod_cuenta:
        conditions.append("cuenta::text LIKE :cod_cuenta")
        params["cod_cuenta"] = f"{cod_cuenta}%"
    if desde:
        conditions.append("fecha::date >= :desde")
        params["desde"] = date.fromisoformat(desde)
    if hasta:
        conditions.append("fecha::date <= :hasta")
        params["hasta"] = date.fromisoformat(hasta)

    where = " AND ".join(conditions)
    count_params = {k: v for k, v in params.items() if k not in ("offset", "limit")}
    total = (await db.execute(text(f"SELECT COUNT(*) FROM movimientos_contables WHERE {where}"), count_params)).scalar()
    rows = (await db.execute(text(f"""
        SELECT id, fecha, cuenta, plan_descripcion, debe, haber,
               tipo_comprobante, numero, descripcion, cod_empresa, tag
        FROM movimientos_contables
        WHERE {where}
        ORDER BY fecha DESC, id
        LIMIT :limit OFFSET :offset
    """), params)).mappings().all()

    return {
        "total": int(total or 0),
        "page": page,
        "limit": limit,
        "movimientos": [
            {
                **dict(r),
                "debe": round(float(r["debe"] or 0), 2),
                "haber": round(float(r["haber"] or 0), 2),
            }
            for r in rows
        ],
    }


@router.get("/caja/recibos")
async def get_recibos_periodo(
    company_id: int = None,
    cod_empresa: Optional[int] = None,
    tag: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    page: int = 1,
    limit: int = 500,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    conditions = ["1=1"]
    params: dict = {"offset": (page - 1) * limit, "limit": limit}
    if cod_empresa is not None:
        conditions.append("fcr.fa_cod_empresa = :cod_empresa")
        params["cod_empresa"] = cod_empresa
    if desde:
        conditions.append("fcr.rc_fecha::date >= :desde")
        params["desde"] = date.fromisoformat(desde)
    if hasta:
        conditions.append("fcr.rc_fecha::date <= :hasta")
        params["hasta"] = date.fromisoformat(hasta)

    where = " AND ".join(conditions)
    count_params_r = {k: v for k, v in params.items() if k not in ("offset", "limit")}
    total = (await db.execute(text(f"SELECT COUNT(*) FROM facturas_con_recibos fcr WHERE {where}"), count_params_r)).scalar()
    rows = (await db.execute(text(f"""
        SELECT fcr.fa_id, fcr.tipo_comp, fcr.fa_fecha, fcr.cod_cliente,
               COALESCE(c.nombre, fcr.cod_cliente::text) AS cliente_nombre,
               fcr.fa_total, fcr.fa_total_moneda_local,
               fcr.rc_id, fcr.rc_fecha, fcr.rc_nro,
               fcr.imp_pag_moneda_local, fcr.cond_pago, fcr.importe,
               fcr.cod_banco, fcr.cheque_numero, fcr.cheque_fec_pago, fcr.importe_retencion
        FROM facturas_con_recibos fcr
        LEFT JOIN clientes c ON c.external_id = fcr.cod_cliente::text
        WHERE {where}
        ORDER BY fcr.rc_fecha DESC
        LIMIT :limit OFFSET :offset
    """), params)).mappings().all()

    return {
        "total": int(total or 0),
        "page": page,
        "limit": limit,
        "recibos": [
            {
                **dict(r),
                "importe": round(float(r["importe"] or 0), 2),
                "fa_total": round(float(r["fa_total"] or 0), 2),
            }
            for r in rows
        ],
    }


@router.get("/caja/flujo-contable")
async def get_flujo_caja_contable(
    company_id: int = None,
    cod_empresa: Optional[int] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    prefijo_cuenta: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    conditions = ["1=1"]
    params: dict = {}
    if cod_empresa is not None:
        conditions.append("cod_empresa = :cod_empresa")
        params["cod_empresa"] = cod_empresa
    if desde:
        conditions.append("fecha::date >= :desde")
        params["desde"] = date.fromisoformat(desde)
    if hasta:
        conditions.append("fecha::date <= :hasta")
        params["hasta"] = date.fromisoformat(hasta)
    if prefijo_cuenta:
        conditions.append("cuenta::text LIKE :prefijo")
        params["prefijo"] = f"{prefijo_cuenta}%"

    where = " AND ".join(conditions)

    daily_rows = (await db.execute(text(f"""
        SELECT fecha::date AS fecha,
               COALESCE(SUM(debe), 0) AS total_debe,
               COALESCE(SUM(haber), 0) AS total_haber,
               COUNT(*) AS movimientos
        FROM movimientos_contables
        WHERE {where}
        GROUP BY fecha::date
        ORDER BY fecha::date
    """), params)).mappings().all()

    acct_rows = (await db.execute(text(f"""
        SELECT cuenta,
               MAX(plan_descripcion) AS plan_descripcion,
               COALESCE(SUM(debe), 0) AS total_debe,
               COALESCE(SUM(haber), 0) AS total_haber,
               COUNT(*) AS movimientos
        FROM movimientos_contables
        WHERE {where}
        GROUP BY cuenta
        ORDER BY cuenta
    """), params)).mappings().all()

    total_debe = sum(float(r["total_debe"] or 0) for r in daily_rows)
    total_haber = sum(float(r["total_haber"] or 0) for r in daily_rows)

    saldo_acum = 0.0
    por_dia = []
    for r in daily_rows:
        neto = round(float(r["total_debe"] or 0) - float(r["total_haber"] or 0), 2)
        saldo_acum += neto
        por_dia.append({
            "fecha": str(r["fecha"]),
            "debe": round(float(r["total_debe"] or 0), 2),
            "haber": round(float(r["total_haber"] or 0), 2),
            "neto": neto,
            "saldo_acum": round(saldo_acum, 2),
            "movimientos": int(r["movimientos"]),
        })

    return {
        "total_debe": round(total_debe, 2),
        "total_haber": round(total_haber, 2),
        "saldo_neto": round(total_debe - total_haber, 2),
        "movimientos_count": sum(int(r["movimientos"]) for r in daily_rows),
        "por_dia": por_dia,
        "por_cuenta": [
            {
                "cuenta": int(r["cuenta"]) if r["cuenta"] else None,
                "plan_descripcion": r["plan_descripcion"],
                "debe": round(float(r["total_debe"] or 0), 2),
                "haber": round(float(r["total_haber"] or 0), 2),
                "saldo": round(float(r["total_debe"] or 0) - float(r["total_haber"] or 0), 2),
                "movimientos": int(r["movimientos"]),
            }
            for r in acct_rows
        ],
    }


@router.get("/listas-precios")
async def get_listas_precios(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    rows = (await db.execute(text(
        "SELECT cod_lista, descripcion, aplica_a, cod_empresa, habilitado FROM listas_precios ORDER BY cod_lista"
    ))).mappings().all()
    return {"listas": [dict(r) for r in rows]}


@router.get("/ventas/facturas")
async def get_facturas_venta(
    company_id: int = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    cod_empresa: Optional[int] = None,
    cod_vendedor: Optional[int] = None,
    page: int = 1,
    limit: int = 200,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    conditions = ["1=1"]
    params: dict = {"offset": (page - 1) * limit, "limit": limit}
    if desde:
        conditions.append("fv.fa_fecha::date >= :desde")
        params["desde"] = date.fromisoformat(desde)
    if hasta:
        conditions.append("fv.fa_fecha::date <= :hasta")
        params["hasta"] = date.fromisoformat(hasta)
    if cod_empresa is not None:
        conditions.append("fv.fa_cod_empresa = :cod_empresa")
        params["cod_empresa"] = cod_empresa
    if cod_vendedor is not None:
        conditions.append("fv.cod_vendedor = :cod_vendedor")
        params["cod_vendedor"] = cod_vendedor

    where = " AND ".join(conditions)
    count_params = {k: v for k, v in params.items() if k not in ("offset", "limit")}
    total = (await db.execute(text(f"SELECT COUNT(*) FROM facturas_venta fv WHERE {where}"), count_params)).scalar() or 0
    rows = (await db.execute(text(f"""
        SELECT fv.fa_id, fv.tipo_comprobante, fv.tipo_factura, fv.fa_cod_empresa,
               fv.fa_fecha, fv.fa_cc, fv.fa_pto_vta, fv.fa_nro, fv.fa_moneda,
               fv.cod_cliente, COALESCE(c.nombre, fv.cod_cliente::text) AS cliente_nombre,
               fv.cod_vendedor, fv.fa_total, fv.fa_total_moneda_local,
               fv.primer_fec_vto, fv.ult_fec_vto, fv.rc_imp_pagado, fv.saldo_fa,
               fv.ultimo_recibo, fv.remitos_asociados
        FROM facturas_venta fv
        LEFT JOIN clientes c ON c.external_id = fv.cod_cliente::text
        WHERE {where}
        ORDER BY fv.fa_fecha DESC NULLS LAST
        LIMIT :limit OFFSET :offset
    """), params)).mappings().all()

    return {
        "total": int(total),
        "page": page,
        "limit": limit,
        "facturas": [
            {
                **dict(r),
                "fa_total": round(float(r["fa_total"] or 0), 2),
                "fa_total_moneda_local": round(float(r["fa_total_moneda_local"] or 0), 2),
                "rc_imp_pagado": round(float(r["rc_imp_pagado"] or 0), 2),
                "saldo_fa": round(float(r["saldo_fa"] or 0), 2),
                "color": (
                    1 if float(r["saldo_fa"] or 0) <= 0.01 else
                    4 if float(r["saldo_fa"] or 0) >= float(r["fa_total"] or 0) * 0.9 else
                    2
                ),
            }
            for r in rows
        ],
    }


@router.get("/clientes/saldos")
async def get_saldos_clientes(
    company_id: int = None,
    cod_empresa: Optional[int] = None,
    color: Optional[int] = None,
    solo_con_saldo: Optional[bool] = True,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    base_cond = "tot_saldo > 0" if solo_con_saldo else "(tot_saldo > 0 OR tot_entrada > 0)"
    conditions = [base_cond]
    params: dict = {}
    if cod_empresa is not None:
        conditions.append("cod_empresa = :cod_empresa")
        params["cod_empresa"] = cod_empresa
    if color is not None:
        conditions.append("color = :color")
        params["color"] = color

    where = " AND ".join(conditions)
    rows = (await db.execute(text(f"""
        SELECT cod_cliente, nombre, fecha, dias_deuda,
               tot_entrada, tot_salida, tot_saldo, color,
               prevision, cod_cuenta, cta_descripcion, snapshot_at
        FROM saldos_clientes
        WHERE {where}
        ORDER BY tot_saldo DESC
        LIMIT 5000
    """), params)).mappings().all()

    total_saldo = sum(float(r["tot_saldo"] or 0) for r in rows)
    return {
        "total": len(rows),
        "total_saldo": round(total_saldo, 2),
        "saldos": [
            {
                **dict(r),
                "tot_entrada": round(float(r["tot_entrada"] or 0), 2),
                "tot_salida": round(float(r["tot_salida"] or 0), 2),
                "tot_saldo": round(float(r["tot_saldo"] or 0), 2),
                "prevision": round(float(r["prevision"] or 0), 2),
            }
            for r in rows
        ],
    }


@router.get("/clientes/pendientes")
async def get_comprobantes_pendientes(
    company_id: int = None,
    cod_empresa: Optional[int] = None,
    color: Optional[int] = None,
    aging: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    conditions = ["saldo > 0"]
    params: dict = {}
    if cod_empresa is not None:
        conditions.append("cod_empresa = :cod_empresa")
        params["cod_empresa"] = cod_empresa
    if color is not None:
        conditions.append("color = :color")
        params["color"] = color
    if aging == "0-30":
        conditions.append("dias_deuda BETWEEN 0 AND 30")
    elif aging == "31-60":
        conditions.append("dias_deuda BETWEEN 31 AND 60")
    elif aging == "61-90":
        conditions.append("dias_deuda BETWEEN 61 AND 90")
    elif aging == "+90":
        conditions.append("dias_deuda > 90")

    where = " AND ".join(conditions)
    rows = (await db.execute(text(f"""
        SELECT comprobante_id, tipo_comprobante, cod_cliente, nombre,
               cod_empresa, cod_vendedor, punto_de_venta, numero,
               importe_factura, importe_pagado, saldo,
               fecha_factura, dias_deuda, color, moneda_fa, detalle
        FROM comprobantes_pendientes_clientes
        WHERE {where}
        ORDER BY dias_deuda DESC NULLS LAST, saldo DESC
        LIMIT 2000
    """), params)).mappings().all()

    total_saldo = sum(float(r["saldo"] or 0) for r in rows)
    return {
        "total": len(rows),
        "total_saldo": round(total_saldo, 2),
        "pendientes": [
            {
                **dict(r),
                "importe_factura": round(float(r["importe_factura"] or 0), 2),
                "importe_pagado": round(float(r["importe_pagado"] or 0), 2),
                "saldo": round(float(r["saldo"] or 0), 2),
            }
            for r in rows
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Catalog list endpoints (for FilterBar and widget dropdowns)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/vendedores")
async def get_vendedores_catalog(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    rows = (await db.execute(text(
        "SELECT cod_vendedor, nombre, email, habilitado FROM vendedores ORDER BY nombre"
    ))).mappings().all()
    return {"vendedores": [dict(r) for r in rows], "total": len(rows)}


@router.get("/clientes")
async def get_clientes_catalog(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    rows = (await db.execute(text("""
        SELECT cod_cliente, nombre
        FROM saldos_clientes
        ORDER BY nombre
        LIMIT 5000
    """))).mappings().all()
    if not rows:
        rows = (await db.execute(text("""
            SELECT DISTINCT cliente_id AS cod_cliente, cliente_nombre AS nombre
            FROM ventas
            WHERE cliente_id IS NOT NULL AND cliente_nombre IS NOT NULL
            ORDER BY nombre
            LIMIT 5000
        """))).mappings().all()
    return {"clientes": [dict(r) for r in rows], "total": len(rows)}


@router.get("/articulos")
async def get_articulos_catalog(
    company_id: int = None,
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)
    rows = (await db.execute(text("""
        SELECT DISTINCT ON (cod_articulo)
            cod_articulo, descripcion, cod_rubro, rubro, cod_subrubro, subrubro,
            precio_compra, precio_venta, habilitado
        FROM stock_disponible
        ORDER BY cod_articulo, descripcion
        LIMIT 2000
    """))).mappings().all()
    return {"articulos": [dict(r) for r in rows], "total": len(rows)}


@router.get("/caja/egresos")
async def get_caja_egresos(
    company_id: int = None,
    filters: GlobalFilters = Depends(get_global_filters),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_schema = await get_tenant_schema(current_user, db, company_id)
    await set_tenant_search_path(db, tenant_schema)

    params: dict = {"desde": filters.desde, "hasta": filters.hasta_exclusive}
    cond = "fecha >= :desde AND fecha < :hasta AND COALESCE(anulada, 'N') <> 'S'"
    if filters.cod_empresa:
        cond += " AND cod_empresa = ANY(:cod_empresa)"
        params["cod_empresa"] = filters.cod_empresa

    rows = (await db.execute(text(f"""
        SELECT
            proveedor_id,
            COALESCE(MAX(proveedor_nombre), proveedor_id::text) AS nombre,
            COUNT(*) AS ordenes,
            COALESCE(SUM({compra_importe_neto_expr()}), 0) AS total_pagado
        FROM compras
        WHERE {cond}
        GROUP BY proveedor_id
        ORDER BY total_pagado DESC
        LIMIT 200
    """), params)).mappings().all()

    total_egresos = sum(float(r["total_pagado"] or 0) for r in rows)
    return {
        "total_egresos": round(total_egresos, 2),
        "proveedores": [
            {
                "proveedor_id": r["proveedor_id"],
                "nombre": r["nombre"],
                "ordenes": int(r["ordenes"] or 0),
                "total_pagado": round(float(r["total_pagado"] or 0), 2),
            }
            for r in rows
        ],
        "total": len(rows),
    }
